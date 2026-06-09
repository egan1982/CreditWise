# pyright: strict
# pyright: reportAny=false, reportExplicitAny=false, reportUnknownMemberType=false, reportUnknownArgumentType=false, reportUnknownVariableType=false, reportUnknownParameterType=false
# pyright: reportImplicitRelativeImport=false, reportConstantRedefinition=false, reportMissingTypeArgument=false, reportOptionalMemberAccess=false, reportArgumentType=false
# pyright: reportAttributeAccessIssue=false, reportCallIssue=false, reportIndexIssue=false, reportAssignmentType=false, reportPossiblyUnboundVariable=false
# pyright: reportDeprecated=false, reportUnusedImport=false, reportUnusedCallResult=false, reportMissingParameterType=false, reportUndefinedVariable=false
# pyright: reportUnusedVariable=false, reportPrivateUsage=false, reportUnusedParameter=false, reportCallInDefaultInitializer=false
"""
SOP Task API - 业务场景SOP任务管理接口

Provides:
- Task type listing and metadata
- Data preview and analysis
- Task execution and progress tracking
- Result retrieval
- Task control (pause/stop/resume)
- Task history and statistics
"""

print("[SOP API] Module loading started...")

import logging
import os
import time
from datetime import datetime
from functools import lru_cache
from pathlib import Path
from typing import Any, Dict, List, Optional

print("[SOP API] Basic imports done")

import numpy as np
import pandas as pd
from fastapi import APIRouter, BackgroundTasks, HTTPException, Query, UploadFile, File, Form
from fastapi.responses import JSONResponse
from pydantic import BaseModel, Field

print("[SOP API] FastAPI imports done")

# Setup logger early to avoid NameError in import blocks
logger = logging.getLogger(__name__)

# Import session_id validation utility
try:
    from utils import validate_session_id as _validate_session_id
except ImportError:
    from API.utils import validate_session_id as _validate_session_id

# Import SOP modules
import sys
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

print("[SOP API] Importing registry...")
from deepanalyze.analysis.task_SOP.registry import (
    get_registry,
    register_builtin_tasks
)
print("[SOP API] Registry imported")

print("[SOP API] Importing executor...")
from deepanalyze.analysis.task_SOP.executor import (
    get_executor,
    get_execution_status,
    get_execution_result,
    ExecutionStore,
    ExecutionStatus,
    StageSnapshot,
)
print("[SOP API] Executor imported")

# Import new LLM+Pipeline architecture components
try:
    from deepanalyze.analysis.task_SOP.llm_param_extractor import (
        LLMParamExtractor,
        create_param_extractor,
        TaskIntent,
    )
    from deepanalyze.analysis.task_SOP.task_router import (
        UnifiedTaskRouter,
        create_router,
        route_from_chat,
        EntrySource,
    )
    from deepanalyze.analysis.task_SOP.code_templates import (
        StageCodeGenerator,
        EquivalentCodeGenerator,
        generate_task_config_summary,
    )
    LLM_PIPELINE_ARCH_AVAILABLE = True
except ImportError as e:
    logger.warning(f"LLM+Pipeline architecture components not available: {e}")
    LLM_PIPELINE_ARCH_AVAILABLE = False

# Import task manager components
try:
    from deepanalyze.core.task_manager import (
        TaskController,
        TaskHistoryService,
        TaskStatus,
        get_result_storage,
        PersistentExecutionStore,
        ExecutionStateRecovery,
        StageAnalysisService,
    )
    TASK_MANAGER_AVAILABLE = True
except ImportError:
    TASK_MANAGER_AVAILABLE = False
    PersistentExecutionStore = None
    ExecutionStateRecovery = None
    StageAnalysisService = None

# Initialize router
router = APIRouter(prefix="/sop", tags=["SOP Tasks"])

# =============================================================================
# Running Tasks Registry - 跟踪正在运行的任务，防止同一 execution_id 的多个实例同时运行
# Phase 19: 解决竞态条件问题
# =============================================================================
import asyncio
from typing import Tuple

# 存储正在运行的任务: {execution_id: (asyncio.Task, timestamp)}
_running_tasks: Dict[str, Tuple[asyncio.Task, float]] = {}
_running_tasks_lock = asyncio.Lock()


async def _cancel_existing_task(execution_id: str) -> bool:
    """取消已存在的任务（如果有）
    
    Returns:
        True 如果取消了一个任务，False 如果没有任务需要取消
    """
    async with _running_tasks_lock:
        if execution_id in _running_tasks:
            task, started_at = _running_tasks[execution_id]
            if not task.done():
                logger.warning(f"[Task Registry] Cancelling existing task for {execution_id} (started at {started_at})")
                task.cancel()
                try:
                    await asyncio.wait_for(asyncio.shield(task), timeout=1.0)
                except (asyncio.CancelledError, asyncio.TimeoutError):
                    pass
            del _running_tasks[execution_id]
            return True
    return False


async def _register_task(execution_id: str, task: asyncio.Task) -> None:
    """注册一个新任务"""
    import time
    async with _running_tasks_lock:
        _running_tasks[execution_id] = (task, time.time())
        logger.info(f"[Task Registry] Registered task for {execution_id}")


async def _unregister_task(execution_id: str) -> None:
    """注销一个任务"""
    async with _running_tasks_lock:
        if execution_id in _running_tasks:
            del _running_tasks[execution_id]
            logger.info(f"[Task Registry] Unregistered task for {execution_id}")


def _is_task_running(execution_id: str) -> bool:
    """检查任务是否正在运行"""
    if execution_id in _running_tasks:
        task, _ = _running_tasks[execution_id]
        return not task.done()
    return False


# =============================================================================
# File Preview Cache - 缓存大文件的列信息，避免重复解析
# =============================================================================
# 缓存结构: {cache_key: {"data": DataPreviewResponse, "mtime": file_mtime, "cached_at": timestamp}}
_preview_cache: Dict[str, Dict[str, Any]] = {}
_CACHE_MAX_SIZE = 50  # 最多缓存50个文件的预览信息


def _get_cache_key(file_path: str, session_id: str) -> str:
    """生成缓存key"""
    return f"{session_id}:{file_path}"


def _get_cached_preview(file_path, session_id: str, file_mtime: float) -> Optional[Dict]:
    """获取缓存的预览数据，如果文件已修改则返回None"""
    cache_key = _get_cache_key(str(file_path), session_id)
    if cache_key in _preview_cache:
        cached = _preview_cache[cache_key]
        # 检查文件是否被修改
        if cached.get("mtime") == file_mtime:
            logger.info(f"Cache hit for file: {file_path}")
            return cached["data"]
        else:
            # 文件已修改，删除旧缓存
            del _preview_cache[cache_key]
            logger.info(f"Cache invalidated (file modified): {file_path}")
    return None


def _set_cached_preview(file_path, session_id: str, file_mtime: float, data: Dict):
    """缓存预览数据"""
    global _preview_cache
    cache_key = _get_cache_key(str(file_path), session_id)
    
    # 如果缓存已满，删除最旧的条目
    if len(_preview_cache) >= _CACHE_MAX_SIZE:
        oldest_key = min(_preview_cache.keys(), key=lambda k: _preview_cache[k].get("cached_at", 0))
        del _preview_cache[oldest_key]
        logger.info(f"Cache evicted oldest entry: {oldest_key}")
    
    _preview_cache[cache_key] = {
        "data": data,
        "mtime": file_mtime,
        "cached_at": datetime.now().timestamp()
    }
    logger.info(f"Cached preview for file: {file_path}")


# =============================================================================
# Task Result Cache - 缓存任务结果序列化数据，避免重复转换
# =============================================================================
# 使用OrderedDict实现LRU缓存，缓存键包含文件修改时间
from collections import OrderedDict

_TASK_RESULT_CACHE: OrderedDict[str, Dict[str, Any]] = OrderedDict()
_TASK_RESULT_CACHE_MAX_SIZE = 64  # 最多缓存64个任务结果
_TASK_RESULT_CACHE_HITS = 0
_TASK_RESULT_CACHE_MISSES = 0


def _get_task_result_cache_key(record_id: str, file_mtime: float) -> str:
    """生成缓存键"""
    return f"{record_id}:{file_mtime}"


def _get_task_result_from_cache(record_id: str, file_mtime: float) -> Optional[Dict[str, Any]]:
    """
    从缓存获取序列化结果
    
    Args:
        record_id: 记录ID
        file_mtime: 文件修改时间（用于缓存失效）
        
    Returns:
        缓存的序列化结果，不存在返回None
    """
    global _TASK_RESULT_CACHE_HITS, _TASK_RESULT_CACHE_MISSES
    
    cache_key = _get_task_result_cache_key(record_id, file_mtime)
    
    if cache_key in _TASK_RESULT_CACHE:
        # 缓存命中，移到末尾（最近使用）
        result = _TASK_RESULT_CACHE.pop(cache_key)
        _TASK_RESULT_CACHE[cache_key] = result
        _TASK_RESULT_CACHE_HITS += 1
        return result
    
    _TASK_RESULT_CACHE_MISSES += 1
    return None


def _set_task_result_cache(record_id: str, file_mtime: float, serialized_result: Dict[str, Any]):
    """
    设置缓存
    
    Args:
        record_id: 记录ID
        file_mtime: 文件修改时间
        serialized_result: 序列化后的结果
    """
    global _TASK_RESULT_CACHE
    
    cache_key = _get_task_result_cache_key(record_id, file_mtime)
    
    # 如果缓存已满，删除最旧的条目
    if len(_TASK_RESULT_CACHE) >= _TASK_RESULT_CACHE_MAX_SIZE:
        oldest_key = next(iter(_TASK_RESULT_CACHE))
        _TASK_RESULT_CACHE.pop(oldest_key)
        logger.debug(f"[TaskResultCache] Evicted oldest entry")
    
    _TASK_RESULT_CACHE[cache_key] = serialized_result


def _get_task_result_cache_stats() -> Dict[str, Any]:
    """获取任务结果缓存统计信息"""
    global _TASK_RESULT_CACHE_HITS, _TASK_RESULT_CACHE_MISSES
    total = _TASK_RESULT_CACHE_HITS + _TASK_RESULT_CACHE_MISSES
    return {
        "hits": _TASK_RESULT_CACHE_HITS,
        "misses": _TASK_RESULT_CACHE_MISSES,
        "maxsize": _TASK_RESULT_CACHE_MAX_SIZE,
        "currsize": len(_TASK_RESULT_CACHE),
        "hit_rate": f"{(_TASK_RESULT_CACHE_HITS / total * 100):.1f}%" if total > 0 else "N/A"
    }


def _clear_task_result_cache():
    """清除任务结果缓存"""
    global _TASK_RESULT_CACHE, _TASK_RESULT_CACHE_HITS, _TASK_RESULT_CACHE_MISSES
    _TASK_RESULT_CACHE.clear()
    _TASK_RESULT_CACHE_HITS = 0
    _TASK_RESULT_CACHE_MISSES = 0
    logger.info("[TaskResultCache] Cache cleared")


# =============================================================================
# Pydantic Models
# =============================================================================

class DataPreviewRequest(BaseModel):
    """数据预览请求"""
    file_path: str = Field(..., description="数据文件路径（相对于工作区）")
    rows: int = Field(default=10, ge=1, le=100, description="预览行数")
    session_id: str = Field(default="default", description="会话ID")


class DataPreviewResponse(BaseModel):
    """数据预览响应"""
    columns: List[Dict[str, Any]]  # [{name, dtype, sample_values, null_count}]
    preview_data: List[Dict]       # 前N行数据
    total_rows: int
    total_columns: int


class TaskExecuteRequest(BaseModel):
    """任务执行请求"""
    task_id: str = Field(..., description="任务类型ID")
    session_id: str = Field(..., description="会话ID")
    file_path: str = Field(..., description="数据文件路径")
    params: Dict[str, Any] = Field(default_factory=dict, description="任务参数")
    interaction_mode: str = Field(default="auto", description="交互模式: auto, expert")
    # LLM配置（来自前端ModelSelector）
    model: str = Field(default="deepseek-chat", description="LLM模型名称")
    api_base: str = Field(default="http://localhost:8200/llm-manager/api/proxy", description="LLM API基础URL")
    system_prompt: Optional[str] = Field(default=None, description="用户自定义系统提示词（与SOP Prompt合并）")


class TaskExecuteResponse(BaseModel):
    """任务执行响应"""
    execution_id: str
    task_id: str
    status: str
    message: str


class TaskStatusResponse(BaseModel):
    """任务状态响应"""
    execution_id: str
    task_id: str
    status: str
    current_stage: str
    overall_progress: float
    message: str
    started_at: Optional[str]
    completed_at: Optional[str]
    stages: Dict[str, Any]
    record_id: Optional[str] = None  # Phase 9: 添加record_id用于AI分析缓存
    file_path: Optional[str] = None  # 任务使用的数据文件路径


class TaskResultResponse(BaseModel):
    """任务结果响应"""
    execution_id: str
    task_id: str
    status: str
    outputs: Dict[str, Any]
    errors: List[str]


class TaskListItem(BaseModel):
    """任务列表项"""
    task_id: str
    task_name: str
    task_name_en: str
    description: str
    category: str
    icon: str
    estimated_time: str


class BuildPromptRequest(BaseModel):
    """构建Prompt请求"""
    task_id: str = Field(..., description="任务类型ID")
    params: Dict[str, Any] = Field(default_factory=dict, description="任务参数")
    workspace_files_info: str = Field(default="", description="工作区文件信息")


# =============================================================================
# LLM+Pipeline 新架构 Pydantic Models
# =============================================================================

class LLMExtractRequest(BaseModel):
    """LLM 参数推断请求（Chat 入口）"""
    user_message: str = Field(..., description="用户自然语言消息")
    session_id: str = Field(..., description="会话ID")
    workspace_files: List[str] = Field(default_factory=list, description="工作区文件列表")
    conversation_history: List[Dict[str, str]] = Field(default_factory=list, description="对话历史")
    model: str = Field(default="deepseek-chat", description="LLM模型名称")
    api_base: str = Field(default="http://localhost:8200/llm-manager/api/proxy", description="LLM API基础URL")


class LLMExtractResponse(BaseModel):
    """LLM 参数推断响应"""
    success: bool
    task_type: Optional[str] = None
    confidence: float = 0.0
    params: Dict[str, Any] = Field(default_factory=dict)
    missing_params: List[str] = Field(default_factory=list)
    clarification_needed: bool = False
    clarification_question: Optional[str] = None
    extraction_code: Optional[str] = None  # 推断过程的伪代码展示


class UnifiedExecuteRequest(BaseModel):
    """统一执行请求（支持 SOP UI 和 Chat 两种入口）"""
    task_id: str = Field(..., description="任务类型ID")
    session_id: str = Field(..., description="会话ID")
    file_path: str = Field(..., description="数据文件路径")
    params: Dict[str, Any] = Field(default_factory=dict, description="任务参数")
    source: str = Field(default="sop_ui", description="入口来源: sop_ui, chat")
    interaction_mode: str = Field(default="auto", description="交互模式: auto, expert")
    # LLM 配置
    model: str = Field(default="deepseek-chat", description="LLM模型名称")
    api_base: str = Field(default="http://localhost:8200/llm-manager/api/proxy", description="LLM API基础URL")
    system_prompt: Optional[str] = Field(default=None, description="用户自定义系统提示词")
    # Chat 入口特有字段
    chat_context: Optional[Dict[str, Any]] = Field(default=None, description="Chat上下文（用户消息、对话历史等）")


class StageCodeEvent(BaseModel):
    """阶段代码事件（用于 Code 栏实时展示）"""
    event_type: str  # config, llm_extraction, stage_start, stage_complete, stage_error, full_code
    stage_id: Optional[str] = None
    stage_name: Optional[str] = None
    content: str  # 代码内容
    status: Optional[str] = None  # pending, running, completed, failed
    result: Optional[str] = None  # 结果摘要
    execution_time_ms: Optional[int] = None
    timestamp: float


# =============================================================================
# Initialization
# =============================================================================

# Register builtin tasks on module load
print("[SOP] Starting builtin task registration...")
try:
    print("[SOP] Calling register_builtin_tasks()...")
    register_builtin_tasks()
    print("[SOP] register_builtin_tasks() completed")
    # Verify registration
    registry = get_registry()
    task_count = len(registry.list_tasks())
    logger.info(f"SOP builtin tasks registered: {task_count} tasks")
    print(f"[SOP] Registered {task_count} tasks: {[t['task_id'] for t in registry.list_tasks()]}")
    print(f"[SOP] Registry instance id: {id(registry)}")
except Exception as e:
    import traceback
    logger.error(f"Failed to register builtin tasks: {e}")
    logger.error(traceback.format_exc())
    print(f"[SOP ERROR] Failed to register builtin tasks: {e}")
    traceback.print_exc()


# =============================================================================
# API Endpoints
# =============================================================================

@router.get("/tasks", response_model=List[TaskListItem])
async def list_available_tasks(
    category: Optional[str] = Query(None, description="按分类筛选")
) -> List[Dict]:
    """
    获取所有可用的SOP任务类型
    
    Args:
        category: 可选，按分类筛选（风控建模/营销分析/运营分析）
        
    Returns:
        任务类型列表
    """
    registry = get_registry()
    print(f"[SOP API] list_tasks called, registry id: {id(registry)}, tasks: {len(registry.list_tasks())}")
    tasks = registry.list_tasks(category=category)
    return tasks


@router.get("/tasks/{task_id}")
async def get_task_definition(task_id: str) -> Dict:
    """
    获取指定任务的详细定义
    
    Args:
        task_id: 任务类型ID
        
    Returns:
        任务定义，包含参数Schema、阶段信息等
    """
    registry = get_registry()
    task_meta = registry.get_task_meta_dict(task_id)
    
    if not task_meta:
        raise HTTPException(status_code=404, detail="Task not found")
    
    return task_meta


@router.post("/data/preview", response_model=DataPreviewResponse)
async def preview_data(request: DataPreviewRequest) -> DataPreviewResponse:
    """
    预览上传的数据文件（带缓存优化）
    
    Args:
        request: 包含file_path、预览行数和session_id
        
    Returns:
        数据预览信息，包含列信息和样本数据
    """
    from pathlib import Path
    
    # P1-D5: 入口层 session_id 校验
    try:
        request.session_id = _validate_session_id(request.session_id)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    
    # 构建完整的文件路径
    workspace_dir = Path("workspace") / request.session_id
    file_path = workspace_dir / request.file_path
    
    logger.info(f"Preview data request: file_path={request.file_path}, session_id={request.session_id}, full_path={file_path}")
    
    # Check file exists
    if not file_path.exists():
        logger.error(f"File not found: {file_path}")
        raise HTTPException(status_code=404, detail="File not found")
    
    # 获取文件修改时间用于缓存验证
    file_mtime = file_path.stat().st_mtime
    
    # 检查缓存（对于大文件特别有用）
    cached_data = _get_cached_preview(file_path, request.session_id, file_mtime)
    if cached_data:
        # 缓存命中，但可能需要调整preview_data的行数
        result = DataPreviewResponse(**cached_data)
        # 如果请求的行数不同，重新切片preview_data
        if len(result.preview_data) > request.rows:
            result.preview_data = result.preview_data[:request.rows]
        return result
    
    try:
        file_str = str(file_path)
        file_size_mb = file_path.stat().st_size / (1024 * 1024)
        
        # 只读取前1000行用于统计（优化大文件性能）
        sample_rows = 1000
        
        if file_str.endswith('.csv'):
            # 先读取少量行获取列信息和统计
            df_sample = pd.read_csv(file_path, nrows=sample_rows)
            # 获取总行数（不加载全部数据）
            with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                total_rows = sum(1 for _ in f) - 1  # 减去header行
        elif file_str.endswith(('.xlsx', '.xls')):
            # 对于大文件（>50MB），使用openpyxl流式读取避免内存问题
            if file_size_mb > 50:
                logger.info(f"Large Excel file detected ({file_size_mb:.1f}MB), using streaming read")
                try:
                    import openpyxl
                    from openpyxl.utils import get_column_letter
                    
                    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
                    ws = wb.active
                    
                    # 流式读取前sample_rows行
                    rows_data = []
                    headers = None
                    row_count = 0
                    
                    for row in ws.iter_rows(values_only=True):
                        if headers is None:
                            headers = [str(cell) if cell is not None else f"Column_{i}" for i, cell in enumerate(row)]
                        else:
                            if row_count < sample_rows:
                                rows_data.append(dict(zip(headers, row)))
                            row_count += 1
                    
                    total_rows = row_count
                    df_sample = pd.DataFrame(rows_data) if rows_data else pd.DataFrame(columns=headers)
                    wb.close()
                    
                except Exception as e:
                    logger.warning(f"Streaming read failed: {e}, falling back to pandas")
                    # 降级到pandas，但限制读取行数
                    df_sample = pd.read_excel(file_path, nrows=sample_rows, engine='openpyxl')
                    total_rows = len(df_sample)  # 无法获取真实总行数
            else:
                # 小文件直接用pandas读取
                df_sample = pd.read_excel(file_path, nrows=sample_rows)
                try:
                    import openpyxl
                    wb = openpyxl.load_workbook(file_path, read_only=True)
                    ws = wb.active
                    total_rows = ws.max_row - 1 if ws.max_row else 0
                    wb.close()
                except Exception:
                    total_rows = len(df_sample)  # 降级方案
        else:
            raise HTTPException(status_code=400, detail="Unsupported file type")
        
        # Build column info（基于采样数据）
        columns = []
        for col in df_sample.columns:
            col_info = {
                "name": col,
                "dtype": str(df_sample[col].dtype),
                "null_count": int(df_sample[col].isnull().sum()),
                "null_ratio": round(df_sample[col].isnull().sum() / len(df_sample), 4) if len(df_sample) > 0 else 0,
                "unique_count": int(df_sample[col].nunique()),
                "sample_values": df_sample[col].dropna().head(5).tolist()
            }
            columns.append(col_info)
        
        # Preview data - 缓存时保存更多行（最多100行），返回时按请求行数切片
        max_cache_rows = 100
        preview_rows = min(max_cache_rows, len(df_sample))
        preview_data_full = df_sample.head(preview_rows).fillna("").to_dict(orient="records")
        
        logger.info(f"Preview data success: {len(columns)} columns, {total_rows} total rows")
        
        # 构建响应数据
        response_data = {
            "columns": columns,
            "preview_data": preview_data_full,
            "total_rows": total_rows,
            "total_columns": len(df_sample.columns)
        }
        
        # 缓存结果（对于大文件特别有用）
        file_size_mb = file_path.stat().st_size / (1024 * 1024)
        if file_size_mb > 10:  # 只缓存>10MB的文件
            _set_cached_preview(file_path, request.session_id, file_mtime, response_data)
        
        # 返回时按请求行数切片
        response_data["preview_data"] = preview_data_full[:request.rows]
        
        return DataPreviewResponse(**response_data)
        
    except Exception as e:
        logger.error(f"Failed to preview data: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail="Failed to preview data")


# =============================================================================
# 敏感信息预检（个保法合规）
# =============================================================================

class SensitiveCheckRequest(BaseModel):
    file_path: str = Field(..., description="数据文件路径（相对于工作区）")
    session_id: str = Field(..., description="Session ID")
    sample_rows: int = Field(default=100, ge=10, le=1000)


@router.post("/data/sensitive-check")
async def sensitive_check(request: SensitiveCheckRequest):
    """
    上传数据集的敏感信息预检接口（个保法合规）

    双层检测：
    - L1: 列名关键词匹配
    - L2: 样本值正则扫描（覆盖匿名列名）

    Returns:
        检测报告，含 has_sensitive / max_level / findings / summary
    """
    from pathlib import Path
    from deepanalyze.analysis.data_validator import SensitiveFieldDetector

    try:
        request.session_id = _validate_session_id(request.session_id)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))

    workspace_dir = Path("workspace") / request.session_id
    file_path = workspace_dir / request.file_path

    if not file_path.exists():
        raise HTTPException(status_code=404, detail="File not found")

    try:
        df = pd.read_csv(file_path, nrows=request.sample_rows)
    except Exception as e:
        logger.error(f"Failed to read file for sensitive check: {e}")
        raise HTTPException(status_code=400, detail=f"Failed to read CSV: {e}")

    detector = SensitiveFieldDetector(sample_rows=request.sample_rows)
    report = detector.detect(df)

    return {
        "has_sensitive": report.has_sensitive,
        "max_level": report.max_level.value if report.max_level else None,
        "findings": [
            {
                "column": f.column,
                "level": f.level.value,
                "rule_name": f.rule_name,
                "detection_method": f.detection_method,
                "hit_rate": round(f.hit_rate, 4) if f.hit_rate is not None else None,
                "sample_values": f.sample_values,
            }
            for f in report.findings
        ],
        "summary": report.summary,
    }


@router.post("/data/analyze")
async def analyze_data(
    file_path: str = Query(..., description="数据文件路径（相对于工作区）"),
    session_id: str = Query("default")
) -> Dict:
    """
    分析数据特征，推荐任务类型和参数
    
    Args:
        file_path: 数据文件路径（相对于工作区）
        session_id: 会话ID
        
    Returns:
        数据分析结果和任务推荐
    """
    # P1-D5: 入口层 session_id 校验
    try:
        session_id = _validate_session_id(session_id)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    
    # P1-D4: 路径限制 — 仅允许访问 workspace 内的文件
    from pathlib import Path
    workspace_dir = Path("workspace") / session_id
    abs_workspace = workspace_dir.resolve()
    full_path = (workspace_dir / file_path).resolve()
    
    if abs_workspace not in full_path.parents and full_path != abs_workspace:
        raise HTTPException(status_code=400, detail="Invalid file path")
    
    if not full_path.exists():
        raise HTTPException(status_code=404, detail="File not found")
    
    try:
        # Load data
        file_str = str(full_path)
        if file_str.endswith('.csv'):
            df = pd.read_csv(full_path)
        elif file_str.endswith(('.xlsx', '.xls')):
            df = pd.read_excel(full_path)
        else:
            raise HTTPException(status_code=400, detail="Unsupported file type")
        
        # Basic analysis
        analysis = {
            "total_rows": len(df),
            "total_columns": len(df.columns),
            "numeric_columns": df.select_dtypes(include=['number']).columns.tolist(),
            "categorical_columns": df.select_dtypes(include=['object', 'category']).columns.tolist(),
            "binary_columns": [],
            "potential_target_columns": [],
            "potential_weight_columns": [],
            "potential_id_columns": []
        }
        
        # Identify binary columns (potential targets)
        for col in df.columns:
            unique_vals = df[col].dropna().unique()
            if len(unique_vals) == 2:
                analysis["binary_columns"].append(col)
                # Check if it looks like a target
                if set(unique_vals).issubset({0, 1}) or set(unique_vals).issubset({'0', '1'}):
                    analysis["potential_target_columns"].append(col)
        
        # Identify potential weight columns
        for col in df.select_dtypes(include=['number']).columns:
            if 'weight' in col.lower() or 'wt' in col.lower():
                analysis["potential_weight_columns"].append(col)
        
        # Identify potential ID columns
        for col in df.columns:
            if 'id' in col.lower() or 'uuid' in col.lower():
                analysis["potential_id_columns"].append(col)
        
        # Task recommendations
        recommendations = []
        if analysis["potential_target_columns"]:
            recommendations.append({
                "task_id": "rule_mining",
                "task_name": "规则挖掘",
                "reason": f"检测到二值目标列: {analysis['potential_target_columns']}",
                "suggested_params": {
                    "target_col": analysis["potential_target_columns"][0],
                    "weight_col": analysis["potential_weight_columns"][0] if analysis["potential_weight_columns"] else None
                }
            })
        
        analysis["recommendations"] = recommendations
        return analysis
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Failed to analyze data: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail="Failed to analyze data")


@router.post("/execute", response_model=TaskExecuteResponse)
async def execute_task(
    request: TaskExecuteRequest,
    background_tasks: BackgroundTasks
) -> TaskExecuteResponse:
    """
    执行SOP任务（异步）
    
    Args:
        request: 任务执行请求
        background_tasks: FastAPI后台任务
        
    Returns:
        执行ID和初始状态
    """
    # P1-D5: 入口层 session_id 校验
    try:
        request.session_id = _validate_session_id(request.session_id)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    
    # Validate task exists
    registry = get_registry()
    task_def = registry.get_task(request.task_id)
    if not task_def:
        raise HTTPException(status_code=404, detail="Task not found")
    
    # Build full file path (relative to workspace)
    from pathlib import Path
    workspace_dir = Path("workspace") / request.session_id
    full_file_path = workspace_dir / request.file_path
    
    logger.info(f"Execute task request: file_path={request.file_path}, session_id={request.session_id}, full_path={full_file_path}")
    logger.info(f"LLM config: model={request.model}, api_base={request.api_base}")
    
    # Validate file exists
    if not full_file_path.exists():
        logger.error(f"File not found: {full_file_path}")
        raise HTTPException(status_code=404, detail="File not found")
    
    # Create execution context with full path and LLM config
    context = ExecutionStore.create(
        task_id=request.task_id,
        session_id=request.session_id,
        params=request.params,
        file_path=str(full_file_path),
        interaction_mode=request.interaction_mode,
        model=request.model,
        api_base=request.api_base,
        system_prompt=request.system_prompt
    )
    execution_id = context.execution_id
    
    # Schedule background execution
    executor = get_executor()
    
    async def run_task():
        try:
            await executor.execute_async(
                task_id=request.task_id,
                session_id=request.session_id,
                params=request.params,
                file_path=str(full_file_path),
                execution_id=execution_id,  # Pass the same execution_id
                interaction_mode=request.interaction_mode,
                model=request.model,
                api_base=request.api_base,
                system_prompt=request.system_prompt
            )
        except asyncio.CancelledError:
            logger.info(f"Task {execution_id} was cancelled")
        except Exception as e:
            import traceback
            logger.error(f"Background task execution failed: {e}\n{traceback.format_exc()}")
            # 更新任务状态为失败
            ctx = ExecutionStore.get(execution_id)
            if ctx:
                ctx.status = ExecutionStatus.FAILED
                ctx.message = f"任务执行失败: {str(e)}"
                ctx.errors.append(traceback.format_exc())
                ExecutionStore.update(ctx)
        finally:
            # Phase 19: 任务完成后注销
            await _unregister_task(execution_id)
    
    # Phase 19: 取消已存在的任务（如果有），然后注册新任务
    await _cancel_existing_task(execution_id)
    task = asyncio.create_task(run_task())
    await _register_task(execution_id, task)
    
    return TaskExecuteResponse(
        execution_id=context.execution_id,
        task_id=request.task_id,
        status="pending",
        message="任务已提交，正在执行中..."
    )


@router.get("/status/{execution_id}", response_model=TaskStatusResponse)
async def get_task_status(execution_id: str) -> TaskStatusResponse:
    """
    获取任务执行状态
    
    Args:
        execution_id: 执行ID
        
    Returns:
        任务状态信息
    """
    status = get_execution_status(execution_id)
    
    if not status:
        raise HTTPException(status_code=404, detail="Execution not found")
    
    return TaskStatusResponse(**status)


def _ensure_json_serializable(obj: Any) -> Any:
    """确保对象可JSON序列化"""
    if obj is None:
        return None
    elif isinstance(obj, bool):
        return obj
    elif isinstance(obj, (int, str)):
        return obj
    elif isinstance(obj, float):
        # 处理NaN和Infinity，JSON不支持这些值
        if np.isnan(obj) or np.isinf(obj):
            return None
        return obj
    elif isinstance(obj, (np.integer,)):
        return int(obj)
    elif isinstance(obj, (np.floating,)):
        val = float(obj)
        if np.isnan(val) or np.isinf(val):
            return None
        return val
    elif isinstance(obj, np.ndarray):
        return _ensure_json_serializable(obj.tolist())
    elif isinstance(obj, dict):
        return {str(k): _ensure_json_serializable(v) for k, v in obj.items()}
    elif isinstance(obj, (list, tuple)):
        return [_ensure_json_serializable(v) for v in obj]
    elif isinstance(obj, pd.DataFrame):
        # 递归处理DataFrame中的数据
        # 先将Categorical列转换为object类型，避免fillna时出现"Cannot setitem on a Categorical with a new category"错误
        df_copy = obj.copy()
        for col in df_copy.columns:
            if pd.api.types.is_categorical_dtype(df_copy[col]):
                df_copy[col] = df_copy[col].astype(object)
        data = df_copy.fillna("").head(100).to_dict(orient="records")
        return {
            "type": "dataframe",
            "columns": [str(c) for c in obj.columns],
            "shape": list(obj.shape),
            "data": _ensure_json_serializable(data)
        }
    elif isinstance(obj, pd.Series):
        # 将Categorical类型转换为object类型，避免fillna时出错
        if pd.api.types.is_categorical_dtype(obj):
            obj = obj.astype(object)
        return _ensure_json_serializable(obj.fillna("").tolist())
    else:
        return str(obj)


def _wrap_stored_outputs(outputs: Dict[str, Any]) -> Dict[str, Any]:
    """
    将存储的结果包装成与内存结果相同的格式
    内存结果格式: {key: {type: "xxx", data: ...}}
    存储结果格式: {key: raw_data}
    """
    wrapped = {}
    for key, value in outputs.items():
        # 检查是否已经有type/data包装
        if isinstance(value, dict) and "type" in value and "data" in value:
            # 已经是包装格式，只需确保数据可序列化
            wrapped[key] = {
                "type": value["type"],
                "data": _ensure_json_serializable(value["data"])
            }
        elif isinstance(value, dict):
            # 普通字典，包装为dict类型
            wrapped[key] = {
                "type": "dict",
                "data": _ensure_json_serializable(value)
            }
        elif isinstance(value, (list, tuple)):
            # 列表，包装为list类型
            wrapped[key] = {
                "type": "list",
                "data": _ensure_json_serializable(value)
            }
        elif isinstance(value, pd.DataFrame):
            # DataFrame
            wrapped[key] = _ensure_json_serializable(value)
        elif value is None:
            wrapped[key] = {"type": "none", "data": None}
        else:
            # 其他类型
            wrapped[key] = {
                "type": "other",
                "data": _ensure_json_serializable(value)
            }
    return wrapped


@router.get("/results/{execution_id}")
async def get_task_results(execution_id: str) -> JSONResponse:
    """
    获取任务执行结果
    
    Args:
        execution_id: 执行ID
        
    Returns:
        任务结果信息
    """
    result = get_execution_result(execution_id)
    
    # 如果内存中找不到，尝试从历史存储加载
    if not result and TASK_MANAGER_AVAILABLE:
        # 通过execution_id查找历史记录
        record = TaskHistoryService.get_record_by_execution_id(execution_id)
        if record:
            record_id = record.get("record_id")
            if record_id:
                storage = get_result_storage()
                try:
                    stored_result = storage.load_result(record_id)
                    if stored_result:
                        # 将存储结果包装成与内存结果相同的格式
                        wrapped_outputs = _wrap_stored_outputs(stored_result)
                        # 构造与内存结果相同的格式，包含TaskResultResponse所需的所有字段
                        # 注意：历史记录中的字段是 task_type，不是 task_id
                        result = {
                            "execution_id": execution_id,
                            "task_id": record.get("task_type") or record.get("task_id", "unknown"),
                            "status": "completed",
                            "outputs": wrapped_outputs,
                            "errors": [],
                            "record_id": record_id  # 添加record_id，用于前端获取历史stages数据
                        }
                except Exception as e:
                    logger.error(f"Failed to load result for {record_id}: {e}")
                    raise HTTPException(
                        status_code=500,
                        detail="Failed to load stored result"
                    )
    
    if not result:
        raise HTTPException(status_code=404, detail="Execution not found")
    
    # 允许completed和paused状态的任务返回结果
    # paused状态的任务可以返回到目前为止的部分结果
    allowed_statuses = ("completed", "paused")
    if result["status"] not in allowed_statuses:
        raise HTTPException(
            status_code=400,
            detail=f"Task not ready for result retrieval. Current status: {result['status']}. Allowed: {allowed_statuses}"
        )
    
    try:
        # 确保所有结果都经过序列化处理（包括从内存加载的结果）
        outputs = result.get("outputs", {})
        # 如果outputs已经是包装格式，直接序列化；否则先包装
        if outputs and isinstance(outputs, dict):
            first_value = next(iter(outputs.values()), None)
            if isinstance(first_value, dict) and "type" in first_value and "data" in first_value:
                # 已经是包装格式
                serializable_outputs = _ensure_json_serializable(outputs)
            else:
                # 需要包装
                serializable_outputs = _wrap_stored_outputs(outputs)
        else:
            serializable_outputs = {}
        
        serializable_result = {
            "execution_id": result.get("execution_id", execution_id),
            "task_id": result.get("task_id", "unknown"),
            "status": result.get("status", "completed"),
            "outputs": serializable_outputs,
            "errors": result.get("errors", []),
            "record_id": result.get("record_id")  # 用于前端获取历史stages数据
        }
        return JSONResponse(content=serializable_result)
    except Exception as e:
        logger.error(f"Failed to serialize result: {e}, result keys: {result.keys() if result else 'None'}")
        raise HTTPException(
            status_code=500,
            detail="Failed to serialize response"
        )




@router.post("/prompt/build")
async def build_sop_prompt(request: BuildPromptRequest) -> Dict[str, str]:
    """
    构建SOP Prompt（用于LLM调用）
    
    Args:
        request: 包含task_id、params和workspace_files_info
        
    Returns:
        填充参数后的Prompt
    """
    # Get task definition
    registry = get_registry()
    task_def = registry.get_task(request.task_id)
    
    if not task_def:
        raise HTTPException(status_code=404, detail="Task not found")
    
    # Get prompt template
    prompt_template = task_def.sop_prompt_template
    if not prompt_template:
        raise HTTPException(status_code=404, detail=f"No prompt template for task: {request.task_id}")
    
    # Build filled prompt
    try:
        # Use the meta module's build function if available
        if request.task_id == "rule_mining":
            from deepanalyze.analysis.task_SOP.rule_mining_meta import build_sop_prompt
            filled_prompt = build_sop_prompt(
                params=request.params,
                workspace_files_info=request.workspace_files_info
            )
        else:
            # Generic filling
            filled_prompt = prompt_template.format(
                **request.params,
                workspace_files_info=request.workspace_files_info
            )
        
        return {"prompt": filled_prompt}
        
    except KeyError as e:
        raise HTTPException(status_code=400, detail=f"Missing parameter: {e.args[0] if e.args else 'unknown'}")
    except Exception as e:
        logger.error(f"Failed to build prompt: {e}")
        raise HTTPException(status_code=500, detail="Failed to build prompt")


# =============================================================================
# LLM+Pipeline 新架构 API 端点
# =============================================================================

@router.post("/llm/extract", response_model=LLMExtractResponse)
async def extract_params_from_chat(request: LLMExtractRequest) -> LLMExtractResponse:
    """
    从自然语言对话中提取任务参数（Chat 入口）
    
    使用 LLM 分析用户消息，识别任务意图并提取参数。
    这是 LLM+Pipeline 新架构的核心入口之一。
    
    Args:
        request: 包含用户消息、会话ID、工作区文件等
        
    Returns:
        提取的任务类型、参数、置信度等
    """
    # P1-D5: 入口层 session_id 校验
    try:
        request.session_id = _validate_session_id(request.session_id)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    
    if not LLM_PIPELINE_ARCH_AVAILABLE:
        raise HTTPException(
            status_code=501, 
            detail="LLM+Pipeline architecture not available"
        )
    
    try:
        # 创建参数提取器
        extractor = create_param_extractor(
            model=request.model,
            api_base=request.api_base
        )
        
        # 执行参数提取
        intent = await extractor.extract(
            user_message=request.user_message,
            workspace_files=request.workspace_files,
            conversation_history=request.conversation_history
        )
        
        # 生成推断过程的伪代码展示
        extraction_code = f'''# === LLM 参数推断 ===
# 用户请求: "{request.user_message[:100]}..."
# 
# 推断结果:
#   - task_type: "{intent.task_type}" (置信度: {intent.confidence:.2f})
#   - params: {intent.params}
#   - missing_params: {intent.missing_params}
'''
        
        return LLMExtractResponse(
            success=True,
            task_type=intent.task_type,
            confidence=intent.confidence,
            params=intent.params,
            missing_params=intent.missing_params,
            clarification_needed=intent.clarification_needed,
            clarification_question=intent.clarification_question if intent.clarification_needed else None,
            extraction_code=extraction_code
        )
        
    except Exception as e:
        logger.error(f"LLM parameter extraction failed: {e}", exc_info=True)
        return LLMExtractResponse(
            success=False,
            clarification_needed=True,
            clarification_question=f"参数提取失败，请手动配置任务参数。错误: {str(e)}"
        )


@router.post("/unified/execute", response_model=TaskExecuteResponse)
async def unified_execute_task(
    request: UnifiedExecuteRequest,
    background_tasks: BackgroundTasks
) -> TaskExecuteResponse:
    """
    统一任务执行入口（支持 SOP UI 和 Chat 两种入口）
    
    无论从哪个入口进入，最终都通过 Pipeline 执行引擎执行任务。
    Chat 入口会额外记录 LLM 推断信息。
    
    Args:
        request: 统一执行请求
        background_tasks: FastAPI后台任务
        
    Returns:
        执行ID和初始状态
    """
    # P1-D5: 入口层 session_id 校验
    try:
        request.session_id = _validate_session_id(request.session_id)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    
    # Validate task exists
    registry = get_registry()
    task_def = registry.get_task(request.task_id)
    if not task_def:
        raise HTTPException(status_code=404, detail="Task not found")
    
    # Build full file path
    from pathlib import Path
    workspace_dir = Path("workspace") / request.session_id
    full_file_path = workspace_dir / request.file_path
    
    logger.info(f"Unified execute: source={request.source}, task={request.task_id}, file={full_file_path}")
    
    # Validate file exists
    if not full_file_path.exists():
        raise HTTPException(status_code=404, detail="File not found")
    
    # 如果有新架构可用，使用统一路由器
    if LLM_PIPELINE_ARCH_AVAILABLE:
        try:
            router_instance = create_router()
            
            # 记录入口来源
            entry_source = EntrySource.CHAT if request.source == "chat" else EntrySource.SOP_UI
            
            # 生成任务配置摘要代码
            config_code = generate_task_config_summary(request.task_id, request.params)
            logger.info(f"Task config code generated for {request.task_id}")
            
        except Exception as e:
            logger.warning(f"Failed to use unified router: {e}, falling back to standard execution")
    
    # Create execution context
    context = ExecutionStore.create(
        task_id=request.task_id,
        session_id=request.session_id,
        params=request.params,
        file_path=str(full_file_path),
        interaction_mode=request.interaction_mode,
        model=request.model,
        api_base=request.api_base,
        system_prompt=request.system_prompt
    )
    execution_id = context.execution_id
    
    # 记录入口来源到 metadata
    context.metadata["entry_source"] = request.source
    if request.source == "chat" and request.chat_context:
        context.metadata["chat_context"] = request.chat_context
    
    # Schedule background execution
    executor = get_executor()
    
    async def run_task():
        try:
            await executor.execute_async(
                task_id=request.task_id,
                session_id=request.session_id,
                params=request.params,
                file_path=str(full_file_path),
                execution_id=execution_id,
                interaction_mode=request.interaction_mode,
                model=request.model,
                api_base=request.api_base,
                system_prompt=request.system_prompt
            )
        except asyncio.CancelledError:
            logger.info(f"Task {execution_id} was cancelled")
        except Exception as e:
            import traceback
            logger.error(f"Background task execution failed: {e}\n{traceback.format_exc()}")
            # 更新任务状态为失败
            ctx = ExecutionStore.get(execution_id)
            if ctx:
                ctx.status = ExecutionStatus.FAILED
                ctx.message = f"任务执行失败: {str(e)}"
                ctx.errors.append(traceback.format_exc())
                ExecutionStore.update(ctx)
        finally:
            # Phase 19: 任务完成后注销
            await _unregister_task(execution_id)
    
    # Phase 19: 取消已存在的任务（如果有），然后注册新任务
    await _cancel_existing_task(execution_id)
    task = asyncio.create_task(run_task())
    await _register_task(execution_id, task)
    
    return TaskExecuteResponse(
        execution_id=execution_id,
        task_id=request.task_id,
        status="pending",
        message=f"任务已提交（入口: {request.source}），正在执行中..."
    )


@router.get("/code/{execution_id}/events")
async def get_code_events(execution_id: str) -> List[StageCodeEvent]:
    """
    获取任务执行的代码事件（用于 Code 栏展示）
    
    返回任务执行过程中产生的所有代码事件，包括：
    - config: 任务配置摘要
    - llm_extraction: LLM 参数推断过程（Chat 入口）
    - stage_start: 阶段开始时的伪代码
    - stage_complete: 阶段完成时的结果
    - full_code: 任务完成后的完整等效代码
    
    Args:
        execution_id: 执行ID
        
    Returns:
        代码事件列表
    """
    context = ExecutionStore.get(execution_id)
    if not context:
        raise HTTPException(status_code=404, detail="Execution not found")
    
    events: List[StageCodeEvent] = []
    
    # 1. 添加任务配置事件
    if LLM_PIPELINE_ARCH_AVAILABLE:
        config_code = generate_task_config_summary(context.task_id, context.params)
        events.append(StageCodeEvent(
            event_type="config",
            content=config_code,
            timestamp=context.started_at.timestamp() if context.started_at else datetime.now().timestamp()
        ))
        
        # 2. 如果是 Chat 入口，添加 LLM 推断事件
        if context.metadata.get("entry_source") == "chat":
            chat_ctx = context.metadata.get("chat_context", {})
            user_msg = chat_ctx.get("user_message", "")
            extraction_code = f'''# === LLM 参数推断 ===
# 用户请求: "{user_msg[:100]}..."
# 
# 调用 Pipeline:
# start_{context.task_id}_task(
#     params={context.params},
#     interaction_mode="{context.interaction_mode}"
# )
'''
            events.append(StageCodeEvent(
                event_type="llm_extraction",
                content=extraction_code,
                timestamp=context.started_at.timestamp() if context.started_at else datetime.now().timestamp()
            ))
    
    # 3. 添加各阶段的代码事件
    for stage_id, stage_info in context.stages.items():
        stage_status = stage_info.get("status", "pending")
        stage_name = stage_info.get("stage_name", stage_id)
        
        # 生成阶段伪代码
        if LLM_PIPELINE_ARCH_AVAILABLE:
            pseudo_code = StageCodeGenerator.generate_pseudo_code(
                context.task_id, 
                stage_id, 
                context.params
            )
        else:
            pseudo_code = f"# 执行阶段: {stage_name}"
        
        # 生成结果注释
        result_comment = None
        if stage_status == "completed":
            if LLM_PIPELINE_ARCH_AVAILABLE:
                result_summary = stage_info.get("output_preview", {})
                exec_time = stage_info.get("execution_time_ms", 0)
                result_comment = StageCodeGenerator.generate_stage_result_comment(
                    stage_id, result_summary, exec_time
                )
            else:
                result_comment = f"# → 完成: {stage_info.get('message', '')}"
        
        events.append(StageCodeEvent(
            event_type="stage_start" if stage_status in ("pending", "running") else "stage_complete",
            stage_id=stage_id,
            stage_name=stage_name,
            content=pseudo_code,
            status=stage_status,
            result=result_comment,
            execution_time_ms=stage_info.get("execution_time_ms"),
            timestamp=datetime.now().timestamp()
        ))
    
    # 4. 如果任务完成，添加完整等效代码
    if context.status.value == "completed" and LLM_PIPELINE_ARCH_AVAILABLE:
        full_code = EquivalentCodeGenerator.generate_equivalent_code(
            context.task_id,
            {**context.params, "file_path": context.file_path}
        )
        events.append(StageCodeEvent(
            event_type="full_code",
            content=full_code,
            timestamp=context.completed_at.timestamp() if context.completed_at else datetime.now().timestamp()
        ))
    
    return events


@router.delete("/executions/{execution_id}")
async def cancel_execution(execution_id: str) -> Dict[str, str]:
    """
    取消/删除执行记录
    
    Args:
        execution_id: 执行ID
        
    Returns:
        操作结果
    """
    if ExecutionStore.delete(execution_id):
        return {"message": f"Execution {execution_id} deleted"}
    else:
        raise HTTPException(status_code=404, detail="Execution not found")


@router.get("/executions")
async def list_executions(
    session_id: Optional[str] = Query(None, description="按会话ID筛选")
) -> List[Dict]:
    """
    列出执行记录
    
    Args:
        session_id: 可选，按会话ID筛选
        
    Returns:
        执行记录列表
    """
    # P1-D5: 入口层 session_id 校验（可选参数，非空时校验）
    if session_id:
        try:
            session_id = _validate_session_id(session_id)
        except ValueError as e:
            raise HTTPException(status_code=400, detail=str(e))
    
    if session_id:
        contexts = ExecutionStore.list_by_session(session_id)
    else:
        contexts = list(ExecutionStore._executions.values())
    
    return [
        {
            "execution_id": ctx.execution_id,
            "task_id": ctx.task_id,
            "session_id": ctx.session_id,
            "status": ctx.status.value,
            "overall_progress": ctx.overall_progress,
            "started_at": ctx.started_at.isoformat() if ctx.started_at else None,
            "completed_at": ctx.completed_at.isoformat() if ctx.completed_at else None
        }
        for ctx in contexts
    ]


@router.get("/docs/{doc_name}")
async def get_task_guide_doc(doc_name: str) -> Dict[str, str]:
    """
    获取任务操作指引文档内容
    
    Args:
        doc_name: 文档名称（不含扩展名），如 rule_mining_workflow
        
    Returns:
        文档内容（Markdown格式）
    """
    # 文档目录路径
    docs_dir = os.path.join(
        os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
        "docs", "taskSOP_solution"
    )
    
    # 安全检查：只允许访问特定文档
    allowed_docs = [
        "rule_mining_workflow",
        "scorecard_dev_workflow",
    ]
    
    if doc_name not in allowed_docs:
        raise HTTPException(status_code=404, detail=f"Document not found: {doc_name}")
    
    doc_path = os.path.join(docs_dir, f"{doc_name}.md")
    
    if not os.path.exists(doc_path):
        raise HTTPException(status_code=404, detail=f"Document file not found: {doc_name}")
    
    try:
        with open(doc_path, "r", encoding="utf-8") as f:
            content = f.read()
        return {"content": content, "doc_name": doc_name}
    except Exception as e:
        logger.error(f"Failed to read document {doc_name}: {e}")
        raise HTTPException(status_code=500, detail="Failed to read document")


# =============================================================================
# Score Conversion API (P0 Feature - Upgrade Plan)
# =============================================================================

class ScoreConvertRequest(BaseModel):
    """评分转换请求"""
    values: List[float] = Field(..., description="要转换的值列表")
    direction: str = Field(default="to_prob", description="转换方向: 'to_prob' (评分→概率) 或 'to_score' (概率→评分)")
    scale_params: Dict[str, float] = Field(
        default=None,
        description="评分刻度参数: base_score, pdo, bad_rate"
    )


class ScoreConvertResponse(BaseModel):
    """评分转换响应"""
    success: bool
    results: List[Dict[str, float]] = []
    scale_info: Dict[str, Any] = {}
    error: str = ""


@router.post("/score/convert", response_model=ScoreConvertResponse)
async def convert_score(request: ScoreConvertRequest) -> ScoreConvertResponse:
    """
    评分与概率双向转换
    
    支持:
    - 评分 → 概率 (direction='to_prob')
    - 概率 → 评分 (direction='to_score')
    
    Args:
        request: 包含values, direction, scale_params
        
    Returns:
        转换结果列表和刻度信息
    """
    try:
        from deepanalyze.analysis.score_transformer import ScoreTransformer
        
        # Parse scale parameters
        scale_params = request.scale_params or {}
        base_score = scale_params.get('base_score', 660)
        pdo = scale_params.get('pdo', 75)
        bad_rate = scale_params.get('bad_rate', 0.15)
        
        # Create transformer
        transformer = ScoreTransformer(
            base_score=base_score,
            pdo=pdo,
            bad_rate=bad_rate
        )
        transformer.fit()
        
        # Convert values
        results = transformer.convert(request.values, request.direction)
        scale_info = transformer.get_scale_info()
        
        return ScoreConvertResponse(
            success=True,
            results=results,
            scale_info=scale_info
        )
        
    except Exception as e:
        logger.error(f"Score conversion failed: {e}")
        return ScoreConvertResponse(
            success=False,
            error="Score conversion failed"
        )


class ReportExportRequest(BaseModel):
    """报告导出请求"""
    execution_id: str = Field(..., description="执行ID")
    format: str = Field(default="html", description="导出格式：html/excel/word")
    sections: List[str] = Field(
        default=None,
        description="导出章节列表（可选，None=全部）。评分卡: overview/bins/scorecard/statistics/evaluation/charts; 规则挖掘: overview/rules/evaluation/amount_analysis/charts"
    )
    style: str = Field(default="professional", description="样式模板: professional/simple/colorful")


class ReportExportResponse(BaseModel):
    """报告导出响应"""
    success: bool
    format: str
    content: str = ""
    filename: str = ""
    download_url: str = ""  # For Excel files
    error: str = ""


def _generate_report_filename(task_id: str, execution_id: str, ext: str, result: dict = None) -> str:
    """
    统一生成报告文件名
    
    格式: {task_type}_{timestamp}_{rec-id}.{ext}
    示例: scorecard_20260226_153045_rec-3b08d5757750.html
    
    统一使用 rec- 前缀的 record_id，确保文件名格式一致
    """
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    
    # Task ID 映射：简化显示名称
    task_name_map = {
        'scorecard_dev': 'scorecard',
        'rule_mining': 'rule_mining',
    }
    display_task_id = task_name_map.get(task_id, task_id)
    
    # 优先从 result 获取 record_id
    record_id = result.get('record_id') if result else None
    
    # 如果 result 中没有，尝试从数据库查询
    if not record_id:
        try:
            from deepanalyze.core.task_manager.task_history_service import TaskHistoryService
            record = TaskHistoryService.get_record_by_execution_id(execution_id)
            if record:
                record_id = record.get('record_id')
        except Exception:
            pass
    
    # 确定最终使用的ID
    if execution_id.startswith('rec:rec-'):
        # 前端传来的历史记录ID格式：rec:rec-xxx → 直接使用 rec-xxx（去除 rec: 前缀）
        clean_id = execution_id[4:]
    elif execution_id.startswith('rec:'):
        # 前端传来的历史记录ID格式：rec:xxx（没有rec-前缀）→ 转换为 rec-xxx
        clean_id = 'rec-' + execution_id[4:]
    elif execution_id.startswith('rec-'):
        # 已经是 rec-xxx 格式，直接使用
        clean_id = execution_id
    elif record_id:
        # 有 record_id，优先使用
        clean_id = record_id
    else:
        # 回退：使用 execution_id，移除 exec- 前缀
        clean_id = execution_id.replace('exec-', '')
        if not clean_id.startswith('rec-'):
            clean_id = 'rec-' + clean_id
    
    return f"{display_task_id}_{timestamp}_{clean_id}.{ext}"


@router.post("/report/export", response_model=ReportExportResponse)
async def export_task_report(request: ReportExportRequest) -> ReportExportResponse:
    """
    导出任务执行报告（HTML格式）
    
    Args:
        request: 包含execution_id和format
        
    Returns:
        HTML报告内容
    """
    execution_id = request.execution_id
    
    # 🔧 处理 rec: 前缀（历史记录ID格式）
    if execution_id.startswith("rec:"):
        record_id = execution_id[4:]  # 移除 "rec:" 前缀
        logger.info(f"[Report Export] Detected history record ID: {record_id}")
    else:
        record_id = None
    
    # Get execution result - 先尝试从内存获取
    result = get_execution_result(execution_id)
    
    # 如果内存中找不到，尝试从历史存储加载
    if not result and TASK_MANAGER_AVAILABLE:
        # 🔧 如果有 record_id，直接加载历史记录
        if record_id:
            record = TaskHistoryService.get_record(record_id)
            if record:
                storage = get_result_storage()
                try:
                    stored_result = storage.load_result(record_id)
                    if stored_result:
                        wrapped_outputs = _wrap_stored_outputs(stored_result)
                        result = {
                            "execution_id": record.get("execution_id", execution_id),
                            "task_id": record.get("task_type") or record.get("task_id", "unknown"),
                            "status": "completed",
                            "outputs": wrapped_outputs,
                            "errors": [],
                            "record_id": record_id
                        }
                        logger.info(f"[Report Export] Loaded from history record: {record_id}")
                except Exception as e:
                    logger.error(f"[Report Export] Failed to load result for record {record_id}: {e}")
        else:
            # 没有 rec: 前缀，按原有逻辑通过 execution_id 查找
            record = TaskHistoryService.get_record_by_execution_id(execution_id)
            if record:
                found_record_id = record.get("record_id")
                if found_record_id:
                    storage = get_result_storage()
                    try:
                        stored_result = storage.load_result(found_record_id)
                        if stored_result:
                            wrapped_outputs = _wrap_stored_outputs(stored_result)
                            # 注意：历史记录中的字段是 task_type，不是 task_id
                            result = {
                                "execution_id": execution_id,
                                "task_id": record.get("task_type") or record.get("task_id", "unknown"),
                                "status": "completed",
                                "outputs": wrapped_outputs,
                                "errors": [],
                                "record_id": found_record_id
                            }
                    except Exception as e:
                        logger.error(f"Failed to load result for export: {e}")
    
    # 如果从内存获取了结果但没有 record_id，尝试从历史记录补充
    if result and not result.get("record_id") and TASK_MANAGER_AVAILABLE:
        # 🔧 优先使用 rec: 前缀中提取的 record_id
        if record_id:
            result["record_id"] = record_id
            logger.info(f"[Report Export] Supplemented record_id from rec: prefix: {record_id}")
        else:
            record = TaskHistoryService.get_record_by_execution_id(execution_id)
            logger.info(f"Report export: Looking up record_id for execution {execution_id}, found record={record is not None}")
            if record:
                result["record_id"] = record.get("record_id")
                logger.info(f"Report export: Supplemented record_id={result.get('record_id')}")
    
    if not result:
        return ReportExportResponse(
            success=False,
            format=request.format,
            error=f"Execution not found: {request.execution_id}"
        )
    
    if result["status"] != "completed":
        return ReportExportResponse(
            success=False,
            format=request.format,
            error=f"Task not completed. Status: {result['status']}"
        )
    
    task_id = result.get("task_id", "")
    outputs = result.get("outputs", {})
    
    # 获取stages数据（用于Excel导出各阶段下载结果）
    stages_data = {}
    record_id = result.get("record_id")
    logger.info(f"[Report Export] record_id from result: {record_id}")
    if record_id and TASK_MANAGER_AVAILABLE:
        record = TaskHistoryService.get_record(record_id)
        if record and record.get("stages"):
            stages_data = record.get("stages", {})
            logger.info(f"[Report Export] stages_data from TaskHistory: {list(stages_data.keys())}")
            # 检查数据结构，确保 output_preview 正确
            for sid, sdata in stages_data.items():
                has_output = 'output_preview' in sdata if isinstance(sdata, dict) else False
                if has_output:
                    output_preview = sdata.get('output_preview', {})
                    # 特别检查 feature_engineering 阶段的关键字段（规则挖掘任务）
                    if sid == 'feature_engineering':
                        logger.info(f"[Report Export] feature_engineering output_preview keys: {list(output_preview.keys()) if isinstance(output_preview, dict) else 'not a dict'}")
                        logger.info(f"[Report Export] feature_engineering.before_count: {output_preview.get('before_count')}")
                        logger.info(f"[Report Export] feature_engineering.after_count: {output_preview.get('after_count')}")
                        logger.info(f"[Report Export] feature_engineering.selection_flow: {output_preview.get('selection_flow')}")
                        logger.info(f"[Report Export] feature_engineering.skipped: {output_preview.get('skipped')}")
                    # 特别检查 data_loading 阶段的关键字段（评分卡任务）
                    if sid == 'data_loading':
                        logger.info(f"[Report Export] data_loading output_preview keys: {list(output_preview.keys()) if isinstance(output_preview, dict) else 'not a dict'}")
                        logger.info(f"[Report Export] data_loading.rows: {output_preview.get('rows')}")
                        logger.info(f"[Report Export] data_loading.target_rate: {output_preview.get('target_rate')}")
                        logger.info(f"[Report Export] data_loading.split_info: {output_preview.get('split_info')}")
                        logger.info(f"[Report Export] data_loading.var_filter_result: {output_preview.get('var_filter_result')}")
                logger.debug(f"[Report Export] Stage {sid}: has_output_preview={has_output}")
    # 如果没有record_id，尝试从execution context获取
    if not stages_data:
        context = get_execution_context(request.execution_id)
        if context and context.stages:
            for stage_id, stage in context.stages.items():
                if stage.output_preview:
                    stages_data[stage_id] = {
                        "stage_id": stage.stage_id,
                        "stage_name": stage.stage_name,
                        "status": stage.status.value if hasattr(stage.status, 'value') else stage.status,
                        "output_preview": stage.output_preview
                    }
            logger.info(f"[Report Export] stages_data from ExecutionContext: {list(stages_data.keys())}")
    
    if not stages_data:
        logger.warning(f"[Report Export] No stages_data available for report generation")
    
    try:
        # Handle Excel format export (new feature)
        if request.format == "excel":
            from deepanalyze.analysis.excel_report import ExcelReportGenerator
            import base64
            from datetime import datetime
            
            generator = ExcelReportGenerator(style_template=request.style)
            
            # 获取 AI 分析（与 Word/MD 报告一致）
            ai_analysis_text = None
            try:
                record_id = result.get('record_id')
                if not record_id:
                    context = get_execution_context(request.execution_id)
                    record_id = getattr(context, 'record_id', None) if context else None
                
                if record_id and StageAnalysisService:
                    # 1. 优先查找任务级 AI 分析
                    analysis = StageAnalysisService.get_analysis(record_id, "_task_analysis")
                    # 2. 如果任务级分析不存在，尝试获取最后阶段的分析
                    if not analysis or not analysis.get('analysis_text'):
                        analysis = StageAnalysisService.get_analysis(record_id, "report_generation")
                    if analysis and analysis.get('analysis_text'):
                        ai_analysis_text = analysis['analysis_text']
                        logger.info(f"Found AI analysis for Excel report: {len(ai_analysis_text)} chars")
            except Exception as e:
                logger.debug(f"AI analysis not available for Excel report (this is OK): {e}")
            
            if task_id == "scorecard_dev":
                # Prepare results dict for Excel generator
                metrics = outputs.get("metrics", {}).get("data", {})
                
                # Extract additional data for comprehensive report
                selected_features = outputs.get("selected_features", {}).get("data", None)
                if isinstance(selected_features, str):
                    try:
                        import json
                        selected_features = json.loads(selected_features.replace("'", '"'))
                    except (json.JSONDecodeError, ValueError):
                        selected_features = [f.strip() for f in selected_features.split(',') if f.strip()]
                
                overfit_warning = outputs.get("overfit_warning", {}).get("data", None)
                selection_detail = outputs.get("selection_detail", {}).get("data", None)
                outlier_info = outputs.get("outlier_info", {}).get("data", None)
                
                # Parse IV table
                iv_table_data = outputs.get("iv_table", {}).get("data", [])
                
                # Parse coefficients
                coef_data = outputs.get("coefficients", {}).get("data", [])
                
                # 获取 stage_configs 配置（用于阶段 Sheet 命名和下载数据字段）
                from deepanalyze.core.task_manager.task_result_config import get_task_result_config
                task_result_config = get_task_result_config(task_id)
                stage_configs = None
                if task_result_config and task_result_config.excel_config:
                    stage_configs = task_result_config.excel_config.stage_data_sheets
                    logger.info(f"Scorecard Excel export: loaded {len(stage_configs)} stage configs")
                
                # Extract multi-dataset metrics and chart data (for overview and evaluation sections)
                multi_dataset_metrics = outputs.get("multi_dataset_metrics", {}).get("data", {})
                multi_dataset_chart_data = outputs.get("multi_dataset_chart_data", {}).get("data", {})
                psi_result_data = outputs.get("psi_result", {}).get("data")
                psi_train_vs_test_data = outputs.get("psi_train_vs_test", {}).get("data")
                psi_train_vs_oot_data = outputs.get("psi_train_vs_oot", {}).get("data")
                
                # 调试日志：检查关键数据是否存在
                logger.info(f"[Excel Export Debug] metrics keys: {list(metrics.keys()) if metrics else 'empty'}")
                logger.info(f"[Excel Export Debug] multi_dataset_metrics keys: {list(multi_dataset_metrics.keys()) if multi_dataset_metrics else 'empty'}")
                logger.info(f"[Excel Export Debug] psi_result: {psi_result_data is not None}, psi_train_vs_test: {psi_train_vs_test_data is not None}, psi_train_vs_oot: {psi_train_vs_oot_data is not None}")
                
                excel_results = {
                    'n_samples': metrics.get("n_samples"),
                    'n_features': metrics.get("n_features"),
                    'train_auc': metrics.get("train_auc") or metrics.get("auc"),
                    'test_auc': metrics.get("test_auc"),
                    'train_ks': metrics.get("train_ks") or metrics.get("ks"),
                    'test_ks': metrics.get("test_ks"),
                    'model_statistics': outputs.get("model_statistics", {}).get("data"),
                    'scorecard': outputs.get("scorecard", {}).get("data"),
                    'woe_results': outputs.get("woe_results", {}).get("data"),
                    # Additional data for comprehensive report
                    'selected_features': selected_features,
                    'overfit_warning': overfit_warning,
                    'selection_detail': selection_detail,
                    'outlier_info': outlier_info,
                    'iv_table': iv_table_data,
                    'coefficients': coef_data,
                    # 添加stages数据用于生成各阶段Sheet
                    'stages': stages_data,
                    # 添加 stage_configs 配置（用于阶段 Sheet 命名和下载数据字段映射）
                    'stage_configs': stage_configs,
                    # 修复：添加概览和评估图表需要的核心数据字段
                    'metrics': metrics,  # 核心指标（KS、AUC、Gini等）
                    'multi_dataset_metrics': multi_dataset_metrics,  # 数据集指标对比
                    'multi_dataset_chart_data': multi_dataset_chart_data,  # 图表数据
                    'psi_result': psi_result_data,  # PSI稳定性指标
                    'psi_train_vs_test': psi_train_vs_test_data,  # 训练集vs测试集PSI
                    'psi_train_vs_oot': psi_train_vs_oot_data,  # 训练集vsOOT的PSI
                }
                
                excel_bytes = generator.generate_scorecard_report(
                    results=excel_results,
                    sections=request.sections,
                    title='评分卡开发报告',
                    ai_analysis=ai_analysis_text
                )
                filename = _generate_report_filename(task_id, request.execution_id, "xlsx", result)

            elif task_id == "rule_mining":
                # Prepare results dict for Excel generator
                # 优先使用 optimal_rules，回退到 rules
                rules_data = outputs.get("optimal_rules", {}).get("data", 
                             outputs.get("rules", {}).get("data", []))
                filtered_rules_data = outputs.get("filtered_rules", {}).get("data", [])
                all_rules_data = outputs.get("rules", {}).get("data", rules_data)
                
                # 获取 stage_configs 配置（用于阶段 Sheet 命名和下载数据字段）
                from deepanalyze.core.task_manager.task_result_config import get_task_result_config
                task_result_config = get_task_result_config(task_id)
                stage_configs = None
                if task_result_config and task_result_config.excel_config:
                    stage_configs = task_result_config.excel_config.stage_data_sheets
                    logger.info(f"Rule mining Excel export: loaded {len(stage_configs)} stage configs")
                
                # 用最新的 all_rules_with_status 更新 rule_filtering 阶段的 output_preview
                # 因为 all_rules_with_status 在 selecting_rules 阶段会被更新（添加 is_optimal、rejection_reason）
                # 但更新后的数据存储在 results['all_rules_with_status']，不在阶段的 output_preview 中
                updated_all_rules_with_status = outputs.get("all_rules_with_status", {}).get("data", [])
                if updated_all_rules_with_status and stages_data.get('rule_filtering'):
                    if 'output_preview' in stages_data['rule_filtering']:
                        stages_data['rule_filtering']['output_preview']['all_rules_with_status'] = updated_all_rules_with_status
                        logger.info(f"Rule mining Excel export: updated rule_filtering.output_preview.all_rules_with_status with {len(updated_all_rules_with_status)} rules")
                    else:
                        stages_data['rule_filtering']['output_preview'] = {'all_rules_with_status': updated_all_rules_with_status}
                
                excel_results = {
                    'n_samples': outputs.get("preprocessing_info", {}).get("data", {}).get("n_samples"),
                    'n_bad': outputs.get("preprocessing_info", {}).get("data", {}).get("n_bad"),
                    'base_bad_rate': outputs.get("preprocessing_info", {}).get("data", {}).get("base_bad_rate"),
                    'n_rules_generated': outputs.get("preprocessing_info", {}).get("data", {}).get("n_rules_generated"),
                    'n_rules_selected': len(rules_data),
                    'total_recall': rules_data[-1].get('dev_cum_recall', rules_data[-1].get('cumulative_recall', 0)) if rules_data else 0,
                    'selected_rules': rules_data,
                    'filtered_rules': filtered_rules_data,
                    'all_rules': all_rules_data,
                    'validation_report': outputs.get("validation_report", {}).get("data"),
                    'psi_report': outputs.get("psi_report", {}).get("data"),
                    'amount_analysis': outputs.get("amount_analysis", {}).get("data"),
                    'prior_analysis': outputs.get("prior_analysis", {}).get("data"),
                    # 添加 optimal_rules 用于任务报告 Sheet
                    'optimal_rules': rules_data,
                    # 添加 AI 分析用于任务报告 Sheet
                    'ai_analysis': ai_analysis_text,
                    # 添加stages数据用于生成各阶段Sheet
                    'stages': stages_data,
                    # 添加 stage_configs 配置（用于阶段 Sheet 命名和下载数据字段映射）
                    'stage_configs': stage_configs,
                    # 添加 all_rules_with_status（规则过滤阶段的完整规则列表，包含过滤状态）
                    'all_rules_with_status': outputs.get("all_rules_with_status", {}).get("data", []),
                }
                
                excel_bytes = generator.generate_rule_mining_report(
                    results=excel_results,
                    sections=request.sections,
                    title='规则挖掘报告'
                )
                filename = _generate_report_filename(task_id, request.execution_id, "xlsx", result)

            else:
                return ReportExportResponse(
                    success=False,
                    format=request.format,
                    error=f"Excel export not supported for task: {task_id}"
                )
            
            # Encode Excel bytes to base64 for JSON response
            content_base64 = base64.b64encode(excel_bytes).decode('utf-8')
            
            return ReportExportResponse(
                success=True,
                format=request.format,
                content=content_base64,
                filename=filename
            )
        
        # Word format export
        if request.format == "word":
            from deepanalyze.analysis.word_report import generate_word_report
            import base64
            from datetime import datetime
            
            # 尝试获取AI分析（可选，失败不影响报告生成）
            # 优先查找 _task_analysis（任务级分析），其次查找 report_generation（最后阶段分析）
            ai_analysis_text = None
            try:
                # 优先使用result中的record_id（从历史记录加载时会有此字段）
                record_id = result.get('record_id')
                if not record_id:
                    context = get_execution_context(request.execution_id)
                    record_id = getattr(context, 'record_id', None) if context else None
                
                if record_id and StageAnalysisService:
                    # 1. 优先查找任务级 AI 分析
                    analysis = StageAnalysisService.get_analysis(record_id, "_task_analysis")
                    # 2. 如果任务级分析不存在，尝试获取最后阶段的分析
                    if not analysis or not analysis.get('analysis_text'):
                        analysis = StageAnalysisService.get_analysis(record_id, "report_generation")
                    if analysis and analysis.get('analysis_text'):
                        ai_analysis_text = analysis['analysis_text']
                        logger.info(f"Found AI analysis for Word report: {len(ai_analysis_text)} chars")
            except Exception as e:
                logger.debug(f"AI analysis not available for report (this is OK): {e}")
            
            if task_id == "scorecard_dev":
                metrics = outputs.get("metrics", {}).get("data", {})
                selected_features = outputs.get("selected_features", {}).get("data", None)
                if isinstance(selected_features, str):
                    try:
                        import json
                        selected_features = json.loads(selected_features.replace("'", '"'))
                    except (json.JSONDecodeError, ValueError):
                        selected_features = [f.strip() for f in selected_features.split(',') if f.strip()]
                
                word_results = {
                    'metrics': metrics,
                    'scorecard': outputs.get("scorecard", {}).get("data"),
                    'iv_table': outputs.get("iv_table", {}).get("data", []),
                    'coefficients': outputs.get("coefficients", {}).get("data", []),
                    'selected_features': selected_features,
                    'selection_detail': outputs.get("selection_detail", {}).get("data"),
                    'outlier_info': outputs.get("outlier_info", {}).get("data"),
                    'model_statistics': outputs.get("model_statistics", {}).get("data"),
                    # 新增：概览章节需要的数据
                    'multi_dataset_metrics': outputs.get("multi_dataset_metrics", {}).get("data", {}),
                    'multi_dataset_chart_data': outputs.get("multi_dataset_chart_data", {}).get("data", {}),  # 第三部分图表数据
                    'psi_result': outputs.get("psi_result", {}).get("data"),
                    'psi_train_vs_test': outputs.get("psi_train_vs_test", {}).get("data"),
                    'psi_train_vs_oot': outputs.get("psi_train_vs_oot", {}).get("data"),
                    'stages': stages_data,  # 样本与特征章节需要
                }
                
                word_bytes = generate_word_report(
                    results=word_results,
                    report_type='scorecard',
                    title='评分卡开发报告',
                    ai_analysis=ai_analysis_text  # 可选，None时优雅处理
                )
                filename = _generate_report_filename(task_id, request.execution_id, "docx", result)

            elif task_id == "rule_mining":
                rules_data = outputs.get("rules", {}).get("data", [])
                word_results = {
                    'preprocessing_info': outputs.get("preprocessing_info", {}).get("data", {}),
                    'rules': rules_data,
                    'optimal_rules': outputs.get("optimal_rules", {}).get("data", rules_data),
                    'filtered_rules': outputs.get("filtered_rules", {}).get("data", []),
                    'all_rules_with_status': outputs.get("all_rules_with_status", {}).get("data", []),  # 用于过滤规则展示
                    'validation_report': outputs.get("validation_report", {}).get("data"),
                    'psi_report': outputs.get("psi_report", {}).get("data"),
                    'amount_analysis': outputs.get("amount_analysis", {}).get("data"),
                    'prior_analysis': outputs.get("prior_analysis", {}).get("data"),  # 先验规则分析
                    'stages': stages_data,  # 添加stages数据用于样本集特征章节
                }
                
                word_bytes = generate_word_report(
                    results=word_results,
                    report_type='rule_mining',
                    title='规则挖掘报告',
                    ai_analysis=ai_analysis_text  # 可选，None时优雅处理
                )
                filename = _generate_report_filename(task_id, request.execution_id, "docx", result)

            else:
                return ReportExportResponse(
                    success=False,
                    format=request.format,
                    error=f"Word export not supported for task: {task_id}"
                )
            
            content_base64 = base64.b64encode(word_bytes).decode('utf-8')
            
            return ReportExportResponse(
                success=True,
                format=request.format,
                content=content_base64,
                filename=filename
            )
        
        # Markdown format export (new: backend-generated, config-driven)
        if request.format == "markdown" or request.format == "md":
            from deepanalyze.analysis.markdown_report import generate_markdown_report
            from deepanalyze.core.task_manager.task_result_config import get_task_result_config
            from datetime import datetime
            
            # 获取任务结果配置
            task_result_config = get_task_result_config(task_id)
            stage_configs = None
            if task_result_config and task_result_config.excel_config:
                stage_configs = task_result_config.excel_config.stage_data_sheets
            
            # 尝试获取AI分析（可选）
            # 优先查找 _task_analysis（任务级分析），其次查找 report_generation（最后阶段分析）
            ai_analysis_text = None
            try:
                # 优先使用result中的record_id（从历史记录加载时会有此字段）
                record_id = result.get('record_id')
                logger.info(f"[MD Export] record_id from result: {record_id}")
                if not record_id:
                    context = get_execution_context(request.execution_id)
                    record_id = getattr(context, 'record_id', None) if context else None
                    logger.info(f"[MD Export] record_id from context: {record_id}")
                
                if record_id and StageAnalysisService:
                    # 1. 优先查找任务级 AI 分析
                    analysis = StageAnalysisService.get_analysis(record_id, "_task_analysis")
                    logger.info(f"[MD Export] _task_analysis found: {analysis is not None}")
                    # 2. 如果任务级分析不存在，尝试获取最后阶段的分析
                    if not analysis or not analysis.get('analysis_text'):
                        analysis = StageAnalysisService.get_analysis(record_id, "report_generation")
                        logger.info(f"[MD Export] report_generation found: {analysis is not None}")
                    if analysis and analysis.get('analysis_text'):
                        ai_analysis_text = analysis['analysis_text']
                        logger.info(f"[MD Export] AI analysis loaded: {len(ai_analysis_text)} chars")
                else:
                    logger.warning(f"[MD Export] No record_id available for AI analysis lookup")
            except Exception as e:
                logger.warning(f"[MD Export] Failed to get AI analysis: {e}")  # AI分析不可用不影响报告生成
            
            if task_id == "scorecard_dev":
                metrics = outputs.get("metrics", {}).get("data", {})
                selected_features = outputs.get("selected_features", {}).get("data", None)
                if isinstance(selected_features, str):
                    try:
                        import json
                        selected_features = json.loads(selected_features.replace("'", '"'))
                    except (json.JSONDecodeError, ValueError):
                        selected_features = [f.strip() for f in selected_features.split(',') if f.strip()]
                
                md_results = {
                    'metrics': metrics,
                    'scorecard': outputs.get("scorecard", {}).get("data"),
                    'iv_table': outputs.get("iv_table", {}).get("data", []),
                    'coefficients': outputs.get("coefficients", {}).get("data", []),
                    'selected_features': selected_features,
                    'selection_detail': outputs.get("selection_detail", {}).get("data"),
                    'outlier_info': outputs.get("outlier_info", {}).get("data"),
                    'model_statistics': outputs.get("model_statistics", {}).get("data"),
                    'overfit_warning': outputs.get("overfit_warning", {}).get("data"),
                    # 2026-02-10: 添加多数据集指标和图表数据（用于数据集指标对比表）
                    'multi_dataset_metrics': outputs.get("multi_dataset_metrics", {}).get("data"),
                    'multi_dataset_chart_data': outputs.get("multi_dataset_chart_data", {}).get("data"),
                    'psi_result': outputs.get("psi_result", {}).get("data"),
                    'psi_train_vs_test': outputs.get("psi_train_vs_test", {}).get("data"),
                    'psi_train_vs_oot': outputs.get("psi_train_vs_oot", {}).get("data"),
                    'stages': stages_data,
                    'stage_configs': stage_configs,
                }
                
                md_content = generate_markdown_report(
                    results=md_results,
                    report_type='scorecard',
                    title='评分卡开发报告',
                    ai_analysis=ai_analysis_text,
                    config=task_result_config
                )
                filename = _generate_report_filename(task_id, request.execution_id, "md", result)

            elif task_id == "rule_mining":
                rules_data = outputs.get("rules", {}).get("data", [])
                md_results = {
                    'preprocessing_info': outputs.get("preprocessing_info", {}).get("data", {}),
                    'rules': rules_data,
                    'optimal_rules': outputs.get("optimal_rules", {}).get("data", rules_data),
                    'filtered_rules': outputs.get("filtered_rules", {}).get("data", []),
                    'all_rules_with_status': outputs.get("all_rules_with_status", {}).get("data", []),  # 用于过滤规则展示
                    'validation_report': outputs.get("validation_report", {}).get("data"),
                    'psi_report': outputs.get("psi_report", {}).get("data"),
                    'amount_analysis': outputs.get("amount_analysis", {}).get("data"),
                    'prior_analysis': outputs.get("prior_analysis", {}).get("data"),  # 先验规则分析
                    'stages': stages_data,
                    'stage_configs': stage_configs,
                }
                
                md_content = generate_markdown_report(
                    results=md_results,
                    report_type='rule_mining',
                    title='规则挖掘分析报告',
                    ai_analysis=ai_analysis_text,
                    config=task_result_config
                )
                filename = _generate_report_filename(task_id, request.execution_id, "md", result)

            else:
                return ReportExportResponse(
                    success=False,
                    format=request.format,
                    error=f"Markdown export not supported for task: {task_id}"
                )
            
            return ReportExportResponse(
                success=True,
                format="markdown",
                content=md_content,  # MD是文本格式，直接返回内容
                filename=filename
            )
        
        # HTML format export (existing logic)
        if task_id == "scorecard_dev":
            # Generate scorecard report HTML
            from deepanalyze.analysis.html_report import generate_html_report
            import numpy as np
            
            # Extract data from outputs
            metrics = outputs.get("metrics", {}).get("data", {})
            
            # Parse IV table
            iv_table_data = outputs.get("iv_table", {}).get("data", [])
            iv_table = pd.DataFrame(iv_table_data) if iv_table_data else pd.DataFrame()
            
            # Parse scorecard - handle {columns: [...], data: [...]} format
            scorecard_data = outputs.get("scorecard", {}).get("data", {})
            scorecard_rows = []
            if isinstance(scorecard_data, dict):
                for variable, var_data in scorecard_data.items():
                    # Handle {columns: [...], data: [...]} format (dict_of_dataframes)
                    if isinstance(var_data, dict) and 'data' in var_data and isinstance(var_data['data'], list):
                        for row in var_data['data']:
                            scorecard_rows.append({
                                'variable': variable,
                                'bin': row.get('bin', row.get('Bin', '')),
                                'woe': row.get('woe', row.get('WOE', 0)),
                                'points': row.get('points', row.get('Points', 0))
                            })
                    # Handle direct list format
                    elif isinstance(var_data, list):
                        for row in var_data:
                            scorecard_rows.append({
                                'variable': variable,
                                'bin': row.get('bin', row.get('Bin', '')),
                                'woe': row.get('woe', row.get('WOE', 0)),
                                'points': row.get('points', row.get('Points', 0))
                            })
            scorecard = pd.DataFrame(scorecard_rows) if scorecard_rows else pd.DataFrame()
            
            # Parse coefficients
            coef_data = outputs.get("coefficients", {}).get("data", [])
            coefficients = pd.DataFrame(coef_data) if coef_data else pd.DataFrame()
            
            # Extract chart_data for ROC/KS/Score distribution charts
            chart_data = outputs.get("chart_data", {}).get("data", None)
            
            # Extract multi-dataset metrics and chart data
            multi_dataset_metrics = outputs.get("multi_dataset_metrics", {}).get("data", None)
            multi_dataset_chart_data = outputs.get("multi_dataset_chart_data", {}).get("data", None)
            
            # Extract additional data for comprehensive report
            # 筛选后的特征列表
            selected_features = outputs.get("selected_features", {}).get("data", None)
            if isinstance(selected_features, str):
                try:
                    import json
                    selected_features = json.loads(selected_features.replace("'", '"'))
                except (json.JSONDecodeError, ValueError):
                    selected_features = [f.strip() for f in selected_features.split(',') if f.strip()]
            
            # 过拟合警告
            overfit_warning = outputs.get("overfit_warning", {}).get("data", None)
            
            # 特征选择详情（逐步回归、系数验证等）
            selection_detail = outputs.get("selection_detail", {}).get("data", None)
            
            # 异常值检测结果
            outlier_info = outputs.get("outlier_info", {}).get("data", None)
            
            # AI analysis - 优先从持久化存储获取（与前端展示一致），回退到outputs
            # 这确保报告中的AI分析与前端展示的完全一致
            ai_analysis = None
            try:
                # 优先使用result中的record_id（从历史记录加载时会有此字段）
                record_id = result.get('record_id')
                if not record_id:
                    # 回退：从执行上下文获取
                    context = get_execution_context(request.execution_id)
                    record_id = getattr(context, 'record_id', None) if context else None
                
                logger.info(f"Scorecard report export: record_id={record_id}")
                if record_id and StageAnalysisService:
                    # 1. 优先查找任务级 AI 分析（与前端展示一致）
                    analysis_result = StageAnalysisService.get_analysis(record_id, "_task_analysis")
                    logger.info(f"Scorecard report export: _task_analysis found={analysis_result is not None}")
                    
                    # 2. 如果任务级分析不存在，尝试获取最后阶段的分析
                    if not analysis_result or not analysis_result.get('analysis_text'):
                        analysis_result = StageAnalysisService.get_analysis(record_id, "report_generation")
                        logger.info(f"Scorecard report export: report_generation found={analysis_result is not None}")
                    
                    if analysis_result and analysis_result.get('analysis_text'):
                        ai_analysis = analysis_result['analysis_text']
                        logger.info(f"Scorecard report export: AI analysis loaded from DB, {len(ai_analysis)} chars")
                else:
                    logger.warning(f"Scorecard report export: No record_id found for execution {request.execution_id}")
            except Exception as e:
                logger.warning(f"Scorecard report export: Failed to get AI analysis from DB: {e}")
            
            # 回退：如果持久化存储中没有，则使用outputs中的ai_analysis
            if not ai_analysis:
                ai_analysis = outputs.get("ai_analysis", {}).get("data", None)
                if ai_analysis:
                    logger.info(f"Scorecard report export: AI analysis fallback to outputs, {len(ai_analysis)} chars")
            
            # Model statistics for comprehensive report
            model_statistics = outputs.get("model_statistics", {}).get("data", None)
            
            # Build results dict for unified report generator
            results = {
                'metrics': metrics,
                'iv_table': iv_table,
                'scorecard': scorecard,
                'coefficients': coefficients,
                'chart_data': chart_data,
                'multi_dataset_metrics': multi_dataset_metrics,
                'multi_dataset_chart_data': multi_dataset_chart_data,
                'selected_features': selected_features,
                'overfit_warning': overfit_warning,
                'selection_detail': selection_detail,
                'outlier_info': outlier_info,
                'model_statistics': model_statistics,  # 添加模型统计检验数据
                'psi_result': outputs.get("psi_result", {}).get("data"),  # PSI稳定性指标（与前端一致）
                'stages': stages_data,  # 添加stages数据，用于样本与特征章节
            }
            
            # Generate HTML report using unified entry point
            html_content = generate_html_report(
                task_type='scorecard',
                results=results,
                title='评分卡开发报告',
                ai_analysis=ai_analysis,
                include_charts=True
            )

            filename = _generate_report_filename(task_id, request.execution_id, "html", result)

        elif task_id == "rule_mining":
            # Generate complete rule mining report HTML (aligned with Word report)
            from deepanalyze.analysis.html_report import generate_html_report
            
            # Build complete results dict from outputs
            results = {
                'preprocessing_info': outputs.get("preprocessing_info", {}).get("data", {}),
                'rules': outputs.get("rules", {}).get("data", []),
                'optimal_rules': outputs.get("optimal_rules", {}).get("data", 
                                 outputs.get("rules", {}).get("data", [])),
                'filtered_rules': outputs.get("filtered_rules", {}).get("data", []),
                'all_rules_with_status': outputs.get("all_rules_with_status", {}).get("data", []),  # 用于过滤规则展示
                'validation_report': outputs.get("validation_report", {}).get("data", None),
                'psi_report': outputs.get("psi_report", {}).get("data", None),
                'amount_analysis': outputs.get("amount_analysis", {}).get("data", None),
                'prior_analysis': outputs.get("prior_analysis", {}).get("data", None),  # 先验规则分析
                'stages': stages_data,  # 添加stages数据用于样本集特征章节
            }
            
            # Get AI analysis from StageAnalysisService (consistent with Word/Markdown)
            # 优先查找 _task_analysis（任务级分析），其次查找 report_generation（最后阶段分析）
            ai_analysis = None
            try:
                # 优先使用result中的record_id（从历史记录加载时会有此字段）
                record_id = result.get('record_id')
                if not record_id:
                    # 回退：从执行上下文获取
                    context = get_execution_context(request.execution_id)
                    record_id = getattr(context, 'record_id', None) if context else None
                
                logger.info(f"Report export: record_id={record_id}")
                if record_id and StageAnalysisService:
                    # 1. 优先查找任务级 AI 分析（自动模式点击"AI 分析评估"按钮保存）
                    analysis_result = StageAnalysisService.get_analysis(record_id, "_task_analysis")
                    logger.info(f"Report export: _task_analysis found={analysis_result is not None}")
                    
                    # 2. 如果任务级分析不存在，尝试获取最后阶段（report_generation）的分析
                    if not analysis_result or not analysis_result.get('analysis_text'):
                        analysis_result = StageAnalysisService.get_analysis(record_id, "report_generation")
                        logger.info(f"Report export: report_generation found={analysis_result is not None}")
                    
                    if analysis_result and analysis_result.get('analysis_text'):
                        ai_analysis = analysis_result['analysis_text']
                        logger.info(f"Report export: AI analysis loaded, {len(ai_analysis)} chars")
                else:
                    logger.warning(f"Report export: No record_id found for execution {request.execution_id}")
            except Exception as e:
                logger.warning(f"Report export: Failed to get AI analysis: {e}")
            
            # Generate complete HTML report
            html_content = generate_html_report(
                task_type='rule_mining',
                results=results,
                title='规则挖掘报告',
                ai_analysis=ai_analysis,
                include_charts=True
            )

            filename = _generate_report_filename(task_id, request.execution_id, "html", result)

        else:
            return ReportExportResponse(
                success=False,
                format=request.format,
                error=f"Report export not supported for task: {task_id}"
            )
        
        return ReportExportResponse(
            success=True,
            format=request.format,
            content=html_content,
            filename=filename
        )
        
    except Exception as e:
        logger.error(f"Failed to generate report: {e}", exc_info=True)
        return ReportExportResponse(
            success=False,
            format=request.format,
            error="Failed to generate report"
        )


# =============================================================================
# Task Control Endpoints (Pause/Stop/Resume)
# =============================================================================

class TaskControlResponse(BaseModel):
    """任务控制响应"""
    success: bool
    message: str
    execution_id: str
    current_status: Optional[str] = None


@router.post("/executions/{execution_id}/pause", response_model=TaskControlResponse)
async def pause_execution(execution_id: str) -> TaskControlResponse:
    """
    暂停任务执行
    
    任务将在当前阶段完成后暂停。
    
    Args:
        execution_id: 执行ID
        
    Returns:
        操作结果
    """
    if not TASK_MANAGER_AVAILABLE:
        raise HTTPException(status_code=501, detail="Task manager not available")
    
    context = ExecutionStore.get(execution_id)
    if not context:
        raise HTTPException(status_code=404, detail="Execution not found")
    
    if context.status.value not in ("running", "pending"):
        raise HTTPException(
            status_code=400, 
            detail=f"Cannot pause task in status: {context.status.value}"
        )
    
    TaskController.request_pause(execution_id)
    
    return TaskControlResponse(
        success=True,
        message="暂停请求已发送，任务将在当前阶段完成后暂停",
        execution_id=execution_id,
        current_status=context.status.value
    )


@router.post("/executions/{execution_id}/stop", response_model=TaskControlResponse)
async def stop_execution(execution_id: str) -> TaskControlResponse:
    """
    停止任务执行
    
    任务将在当前阶段完成后停止，已完成的阶段结果会被保留。
    
    Args:
        execution_id: 执行ID
        
    Returns:
        操作结果
    """
    if not TASK_MANAGER_AVAILABLE:
        raise HTTPException(status_code=501, detail="Task manager not available")
    
    context = ExecutionStore.get(execution_id)
    if not context:
        raise HTTPException(status_code=404, detail="Execution not found")
    
    if context.status.value not in ("running", "pending", "paused"):
        raise HTTPException(
            status_code=400,
            detail=f"Cannot stop task in status: {context.status.value}"
        )
    
    TaskController.request_stop(execution_id)
    
    return TaskControlResponse(
        success=True,
        message="停止请求已发送，任务将在当前阶段完成后停止",
        execution_id=execution_id,
        current_status=context.status.value
    )


@router.post("/executions/{execution_id}/resume", response_model=TaskControlResponse)
async def resume_execution(execution_id: str) -> TaskControlResponse:
    """
    恢复已暂停的任务
    
    支持跨重启恢复：如果ExecutionStore中没有context，会从持久化存储中加载。
    
    Args:
        execution_id: 执行ID
        
    Returns:
        操作结果
    """
    logger.info(f"[Resume] Resume request received for execution: {execution_id}")
    
    if not TASK_MANAGER_AVAILABLE:
        raise HTTPException(status_code=501, detail="Task manager not available")
    
    # Check if task is paused (via persistent record OR in-memory context)
    record = TaskHistoryService.get_record_by_execution_id(execution_id)
    record_status = record.get("status") if record else None
    
    # 也检查内存中的 context 状态（可能比数据库更新）
    mem_context = ExecutionStore.get(execution_id)
    mem_status = mem_context.status.value if mem_context else None
    
    is_paused = (record_status == "paused" or mem_status == "paused")
    
    if not record or not is_paused:
        logger.error(f"[Resume] Task is not paused, record_status={record_status}, mem_status={mem_status}")
        raise HTTPException(
            status_code=400,
            detail=f"只能恢复已暂停的任务（当前状态: {record_status or mem_status or '未知'}）"
        )
    
    logger.info(f"[Resume] Task is paused, record: {record.get('record_id')}")
    
    # 优先检查内存中的 context（Pipeline 线程可能仍在 _check_control 中阻塞等待 RESUME）
    context = ExecutionStore.get(execution_id)
    
    # 关键：不仅要检查 context 存在且 PAUSED，还要确认有活跃的 Pipeline 线程在等待信号
    # get_execution_status 会自动从持久化存储恢复 context 到 ExecutionStore（用于前端展示），
    # 但此时没有活跃线程在 _check_control 中阻塞等待 RESUME 信号。
    # 如果只靠 context.status == PAUSED 判断，会误走"发信号"分支导致 resume 无效。
    has_active_thread = _is_task_running(execution_id)
    
    if context and context.status == ExecutionStatus.PAUSED and has_active_thread:
        # Pipeline 线程仍在阻塞等待，直接发 RESUME 信号即可
        logger.info(f"[Resume] Found active in-memory context (PAUSED) with active thread, sending resume signal: {execution_id}")
        TaskController.request_resume(execution_id)
        
        # 等待 Pipeline 线程实际处理 RESUME 信号（最多 500ms）
        # 避免前端轮询时拉到旧的 paused 状态
        import time
        for _ in range(10):
            await asyncio.sleep(0.05)  # 50ms 间隔检查
            refreshed = ExecutionStore.get(execution_id)
            if refreshed and refreshed.status != ExecutionStatus.PAUSED:
                logger.info(f"[Resume] Status confirmed changed to: {refreshed.status.value}")
                break
        
        return TaskControlResponse(
            success=True,
            message="恢复请求已发送",
            execution_id=execution_id,
            current_status="running"
        )
    
    # 内存中没有 context 或 status 不是 PAUSED（如服务重启后），从持久化存储加载
    logger.info(f"[Resume] No active in-memory context, loading from persistent storage: {execution_id}")
    context = PersistentExecutionStore.load_full_state(execution_id)
    
    if not context:
        logger.error(f"[Resume] Failed to load context from persistent storage: {execution_id}")
        raise HTTPException(
            status_code=404, 
            detail=f"Execution not found in persistent storage: {execution_id}"
        )
    
    # 恢复 context 到 ExecutionStore
    ExecutionStore.update(context)
    logger.info(f"[Resume] Loaded and restored context: {execution_id}, status={context.status}, current_stage={context.current_stage}")
    
    # 设置恢复标志，让正在运行的执行器检测到并继续执行
    # 如果没有正在运行的执行器（比如重启后），则重新启动执行器
    logger.info(f"[Resume] Context status: {context.status}, current_stage: {context.current_stage}")
    
    if context.status != ExecutionStatus.RUNNING:
        # 没有正在运行的执行器，需要重新启动
        logger.info(f"[Resume] No running executor found, restarting task: {execution_id}")
        
        # 获取执行器并启动后台任务
        executor = get_executor()
        
        async def resume_task():
            logger.info(f"[Resume] Background task started for execution: {execution_id}")
            try:
                await executor.execute_async(
                    task_id=context.task_id,
                    session_id=context.session_id,
                    params=context.params,
                    data=context.data,
                    file_path=context.file_path,
                    execution_id=execution_id,  # 使用相同的execution_id
                    interaction_mode=context.interaction_mode,
                    model=context.model,
                    api_base=context.api_base,
                    system_prompt=context.system_prompt
                )
                logger.info(f"[Resume] Background task completed successfully: {execution_id}")
            except asyncio.CancelledError:
                logger.info(f"[Resume] Task {execution_id} was cancelled")
            except Exception as e:
                import traceback
                logger.error(f"[Resume] Background task execution failed: {e}\n{traceback.format_exc()}")
                ctx = ExecutionStore.get(execution_id)
                if ctx:
                    ctx.status = ExecutionStatus.FAILED
                    ctx.message = f"任务执行失败: {str(e)}"
                    ctx.errors.append(traceback.format_exc())
                    ExecutionStore.update(ctx)
            finally:
                # Phase 19: 任务完成后注销
                await _unregister_task(execution_id)
        
        # Phase 19: 取消已存在的任务（如果有），然后注册新任务
        await _cancel_existing_task(execution_id)
        logger.info(f"[Resume] Creating asyncio task for execution: {execution_id}")
        task = asyncio.create_task(resume_task())
        await _register_task(execution_id, task)
        logger.info(f"[Resume] Asyncio task created and registered: {execution_id}")
        
        return TaskControlResponse(
            success=True,
            message="任务已重新启动",
            execution_id=execution_id,
            current_status="running"
        )
    else:
        # 执行器正在运行，只是暂停了，发送恢复信号即可
        logger.info(f"[Resume] Executor is running, sending resume signal: {execution_id}")
        TaskController.request_resume(execution_id)
        
        return TaskControlResponse(
            success=True,
            message="恢复请求已发送",
            execution_id=execution_id,
            current_status="paused"
        )







# =============================================================================
# AI 分析辅助：SUGGESTED_PARAMS 解析
# =============================================================================

def _parse_suggested_params(analysis_text: str) -> Optional[Dict[str, Any]]:
    """
    从 AI 分析文本末尾解析 SUGGESTED_PARAMS: {...} 块。

    支持两种格式：
      1. 单行：SUGGESTED_PARAMS: {"k": v}
      2. 多行：SUGGESTED_PARAMS: {\n  "k": v\n}   （LLM 有时格式化 JSON）

    从文本末尾向上扫描，找到 SUGGESTED_PARAMS: 行后，
    收集从该行到文本末尾的所有内容尝试解析 JSON。
    """
    if not analysis_text:
        return None

    import json as _json

    lines = analysis_text.splitlines()
    # 去掉末尾空行
    while lines and not lines[-1].strip():
        lines.pop()

    # 从末尾向上找 SUGGESTED_PARAMS: 行
    marker_idx = None
    for i in range(len(lines) - 1, -1, -1):
        stripped = lines[i].strip()
        if stripped.startswith("SUGGESTED_PARAMS:"):
            marker_idx = i
            break
        # 只允许跳过空行和纯 JSON 行（} ] , 数字 字符串等）
        # 如果遇到普通文本行，说明没有 SUGGESTED_PARAMS
        if stripped and not stripped.startswith(("{", "}", "[", "]", '"', "'")) and not stripped[-1] in (",", "}"):
            break

    if marker_idx is None:
        return None

    # 提取从 SUGGESTED_PARAMS: 开始的内容
    first_line = lines[marker_idx].strip()
    json_start = first_line[len("SUGGESTED_PARAMS:"):].strip()

    # 拼接后续行（多行 JSON）
    remaining = "\n".join(lines[marker_idx + 1:]).strip()
    json_str = (json_start + "\n" + remaining).strip() if remaining else json_start

    try:
        params = _json.loads(json_str)
        if isinstance(params, dict) and params:
            return params
    except Exception:
        # 尝试只用第一行（可能后续行不是 JSON 的一部分）
        try:
            params = _json.loads(json_start)
            if isinstance(params, dict) and params:
                return params
        except Exception:
            pass
        logger.warning(f"[AI Analysis] Failed to parse SUGGESTED_PARAMS: {json_str!r}")
    return None


def _strip_suggested_params(analysis_text: str) -> str:
    """
    从 AI 分析文本中剥离 SUGGESTED_PARAMS: 块（含多行 JSON），返回干净的展示文本。
    """
    if not analysis_text:
        return analysis_text

    lines = analysis_text.splitlines()
    # 去掉末尾空行
    while lines and not lines[-1].strip():
        lines.pop()

    # 从末尾向上找 SUGGESTED_PARAMS: 行
    marker_idx = None
    for i in range(len(lines) - 1, -1, -1):
        stripped = lines[i].strip()
        if stripped.startswith("SUGGESTED_PARAMS:"):
            marker_idx = i
            break
        # 只允许跳过空行和纯 JSON 结构行
        if stripped and not stripped.startswith(("{", "}", "[", "]", '"', "'")) and not stripped[-1] in (",", "}"):
            break

    if marker_idx is None:
        return "\n".join(lines)

    # 截断到 marker 之前，再去末尾空行
    lines = lines[:marker_idx]
    while lines and not lines[-1].strip():
        lines.pop()
    return "\n".join(lines)





# =============================================================================
# Phase 6: 阶段重试和恢复 API
# =============================================================================



class StageRetryRequest(BaseModel):
    """阶段重试请求"""
    new_params: Optional[Dict[str, Any]] = Field(None, description="新参数（可选，覆盖原参数）")
    retry_reason: Optional[str] = Field(None, description="重试原因，如'接受AI建议'（可选，用于版本历史展示）")


class StageRetryResponse(BaseModel):
    """阶段重试响应"""
    success: bool
    message: str
    execution_id: str
    retry_stage_id: str
    previous_stage_id: Optional[str] = None
    needs_restart: bool = False  # 是否需要重新启动任务（Pipeline模式下为True）


class RecoverableExecutionItem(BaseModel):
    """可恢复的执行项"""
    execution_id: str
    task_id: str
    session_id: Optional[str]
    record_id: Optional[str]
    status: str
    pause_stage_id: Optional[str]
    current_stage_id: Optional[str]
    completed_stages_count: int
    paused_at: Optional[str]
    created_at: Optional[str]


class RecoverableExecutionsResponse(BaseModel):
    """可恢复执行列表响应"""
    executions: List[RecoverableExecutionItem]


class StageCheckpointItem(BaseModel):
    """阶段检查点项"""
    stage_id: str
    stage_index: int
    stage_status: str
    outputs_summary: Optional[Dict[str, Any]] = None
    params: Optional[Dict[str, Any]] = None
    started_at: Optional[str]
    completed_at: Optional[str]


class ExecutionCheckpointsResponse(BaseModel):
    """执行检查点列表响应"""
    execution_id: str
    checkpoints: List[StageCheckpointItem]


@router.get("/executions/recoverable", response_model=RecoverableExecutionsResponse)
async def list_recoverable_executions(
    session_id: Optional[str] = Query(None, description="按会话筛选")
) -> RecoverableExecutionsResponse:
    """
    列出可恢复的任务
    
    返回所有暂停中的任务，支持跨后端重启恢复。
    """
    # P1-D5: 入口层 session_id 校验（可选参数，非空时校验）
    if session_id:
        try:
            session_id = _validate_session_id(session_id)
        except ValueError as e:
            raise HTTPException(status_code=400, detail=str(e))
    
    if not TASK_MANAGER_AVAILABLE or ExecutionStateRecovery is None:
        raise HTTPException(status_code=501, detail="Task manager not available")
    
    try:
        recoverable = ExecutionStateRecovery.list_recoverable(session_id)
        
        items = [
            RecoverableExecutionItem(
                execution_id=r["execution_id"],
                task_id=r["task_id"],
                session_id=r.get("session_id"),
                record_id=r.get("record_id"),
                status=r["status"],
                pause_stage_id=r.get("pause_stage_id"),
                current_stage_id=r.get("current_stage_id"),
                completed_stages_count=r.get("completed_stages_count", 0),
                paused_at=r.get("paused_at"),
                created_at=r.get("created_at"),
            )
            for r in recoverable
        ]
        
        return RecoverableExecutionsResponse(executions=items)
    except Exception as e:
        logger.error(f"Failed to list recoverable executions: {e}")
        raise HTTPException(status_code=500, detail="Failed to list recoverable executions")


@router.get("/executions/{execution_id}/checkpoints", response_model=ExecutionCheckpointsResponse)
async def get_execution_checkpoints(execution_id: str) -> ExecutionCheckpointsResponse:
    """
    获取执行的所有检查点
    
    用于查看已完成阶段的信息，支持阶段重试。
    """
    if not TASK_MANAGER_AVAILABLE or PersistentExecutionStore is None:
        raise HTTPException(status_code=501, detail="Task manager not available")
    
    try:
        checkpoints = PersistentExecutionStore.get_checkpoints(execution_id)
        
        items = [
            StageCheckpointItem(
                stage_id=cp["stage_id"],
                stage_index=cp["stage_index"],
                stage_status=cp["stage_status"],
                outputs_summary=cp.get("outputs_summary"),
                params=cp.get("params"),
                started_at=cp.get("started_at"),
                completed_at=cp.get("completed_at"),
            )
            for cp in checkpoints
        ]
        
        return ExecutionCheckpointsResponse(
            execution_id=execution_id,
            checkpoints=items
        )
    except Exception as e:
        logger.error(f"Failed to get checkpoints: {e}")
        raise HTTPException(status_code=500, detail="Failed to get checkpoints")


@router.get("/executions/{execution_id}/recovery-info")
async def get_recovery_info(execution_id: str) -> Dict[str, Any]:
    """
    获取任务恢复信息
    
    检查任务是否可以恢复，返回恢复所需的信息。
    """
    if not TASK_MANAGER_AVAILABLE or ExecutionStateRecovery is None:
        raise HTTPException(status_code=501, detail="Task manager not available")
    
    try:
        check_result = ExecutionStateRecovery.can_resume(execution_id)
        
        if not check_result["can_resume"]:
            return {
                "can_resume": False,
                "reason": check_result["reason"],
                "execution_id": execution_id,
            }
        
        return {
            "can_resume": True,
            "execution_id": execution_id,
            "resume_stage_id": check_result["resume_stage_id"],
            "completed_stages": check_result["completed_stages"],
        }
    except Exception as e:
        logger.error(f"Failed to get recovery info: {e}")
        raise HTTPException(status_code=500, detail="Failed to get recovery info")


@router.post("/executions/{execution_id}/recover")
async def recover_execution(execution_id: str) -> Dict[str, Any]:
    """
    恢复暂停的任务执行（Phase 6）
    
    从暂停点恢复任务执行，使用已保存的检查点数据。
    支持跨后端重启后的任务恢复。
    
    Args:
        execution_id: 执行ID
        
    Returns:
        恢复结果，包含恢复的阶段ID等信息
    """
    if not TASK_MANAGER_AVAILABLE or ExecutionStateRecovery is None:
        raise HTTPException(status_code=501, detail="Task manager not available")
    
    try:
        # 检查是否可以恢复
        check_result = ExecutionStateRecovery.can_resume(execution_id)
        if not check_result["can_resume"]:
            raise HTTPException(status_code=400, detail=check_result["reason"])
        
        # 获取恢复上下文
        resume_context = ExecutionStateRecovery.get_resume_context(execution_id)
        if not resume_context:
            raise HTTPException(status_code=400, detail="无法获取恢复上下文")
        
        # 检查内存中是否有执行上下文
        context = ExecutionStore.get(execution_id)
        
        if not context:
            # 尝试从持久化存储恢复完整状态
            if PersistentExecutionStore is not None:
                full_state = PersistentExecutionStore.load_full_state(execution_id)
                if full_state:
                    # 将恢复的上下文添加到内存存储
                    ExecutionStore._executions[execution_id] = full_state
                    context = full_state
                    logger.info(f"Restored execution context from persistent storage: {execution_id}")
        
        if not context:
            raise HTTPException(
                status_code=400, 
                detail="无法恢复执行上下文，请重新创建任务"
            )
        
        # 设置恢复阶段
        resume_stage_id = check_result["resume_stage_id"]
        context.retry_from_stage = resume_stage_id
        
        # 更新状态
        context.status = ExecutionStatus.PAUSED  # 保持暂停状态，等待 resume 信号
        ExecutionStore.update(context)
        
        # 发送恢复信号
        TaskController.request_resume(execution_id)
        
        logger.info(f"Recovered execution {execution_id}, resuming from stage: {resume_stage_id}")
        
        return {
            "success": True,
            "execution_id": execution_id,
            "resume_stage_id": resume_stage_id,
            "completed_stages": check_result["completed_stages"],
            "message": f"任务已恢复，从阶段 {resume_stage_id} 继续执行"
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Failed to recover execution: {e}")
        raise HTTPException(status_code=500, detail="Failed to recover execution")


@router.post("/executions/{execution_id}/stages/{stage_id}/retry", response_model=StageRetryResponse)
async def retry_stage(
    execution_id: str,
    stage_id: str,
    request: Optional[StageRetryRequest] = None
) -> StageRetryResponse:
    """
    重试指定阶段
    
    设置重试标记并恢复任务执行。Pipeline 会从该阶段重新开始执行，
    之前的阶段会快速执行（不触发专家模式暂停），重试阶段及之后的阶段会正常执行。
    
    支持跨重启重试：如果ExecutionStore中没有context，会从持久化存储中加载。
    
    Args:
        execution_id: 执行ID
        stage_id: 要重试的阶段ID
        request: 可选的新参数
    """
    if not TASK_MANAGER_AVAILABLE:
        raise HTTPException(status_code=501, detail="Task manager not available")
    
    # Try to get context from ExecutionStore (memory cache)
    context = ExecutionStore.get(execution_id)
    
    # If not found in memory, try to load from persistent storage
    if not context:
        logger.info(f"[Retry Stage] Context not found in ExecutionStore, loading from persistent storage: {execution_id}")
        context = PersistentExecutionStore.load_full_state(execution_id)
        
        if not context:
            raise HTTPException(
                status_code=404, 
                detail=f"Execution not found in persistent storage: {execution_id}"
            )
        
        # Restore context to ExecutionStore for subsequent operations
        ExecutionStore.update(context)
        logger.info(f"[Retry Stage] Restored context to ExecutionStore: {execution_id}")
    
    # 检查任务是否处于暂停状态（内存 context 或数据库记录，任一为 paused 即可）
    record = TaskHistoryService.get_record_by_execution_id(execution_id)
    record_status = record.get("status") if record else None
    mem_status = context.status.value if context else None
    is_paused = (mem_status == "paused" or record_status == "paused")
    
    if not is_paused:
        raise HTTPException(
            status_code=400, 
            detail=f"任务不是暂停状态，无法重试（内存状态: {mem_status}, 数据库状态: {record_status}）"
        )
    
    try:
        # 如果提供了新参数，更新上下文
        new_params = request.new_params if request else None
        retry_reason = request.retry_reason if request else None
        if new_params:
            context.params.update(new_params)
            logger.info(f"[Retry Stage] Updated context.params for {execution_id}: {new_params}")

        # ── 版本快照：重置前将当前阶段状态存档 ──────────────────────────
        target_stage = context.stages.get(stage_id)
        if (
            target_stage is not None
            and target_stage.status == ExecutionStatus.COMPLETED
            and target_stage.output_preview is not None
        ):
            # 从 DB 读取该阶段的 AI 分析文本（异步安全地同步读取）
            ai_analysis_text: str | None = None
            try:
                record_id_for_snap = getattr(context, 'record_id', None)
                if record_id_for_snap and StageAnalysisService:
                    analysis_rec = StageAnalysisService.get_analysis(record_id_for_snap, stage_id)
                    ai_analysis_text = analysis_rec.get('analysis_text') if analysis_rec else None
            except Exception as e:
                logger.warning(f"[Retry Stage] Could not load AI analysis for snapshot: {e}")

            # 计算版本号 = 已有快照数 + 1
            snapshot = StageSnapshot(
                version=len(target_stage.snapshots) + 1,
                params_used=dict(target_stage.params_used),
                output_preview=dict(target_stage.output_preview),
                ai_analysis=ai_analysis_text,
                suggested_params=None,   # 前端解析后如需保存，可通过单独接口写入
                execution_time_ms=target_stage.execution_time_ms,
                completed_at=target_stage.completed_at.isoformat() if target_stage.completed_at else None,
                retry_reason=retry_reason,
            )
            # FIFO：保留最近 10 个快照
            target_stage.snapshots.append(snapshot)
            if len(target_stage.snapshots) > 10:
                target_stage.snapshots = target_stage.snapshots[-10:]
            logger.info(
                f"[Retry Stage] Created snapshot v{snapshot.version} for {execution_id}/{stage_id}"
                f"  reason={retry_reason}"
            )

            # 快照已保存旧分析，删除 DB 中的 AI 分析记录
            # 这样前端重试完成后 fetchAnalysis 返回 null，触发重新生成新版本分析
            if record_id_for_snap and StageAnalysisService:
                try:
                    StageAnalysisService.delete_analysis(record_id_for_snap, stage_id)
                    logger.info(f"[Retry Stage] Deleted AI analysis for {record_id_for_snap}/{stage_id} (saved in snapshot v{snapshot.version})")
                except Exception as e:
                    logger.warning(f"[Retry Stage] Failed to delete AI analysis: {e}")
        # ──────────────────────────────────────────────────────────────────
        
        # 设置重试阶段标记
        context.retry_from_stage = stage_id
        logger.info(f"Set retry_from_stage for {execution_id}: {stage_id}")
        
        # 重置该阶段及后续阶段的状态
        # 2026-02-10: 修复逻辑 - 只对已执行过的阶段显示"阶段已重置"
        # - 从未执行过的阶段（status=PENDING）不应显示重置信息
        task_def = get_registry().get_task(context.task_id)
        if task_def:
            stage_ids = [s.id for s in task_def.stages]
            if stage_id in stage_ids:
                start_idx = stage_ids.index(stage_id)
                for sid in stage_ids[start_idx:]:
                    if sid in context.stages:
                        stage = context.stages[sid]
                        old_params_used = dict(stage.params_used) if stage.params_used else {}
                        
                        # 判断阶段是否曾经执行过（非PENDING状态或有output_preview）
                        was_executed = (
                            stage.status != ExecutionStatus.PENDING or
                            stage.output_preview is not None or
                            stage.started_at is not None
                        )
                        
                        # 重置阶段状态
                        stage.status = ExecutionStatus.PENDING
                        stage.progress = 0.0
                        stage.started_at = None
                        stage.completed_at = None
                        stage.output_preview = None
                        # 清除 params_used，让阶段重新执行时从 context.params 获取最新参数
                        stage.params_used = {}
                        # 后续阶段（非重试阶段本身）清空快照历史：上游重试后旧快照已无效
                        if sid != stage_id and hasattr(stage, 'snapshots'):
                            stage.snapshots = []
                        # 后续阶段（非重试阶段本身）删除 DB 中的 AI 分析：避免旧分析被展示
                        if sid != stage_id and was_executed:
                            record_id_for_downstream = getattr(context, 'record_id', None)
                            if record_id_for_downstream and StageAnalysisService:
                                try:
                                    StageAnalysisService.delete_analysis(record_id_for_downstream, sid)
                                    logger.info(f"[Retry Stage] Deleted downstream AI analysis for {record_id_for_downstream}/{sid}")
                                except Exception as e:
                                    logger.warning(f"[Retry Stage] Failed to delete downstream AI analysis for {sid}: {e}")
                        
                        # 只有曾经执行过的阶段才记录"已重置"日志
                        if was_executed:
                            stage.logs.append(f"[{datetime.now().strftime('%H:%M:%S')}] 阶段已重置（准备重试）")
                            logger.info(f"[Retry Stage] Reset stage {sid} (was executed): cleared params_used (was: {list(old_params_used.keys())})")
                        else:
                            logger.info(f"[Retry Stage] Stage {sid} was never executed, no reset log added")
        
        ExecutionStore.update(context)
        
        # Phase 17: 保存到持久化存储，确保 retry_from_stage 在跨重启时也能正确恢复
        # 这是必要的，因为 resume_execution 会从持久化存储加载 context
        PersistentExecutionStore.save_full_state(execution_id, context)
        logger.info(f"Saved retry_from_stage to persistent storage: {execution_id}")
        
        # 2026-02-10: 清理 AI 分析缓存
        # 阶段重试后需要重新生成 AI 评估，而不是使用旧的缓存
        # 使用 invalidate_checkpoints_from 清理从重试阶段开始的所有 AI 分析结果
        if TASK_MANAGER_AVAILABLE and PersistentExecutionStore:
            try:
                invalidated_count = PersistentExecutionStore.invalidate_checkpoints_from(
                    execution_id, stage_id
                )
                logger.info(f"[Retry Stage] Invalidated {invalidated_count} checkpoints and AI analysis from {execution_id}/{stage_id}")
            except Exception as e:
                logger.warning(f"[Retry Stage] Failed to invalidate checkpoints: {e}")
        
        # Phase 19: 终止旧线程并启动新任务
        # 递增 execution_version，旧线程在 progress_callback/stop_check_callback 中
        # 检测到版本不匹配后自动抛出 TaskStoppedException 退出。
        # 这比 STOP 信号更可靠，因为版本检查在 Pipeline 的每个 progress_callback 调用点都会执行，
        # 不依赖异常穿透 try/except 层级。
        context._execution_version += 1
        context._aborted_for_retry = True
        logger.info(f"[Retry Stage] Incremented execution_version to {context._execution_version}")
        
        if _is_task_running(execution_id):
            logger.info(f"[Retry Stage] Found running task for {execution_id}, sending STOP + version bump to terminate old thread")
            TaskController.request_stop(execution_id)
            # 等待旧线程退出（最多等3秒）
            for _ in range(30):
                await asyncio.sleep(0.1)
                if not _is_task_running(execution_id):
                    break
            if _is_task_running(execution_id):
                logger.warning(f"[Retry Stage] Old thread didn't exit after 3s, force cancelling")
                await _cancel_existing_task(execution_id)
            else:
                logger.info(f"[Retry Stage] Old thread exited cleanly")
            TaskController.clear_control(execution_id)
            context._aborted_for_retry = False
        else:
            logger.info(f"[Retry Stage] No running task for {execution_id}")
            context._aborted_for_retry = False
        
        # 重置 context 状态为 RUNNING
        context.status = ExecutionStatus.RUNNING
        context.message = "阶段重试中..."
        ExecutionStore.update(context)
        
        # 启动新任务执行重试
        executor = get_executor()
        
        async def retry_task():
            logger.info(f"[Retry Stage] Background task started for execution: {execution_id}")
            try:
                await executor.execute_async(
                    task_id=context.task_id,
                    session_id=context.session_id,
                    params=context.params,
                    data=context.data,
                    file_path=context.file_path,
                    execution_id=execution_id,
                    interaction_mode=context.interaction_mode,
                    model=context.model,
                    api_base=context.api_base,
                    system_prompt=context.system_prompt
                )
                logger.info(f"[Retry Stage] Background task completed: {execution_id}")
            except asyncio.CancelledError:
                logger.info(f"[Retry Stage] Task {execution_id} was cancelled")
            except Exception as e:
                import traceback
                logger.error(f"[Retry Stage] Background task failed: {e}\n{traceback.format_exc()}")
                ctx = ExecutionStore.get(execution_id)
                if ctx:
                    ctx.status = ExecutionStatus.FAILED
                    ctx.message = f"任务执行失败: {str(e)}"
                    ctx.errors.append(traceback.format_exc())
                    ExecutionStore.update(ctx)
            finally:
                await _unregister_task(execution_id)
        
        task = asyncio.create_task(retry_task())
        await _register_task(execution_id, task)
        
        return StageRetryResponse(
            success=True,
            message=f"阶段 {stage_id} 正在重试，任务已恢复执行",
            execution_id=execution_id,
            retry_stage_id=stage_id,
            previous_stage_id=None,
            needs_restart=False
        )
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Failed to retry stage: {e}")
        raise HTTPException(status_code=500, detail="Failed to retry stage")


@router.post("/executions/{execution_id}/stages/{stage_id}/reset", response_model=TaskControlResponse)
async def reset_stage(execution_id: str, stage_id: str) -> TaskControlResponse:
    """
    重置指定阶段状态为 pending
    
    仅重置状态，不触发执行。用于手动控制阶段状态。
    支持跨重启重置：如果ExecutionStore中没有context，会从持久化存储中加载。
    """
    if not TASK_MANAGER_AVAILABLE or PersistentExecutionStore is None:
        raise HTTPException(status_code=501, detail="Task manager not available")
    
    # Try to get context from ExecutionStore (memory cache)
    context = ExecutionStore.get(execution_id)
    
    # If not found in memory, try to load from persistent storage
    if not context:
        logger.info(f"[Reset Stage] Context not found in ExecutionStore, loading from persistent storage: {execution_id}")
        context = PersistentExecutionStore.load_full_state(execution_id)
        
        if not context:
            raise HTTPException(
                status_code=404, 
                detail=f"Execution not found in persistent storage: {execution_id}"
            )
        
        # Restore context to ExecutionStore for subsequent operations
        ExecutionStore.update(context)
        logger.info(f"[Reset Stage] Restored context to ExecutionStore: {execution_id}")
    
    if stage_id not in context.stages:
        raise HTTPException(status_code=404, detail=f"Stage not found: {stage_id}")
    
    try:
        # 重置检查点
        PersistentExecutionStore.reset_checkpoint(execution_id, stage_id)
        
        # 重置内存中的阶段状态
        stage = context.stages[stage_id]
        stage.status = ExecutionStatus.PENDING
        stage.progress = 0.0
        stage.started_at = None
        stage.completed_at = None
        stage.output_preview = None
        stage.logs.append(f"[{datetime.now().strftime('%H:%M:%S')}] 阶段已手动重置")
        
        ExecutionStore.update(context)
        
        return TaskControlResponse(
            success=True,
            message=f"阶段 {stage_id} 已重置",
            execution_id=execution_id,
            current_status=context.status.value
        )
    except Exception as e:
        logger.error(f"Failed to reset stage: {e}")
        raise HTTPException(status_code=500, detail="Failed to reset stage")


# =============================================================================
# Task History Endpoints
# =============================================================================

class TaskHistoryItem(BaseModel):
    """任务历史记录项"""
    record_id: str
    task_type: str
    task_category: str
    execution_id: Optional[str]
    session_id: Optional[str]
    interaction_mode: str
    status: str
    progress: float
    current_stage: Optional[str]
    message: Optional[str]
    created_at: Optional[str]
    started_at: Optional[str]
    completed_at: Optional[str]
    duration_seconds: Optional[float]


class TaskHistoryListResponse(BaseModel):
    """任务历史列表响应"""
    total: int
    limit: int
    offset: int
    records: List[TaskHistoryItem]


class TaskHistoryDetailResponse(BaseModel):
    """任务历史详情响应"""
    record_id: str
    task_type: str
    task_category: str
    execution_id: Optional[str]
    session_id: Optional[str]
    interaction_mode: str
    status: str
    progress: float
    current_stage: Optional[str]
    message: Optional[str]
    params: Optional[Dict[str, Any]]
    inputs_summary: Optional[Dict[str, Any]]
    outputs_summary: Optional[Dict[str, Any]]
    stages: Optional[Dict[str, Any]]
    error_message: Optional[str]
    created_at: Optional[str]
    started_at: Optional[str]
    completed_at: Optional[str]
    duration_seconds: Optional[float]


class TaskStatisticsResponse(BaseModel):
    """任务统计响应"""
    total: int
    completed: int
    failed: int
    stopped: int
    running: int
    paused: int
    success_rate: float
    avg_duration_seconds: float
    period_days: int


@router.get("/history", response_model=TaskHistoryListResponse)
async def list_task_history(
    task_type: Optional[str] = Query(None, description="任务类型筛选"),
    task_category: Optional[str] = Query(None, description="任务类别筛选"),
    session_id: Optional[str] = Query(None, description="会话ID筛选"),
    status: Optional[str] = Query(None, description="状态筛选"),
    start_date: Optional[str] = Query(None, description="开始日期（ISO格式）"),
    end_date: Optional[str] = Query(None, description="结束日期（ISO格式）"),
    limit: int = Query(50, ge=1, le=200, description="每页数量"),
    offset: int = Query(0, ge=0, description="偏移量")
) -> TaskHistoryListResponse:
    """
    查询任务历史记录列表
    
    支持多条件筛选和分页。
    """
    # P1-D5: 入口层 session_id 校验（可选参数，非空时校验）
    if session_id:
        try:
            session_id = _validate_session_id(session_id)
        except ValueError as e:
            raise HTTPException(status_code=400, detail=str(e))
    
    if not TASK_MANAGER_AVAILABLE:
        raise HTTPException(status_code=501, detail="Task manager not available")
    
    # Parse dates
    start_dt = datetime.fromisoformat(start_date) if start_date else None
    end_dt = datetime.fromisoformat(end_date) if end_date else None
    
    records = TaskHistoryService.list_records(
        task_type=task_type,
        task_category=task_category,
        session_id=session_id,
        status=status,
        start_date=start_dt,
        end_date=end_dt,
        limit=limit,
        offset=offset
    )
    
    total = TaskHistoryService.count_records(
        task_type=task_type,
        task_category=task_category,
        session_id=session_id,
        status=status,
        start_date=start_dt,
        end_date=end_dt
    )
    
    return TaskHistoryListResponse(
        total=total,
        limit=limit,
        offset=offset,
        records=[TaskHistoryItem(**r) for r in records]
    )


@router.get("/history/{record_id}", response_model=TaskHistoryDetailResponse)
async def get_task_history_detail(record_id: str) -> TaskHistoryDetailResponse:
    """
    获取任务历史记录详情
    
    Args:
        record_id: 记录ID
        
    Returns:
        记录详情
    """
    if not TASK_MANAGER_AVAILABLE:
        raise HTTPException(status_code=501, detail="Task manager not available")
    
    record = TaskHistoryService.get_record(record_id)
    if not record:
        raise HTTPException(status_code=404, detail=f"Record not found: {record_id}")
    
    # 从任务定义中获取参数元数据，填充到stages中
    # 这样历史任务恢复时也能显示参数配置面板（而不是JSON编辑器）
    task_type = record.get("task_type")
    stages = record.get("stages")
    if task_type and stages:
        try:
            from deepanalyze.analysis.task_SOP.registry import get_registry
            registry = get_registry()
            task_def = registry.get_task(task_type)
            
            if task_def:
                # 构建阶段参数映射（从任务定义中获取每个阶段的参数）
                stage_params_meta: Dict[str, List[Dict[str, Any]]] = {}
                all_params = list(task_def.required_params) + list(task_def.optional_params)
                for param in all_params:
                    if param.stage:
                        if param.stage not in stage_params_meta:
                            stage_params_meta[param.stage] = []
                        stage_params_meta[param.stage].append({
                            "name": param.name,
                            "label": param.label,
                            "type": param.param_type.value,
                            "default": param.default,
                            "description": param.description,
                            "options": param.options,
                            "validation": param.validation,
                            "required": param.required,
                            "show_when": param.show_when,
                        })
                
                # 填充每个阶段的params_meta（如果为空）
                for stage_id, stage_data in stages.items():
                    if isinstance(stage_data, dict):
                        # 如果stages中没有params_meta或为空，从任务定义中获取
                        if not stage_data.get("params_meta"):
                            stage_data["params_meta"] = stage_params_meta.get(stage_id, [])
                
                record["stages"] = stages
        except Exception as e:
            logger.warning(f"Failed to fill params_meta from task definition: {e}")
    
    return TaskHistoryDetailResponse(**record)


def _serialize_task_result(result: Dict[str, Any], record_id: str) -> Dict[str, Any]:
    """
    序列化任务结果（将被缓存）
    
    Args:
        result: 原始结果字典（含DataFrame等）
        record_id: 记录ID（用于日志）
        
    Returns:
        序列化后的结果字典
    """
    def safe_serialize(value: Any, key: str = "") -> Dict[str, Any]:
        """安全地序列化值，处理各种类型"""
        try:
            if isinstance(value, pd.DataFrame):
                return {
                    "type": "dataframe",
                    "columns": list(value.columns),
                    "shape": list(value.shape),
                    "data": value.fillna("").head(100).to_dict(orient="records")
                }
            elif isinstance(value, dict):
                # 非空字典且所有值都是 DataFrame
                if value and all(isinstance(v, pd.DataFrame) for v in value.values()):
                    return {
                        "type": "dict_of_dataframes",
                        "data": {
                            str(k): {
                                "columns": list(v.columns),
                                "data": v.fillna("").to_dict(orient="records")
                            }
                            for k, v in value.items()
                        }
                    }
                else:
                    # 尝试递归序列化字典内容
                    try:
                        import json
                        json.dumps(value)  # 测试是否可序列化
                        return {"type": "dict", "data": value}
                    except (TypeError, ValueError):
                        # 不可直接序列化，递归处理
                        return {
                            "type": "dict",
                            "data": {str(k): safe_serialize(v, f"{key}.{k}") for k, v in value.items()}
                        }
            elif isinstance(value, list):
                try:
                    import json
                    json.dumps(value)  # 测试是否可序列化
                    return {"type": "list", "data": value}
                except (TypeError, ValueError):
                    return {"type": "list", "data": [safe_serialize(item, f"{key}[{i}]") for i, item in enumerate(value)]}
            elif hasattr(value, 'item'):  # numpy 类型
                return {"type": "number", "data": value.item()}
            elif hasattr(value, 'tolist'):  # numpy array
                return {"type": "list", "data": value.tolist()}
            else:
                return {"type": "other", "data": str(value)}
        except Exception as e:
            logger.warning(f"Failed to serialize key '{key}': {e}")
            return {"type": "error", "data": f"Serialization failed: {str(e)}"}
    
    serialized_result = {}
    for key, value in result.items():
        serialized_result[key] = safe_serialize(value, key)
    
    return serialized_result


@router.get("/history/{record_id}/result")
async def get_task_history_result(record_id: str) -> Dict[str, Any]:
    """
    获取历史任务的完整执行结果
    
    从文件存储中加载完整结果数据。
    使用LRU缓存避免重复的DataFrame序列化操作。
    
    Args:
        record_id: 记录ID
        
    Returns:
        完整结果数据
    """
    if not TASK_MANAGER_AVAILABLE:
        raise HTTPException(status_code=501, detail="Task manager not available")
    
    record = TaskHistoryService.get_record(record_id)
    if not record:
        raise HTTPException(status_code=404, detail=f"Record not found: {record_id}")
    
    storage = get_result_storage()
    result = storage.load_result(record_id)
    
    if not result:
        raise HTTPException(status_code=404, detail="Result file not found")
    
    # 获取结果文件的修改时间，用于缓存失效判断
    result_dir = Path(storage.base_dir) / record_id
    pickle_path = result_dir / "outputs.pkl"
    
    if pickle_path.exists():
        file_mtime = pickle_path.stat().st_mtime
    else:
        # 如果没有pickle文件，使用result.json的修改时间
        json_path = result_dir / "result.json"
        file_mtime = json_path.stat().st_mtime if json_path.exists() else time.time()
    
    # 检查缓存
    cached_result = _get_task_result_from_cache(record_id, file_mtime)
    
    if cached_result is not None:
        logger.info(f"[TaskResultCache] Hit for {record_id}")
        serialized_result = cached_result
    else:
        logger.info(f"[TaskResultCache] Miss for {record_id}, serializing...")
        start_time = time.time()
        
        # 执行序列化
        serialized_result = _serialize_task_result(result, record_id)
        
        # 存入缓存
        _set_task_result_cache(record_id, file_mtime, serialized_result)
        
        elapsed_ms = (time.time() - start_time) * 1000
        logger.info(f"[TaskResultCache] Serialized {record_id} in {elapsed_ms:.1f}ms")
    
    # 同时返回stages数据（用于"样本及特征"Tab）
    stages = record.get("stages", {}) if record else {}
    
    return {
        "record_id": record_id,
        "result": serialized_result,
        "stages": stages  # 一次性返回stages数据，避免前端需要额外调用getTaskHistoryDetail
    }


@router.get("/cache/stats")
async def get_cache_statistics() -> Dict[str, Any]:
    """
    获取任务结果缓存统计信息
    
    用于监控缓存命中率和性能优化。
    
    Returns:
        缓存统计信息
    """
    return {
        "task_result_cache": _get_task_result_cache_stats(),
        "preview_cache": {
            "size": len(_preview_cache),
            "max_size": _CACHE_MAX_SIZE
        }
    }


@router.post("/cache/clear")
async def clear_all_caches() -> Dict[str, Any]:
    """
    清除所有缓存
    
    包括任务结果缓存和文件预览缓存。
    
    Returns:
        清除结果
    """
    _clear_task_result_cache()
    _preview_cache.clear()
    
    return {
        "success": True,
        "message": "All caches cleared"
    }


class BatchDeleteRequest(BaseModel):
    """批量删除请求（Phase 25）"""
    record_ids: List[str] = Field(..., description="要删除的记录ID列表", min_length=1, max_length=100)
    cleanup_files: bool = Field(True, description="是否同时清理关联文件")


@router.post("/history/batch-delete")
async def batch_delete_task_history(request: BatchDeleteRequest) -> Dict[str, Any]:
    """
    批量删除历史记录（Phase 25）
    
    同时删除数据库记录、AI分析、执行状态和结果文件。
    运行中/待执行的任务会被跳过。
    
    Args:
        request: 包含 record_ids 列表和 cleanup_files 标志
        
    Returns:
        删除统计（deleted/failed/skipped_running）
    """
    if not TASK_MANAGER_AVAILABLE:
        raise HTTPException(status_code=501, detail="Task manager not available")
    
    # 批量删除数据库记录 + 执行状态
    result = TaskHistoryService.batch_delete_records(
        record_ids=request.record_ids,
        cleanup_files=request.cleanup_files
    )
    
    # 批量删除结果文件
    if request.cleanup_files:
        storage = get_result_storage()
        for record_id in request.record_ids:
            if record_id not in result.get("skipped_running", []) and record_id not in result.get("failed_ids", []):
                try:
                    storage.delete_result(record_id)
                except Exception as e:
                    logger.warning(f"Failed to delete result files for {record_id}: {e}")
    
    return {
        "success": True,
        "deleted": result["deleted"],
        "failed": result["failed"],
        "skipped_running": result.get("skipped_running", []),
        "cleanup_files": request.cleanup_files,
        "message": f"已删除 {result['deleted']} 条记录"
            + (f"，{len(result.get('skipped_running', []))} 条运行中已跳过" if result.get("skipped_running") else "")
            + (f"，{result['failed']} 条失败" if result["failed"] > 0 else "")
    }


@router.delete("/history/{record_id}")
async def delete_task_history(record_id: str) -> Dict[str, Any]:
    """
    删除历史记录
    
    同时删除数据库记录和结果文件。
    
    Args:
        record_id: 记录ID
        
    Returns:
        操作结果
    """
    if not TASK_MANAGER_AVAILABLE:
        raise HTTPException(status_code=501, detail="Task manager not available")
    
    # Delete database record
    success = TaskHistoryService.delete_record(record_id)
    if not success:
        raise HTTPException(status_code=404, detail=f"Record not found: {record_id}")
    
    # Delete result files
    storage = get_result_storage()
    storage.delete_result(record_id)
    
    return {
        "success": True,
        "message": "记录已删除",
        "record_id": record_id
    }


# =============================================================================
# Stage AI Analysis Endpoints (Phase 7)
# =============================================================================

class StageAnalysisRequest(BaseModel):
    """阶段 AI 分析保存请求"""
    analysis_text: str = Field(..., description="AI 分析文本")
    model_used: Optional[str] = Field(None, description="使用的模型名称")


class StageAnalysisResponse(BaseModel):
    """阶段 AI 分析响应"""
    record_id: str
    stage_id: str
    analysis_text: str
    model_used: Optional[str] = None
    created_at: Optional[str] = None
    updated_at: Optional[str] = None


@router.get("/history/{record_id}/stages/{stage_id}/analysis")
async def get_stage_analysis(
    record_id: str,
    stage_id: str
) -> Dict[str, Any]:
    """
    获取阶段 AI 分析结果
    
    从数据库读取缓存的 AI 分析文本。如果不存在，返回 null。
    
    Args:
        record_id: 任务记录ID
        stage_id: 阶段ID
        
    Returns:
        分析结果或 null
    """
    if not TASK_MANAGER_AVAILABLE or StageAnalysisService is None:
        raise HTTPException(status_code=501, detail="Task manager not available")
    
    analysis = StageAnalysisService.get_analysis(record_id, stage_id)
    
    # 从分析文本中提取 suggested_params，并确保返回给前端的文本是干净的（无 SUGGESTED_PARAMS 行）
    if analysis and analysis.get("analysis_text"):
        raw_text = analysis["analysis_text"]
        analysis["suggested_params"] = _parse_suggested_params(raw_text)
        # 防御性剥离：若 DB 中存储了含标记行的旧数据，在返回时一并剥离
        analysis["analysis_text"] = _strip_suggested_params(raw_text)
    
    return {
        "record_id": record_id,
        "stage_id": stage_id,
        "analysis": analysis  # 可能为 None
    }


@router.post("/history/{record_id}/stages/{stage_id}/analysis")
async def save_stage_analysis(
    record_id: str,
    stage_id: str,
    request: StageAnalysisRequest
) -> Dict[str, Any]:
    """
    保存阶段 AI 分析结果
    
    将 AI 生成的分析文本保存到数据库。如果已存在则覆盖。
    
    Args:
        record_id: 任务记录ID
        stage_id: 阶段ID
        request: 分析内容
        
    Returns:
        操作结果
    """
    if not TASK_MANAGER_AVAILABLE or StageAnalysisService is None:
        raise HTTPException(status_code=501, detail="Task manager not available")
    
    logger.info(f"[AI Analysis] Saving analysis for record={record_id}, stage={stage_id}, model={request.model_used}")
    
    # 验证任务记录存在
    record = TaskHistoryService.get_record(record_id)
    if not record:
        logger.warning(f"[AI Analysis] Record not found: {record_id}")
        raise HTTPException(status_code=404, detail=f"Record not found: {record_id}")

    # 解析 SUGGESTED_PARAMS 行（用于返回给前端），但存储原始文本（保留标记行，供 GET 接口反解析）
    suggested_params = _parse_suggested_params(request.analysis_text)
    
    success = StageAnalysisService.save_analysis(
        record_id=record_id,
        stage_id=stage_id,
        analysis_text=request.analysis_text,  # 存原始文本，GET 接口返回时再剥离
        model_used=request.model_used
    )
    
    if not success:
        logger.error(f"[AI Analysis] Failed to save analysis for record={record_id}, stage={stage_id}")
        raise HTTPException(status_code=500, detail="Failed to save analysis")
    
    logger.info(f"[AI Analysis] Successfully saved analysis for record={record_id}, stage={stage_id}, suggested_params={suggested_params}")
    return {
        "success": True,
        "message": "分析结果已保存",
        "record_id": record_id,
        "stage_id": stage_id,
        "suggested_params": suggested_params,  # 解析到的参数建议（可能为 None）
    }


@router.delete("/history/{record_id}/stages/{stage_id}/analysis")
async def delete_stage_analysis(
    record_id: str,
    stage_id: str
) -> Dict[str, Any]:
    """
    删除阶段 AI 分析结果
    
    删除指定阶段的 AI 分析缓存（用于重新生成分析）。
    
    Args:
        record_id: 任务记录ID
        stage_id: 阶段ID
        
    Returns:
        操作结果
    """
    if not TASK_MANAGER_AVAILABLE or StageAnalysisService is None:
        raise HTTPException(status_code=501, detail="Task manager not available")
    
    deleted_count = StageAnalysisService.delete_analysis(record_id, stage_id)
    
    return {
        "success": True,
        "deleted_count": deleted_count,
        "record_id": record_id,
        "stage_id": stage_id
    }


@router.get("/history/{record_id}/analyses")
async def get_all_stage_analyses(
    record_id: str
) -> Dict[str, Any]:
    """
    获取任务所有阶段的 AI 分析结果
    
    批量获取任务所有已保存的 AI 分析。
    
    Args:
        record_id: 任务记录ID
        
    Returns:
        所有阶段的分析结果列表
    """
    if not TASK_MANAGER_AVAILABLE or StageAnalysisService is None:
        raise HTTPException(status_code=501, detail="Task manager not available")
    
    # 验证任务记录存在
    record = TaskHistoryService.get_record(record_id)
    if not record:
        raise HTTPException(status_code=404, detail=f"Record not found: {record_id}")
    
    analyses = StageAnalysisService.get_all_analyses_for_record(record_id)
    
    return {
        "record_id": record_id,
        "analyses": analyses,
        "count": len(analyses)
    }


@router.get("/statistics", response_model=TaskStatisticsResponse)
async def get_task_statistics(
    task_type: Optional[str] = Query(None, description="任务类型筛选"),
    task_category: Optional[str] = Query(None, description="任务类别筛选"),
    days: int = Query(30, ge=1, le=365, description="统计天数")
) -> TaskStatisticsResponse:
    """
    获取任务统计信息
    
    包括任务数量、成功率、平均耗时等。
    """
    if not TASK_MANAGER_AVAILABLE:
        raise HTTPException(status_code=501, detail="Task manager not available")
    
    stats = TaskHistoryService.get_statistics(
        task_type=task_type,
        task_category=task_category,
        days=days
    )
    
    return TaskStatisticsResponse(**stats)


# =============================================================================
# Expert Mode Endpoints
# =============================================================================

# Expert mode is now integrated into the main ExecutionStore
# The separate ExpertExecutionStore has been deprecated


def get_execution_context(execution_id: str):
    """
    统一获取执行上下文（从 ExecutionStore）
    
    Args:
        execution_id: 执行ID
        
    Returns:
        ExecutionContext 或 None
    """
    return ExecutionStore.get(execution_id)


class ExpertExecutionRequest(BaseModel):
    """专家模式执行请求"""
    task_id: str = Field(..., description="任务类型ID")
    session_id: str = Field(..., description="会话ID")
    file_path: Optional[str] = Field(None, description="数据文件路径")
    params: Dict[str, Any] = Field(default_factory=dict, description="任务参数")


class ExpertStageParamsRequest(BaseModel):
    """阶段参数更新请求"""
    params: Dict[str, Any] = Field(..., description="阶段参数")


class ExpertStageCodeRequest(BaseModel):
    """阶段代码更新请求"""
    code: str = Field(..., description="阶段代码")


@router.post("/expert/create")
async def create_expert_execution(request: ExpertExecutionRequest) -> Dict[str, Any]:
    """
    创建专家模式执行上下文
    
    注意：此API已废弃，建议使用 /sop/execute 启动任务（设置 interaction_mode=expert）
    保留此端点是为了向后兼容
    
    Args:
        request: 包含task_id, session_id, file_path, params
        
    Returns:
        执行上下文信息
    """
    # P1-D5: 入口层 session_id 校验
    try:
        request.session_id = _validate_session_id(request.session_id)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    
    # 重定向到统一的执行API
    logger.warning("Deprecated: /sop/expert/create is deprecated, use /sop/execute with interaction_mode=expert")
    
    # Resolve file path
    full_file_path = None
    if request.file_path:
        full_file_path = resolve_file_path(request.file_path)
        if not os.path.exists(full_file_path):
            raise HTTPException(status_code=404, detail="File not found")
    
    # 使用统一的执行逻辑
    try:
        from deepanalyze.analysis.task_SOP.executor import TaskExecutor
        
        executor = TaskExecutor(os.path.dirname(full_file_path) if full_file_path else ".")
        context = executor.create_execution(
            task_id=request.task_id,
            session_id=request.session_id,
            file_path=str(full_file_path) if full_file_path else None,
            params=request.params,
            interaction_mode="expert"
        )
        
        return context.to_dict()
        
    except ValueError as e:
        logger.warning(f"Expert execution validation error: {e}")
        raise HTTPException(status_code=400, detail="Invalid parameters for expert execution")
    except Exception as e:
        logger.error(f"Failed to create expert execution: {e}")
        raise HTTPException(status_code=500, detail="Failed to create expert execution")


@router.get("/expert/{execution_id}")
async def get_expert_execution(execution_id: str) -> Dict[str, Any]:
    """
    获取专家模式执行上下文
    
    Args:
        execution_id: 执行ID
        
    Returns:
        执行上下文信息
    """
    context = get_execution_context(execution_id)
    if not context:
        raise HTTPException(status_code=404, detail="Execution not found")
    
    return context.to_dict()


@router.post("/expert/{execution_id}/stages/{stage_id}/execute")
async def execute_expert_stage(
    execution_id: str,
    stage_id: str,
    background_tasks: BackgroundTasks
) -> Dict[str, Any]:
    """
    执行专家模式单个阶段
    
    Args:
        execution_id: 执行ID
        stage_id: 阶段ID
        
    Returns:
        阶段执行结果
    """
    context = get_execution_context(execution_id)
    if not context:
        raise HTTPException(status_code=404, detail="Execution not found")
    
    if stage_id not in context.stages:
        raise HTTPException(status_code=404, detail=f"Stage not found: {stage_id}")
    
    # 使用统一的执行器
    from deepanalyze.analysis.task_SOP.executor import TaskExecutor
    executor = TaskExecutor(context.file_path or ".")
    
    # Execute stage in background
    async def run_stage():
        try:
            await executor.execute_stage(execution_id, stage_id)
        except Exception as e:
            logger.error(f"Expert stage execution failed: {e}")
    
    background_tasks.add_task(run_stage)
    
    return {
        "message": "阶段执行已启动",
        "execution_id": execution_id,
        "stage_id": stage_id,
        "status": "running"
    }


@router.post("/expert/{execution_id}/stages/{stage_id}/skip")
async def skip_expert_stage(
    execution_id: str,
    stage_id: str,
    reason: str = Query("", description="跳过原因")
) -> Dict[str, Any]:
    """
    跳过专家模式阶段
    
    Args:
        execution_id: 执行ID
        stage_id: 阶段ID
        reason: 跳过原因
        
    Returns:
        操作结果
    """
    context = get_execution_context(execution_id)
    if not context:
        raise HTTPException(status_code=404, detail="Execution not found")
    
    if stage_id not in context.stages:
        raise HTTPException(status_code=404, detail=f"Stage not found: {stage_id}")
    
    try:
        stage = context.stages[stage_id]
        stage.status = "skipped"
        stage.error = reason if reason else "用户跳过"
        ExecutionStore.update(context)
        
        return {
            "success": True,
            "message": "阶段已跳过",
            "stage": {
                "stage_id": stage.stage_id,
                "stage_name": stage.stage_name,
                "status": stage.status,
                "error": stage.error
            }
        }
        
    except ValueError as e:
        logger.warning(f"Expert stage skip validation error: {e}")
        raise HTTPException(status_code=400, detail="Invalid operation")


@router.post("/expert/{execution_id}/stages/{stage_id}/reset")
async def reset_expert_stage(
    execution_id: str,
    stage_id: str
) -> Dict[str, Any]:
    """
    重置专家模式阶段
    
    Args:
        execution_id: 执行ID
        stage_id: 阶段ID
        
    Returns:
        操作结果
    """
    context = get_execution_context(execution_id)
    if not context:
        raise HTTPException(status_code=404, detail="Execution not found")
    
    if stage_id not in context.stages:
        raise HTTPException(status_code=404, detail=f"Stage not found: {stage_id}")
    
    try:
        stage = context.stages[stage_id]
        stage.status = "pending"
        stage.error = None
        stage.output_preview = None
        stage.execution_time_ms = None
        ExecutionStore.update(context)
        
        return {
            "success": True,
            "message": "阶段已重置",
            "stage": {
                "stage_id": stage.stage_id,
                "stage_name": stage.stage_name,
                "status": stage.status
            }
        }
        
    except ValueError as e:
        logger.warning(f"Expert stage reset validation error: {e}")
        raise HTTPException(status_code=400, detail="Invalid operation")


@router.put("/expert/{execution_id}/stages/{stage_id}/params")
async def update_expert_stage_params(
    execution_id: str,
    stage_id: str,
    request: ExpertStageParamsRequest
) -> Dict[str, Any]:
    """
    更新专家模式阶段参数
    
    Args:
        execution_id: 执行ID
        stage_id: 阶段ID
        request: 参数更新请求
        
    Returns:
        操作结果
    """
    context = get_execution_context(execution_id)
    if not context:
        raise HTTPException(status_code=404, detail="Execution not found")
    
    if stage_id not in context.stages:
        raise HTTPException(status_code=404, detail=f"Stage not found: {stage_id}")
    
    stage = context.stages[stage_id]
    # 更新阶段使用的参数
    stage.params_used.update(request.params)
    # 同时更新全局参数（供下次执行使用）
    context.params.update(request.params)
    ExecutionStore.update(context)
    
    return {
        "success": True,
        "message": "参数已更新",
        "stage": {
            "stage_id": stage.stage_id,
            "stage_name": stage.stage_name,
            "status": stage.status,
            "params_used": stage.params_used
        }
    }


@router.put("/expert/{execution_id}/stages/{stage_id}/code")
async def update_expert_stage_code(
    execution_id: str,
    stage_id: str,
    request: ExpertStageCodeRequest
) -> Dict[str, Any]:
    """
    更新专家模式阶段代码
    
    Args:
        execution_id: 执行ID
        stage_id: 阶段ID
        request: 代码更新请求
        
    Returns:
        操作结果
    """
    context = get_execution_context(execution_id)
    if not context:
        raise HTTPException(status_code=404, detail="Execution not found")
    
    if stage_id not in context.stages:
        raise HTTPException(status_code=404, detail=f"Stage not found: {stage_id}")
    
    try:
        stage = context.stages[stage_id]
        # 存储自定义代码到阶段参数中
        stage.params_used["custom_code"] = request.code
        ExecutionStore.update(context)
        
        return {
            "success": True,
            "message": "代码已更新",
            "stage": {
                "stage_id": stage.stage_id,
                "stage_name": stage.stage_name,
                "status": stage.status,
                "has_custom_code": True
            }
        }
        
    except ValueError as e:
        logger.warning(f"Expert stage code validation error: {e}")
        raise HTTPException(status_code=400, detail="Invalid operation")


@router.get("/expert/{execution_id}/stages/{stage_id}/result")
async def get_expert_stage_result(
    execution_id: str,
    stage_id: str
) -> Dict[str, Any]:
    """
    获取专家模式阶段结果
    
    Args:
        execution_id: 执行ID
        stage_id: 阶段ID
        
    Returns:
        阶段结果
    """
    context = get_execution_context(execution_id)
    if not context:
        raise HTTPException(status_code=404, detail="Execution not found")
    
    if stage_id not in context.stages:
        raise HTTPException(status_code=404, detail=f"Stage not found: {stage_id}")
    
    stage = context.stages[stage_id]
    
    return {
        "execution_id": execution_id,
        "stage_id": stage_id,
        "has_result": stage.output_preview is not None,
        "result": stage.output_preview
    }


# =============================================================================
# Overall AI Analysis API
# =============================================================================

class OverallAnalysisRequest(BaseModel):
    """整体分析请求"""
    record_id: str
    task_type: str
    analysis_text: Optional[str] = None  # AI分析文本
    model_used: Optional[str] = None  # 使用的模型
    results: Optional[Dict[str, Any]] = None
    stages_data: Optional[Dict[str, Any]] = None


class OverallAnalysisGenerateRequest(BaseModel):
    """生成整体分析请求"""
    record_id: str
    task_type: str
    execution_id: Optional[str] = None


@router.get("/history/{record_id}/overall-analysis")
async def get_overall_analysis(record_id: str) -> Dict[str, Any]:
    """
    获取任务整体 AI 分析结果
    
    Args:
        record_id: 任务记录ID
        
    Returns:
        分析结果
    """
    from deepanalyze.core.task_manager.overall_analysis_service import OverallAnalysisService
    
    analysis = OverallAnalysisService.get_analysis(record_id)
    
    if analysis:
        return {
            "success": True,
            "exists": True,
            "analysis": analysis
        }
    else:
        return {
            "success": True,
            "exists": False,
            "analysis": None
        }


@router.post("/history/{record_id}/overall-analysis")
async def save_overall_analysis(
    record_id: str,
    request: OverallAnalysisRequest
) -> Dict[str, Any]:
    """
    保存任务整体 AI 分析结果
    
    Args:
        record_id: 任务记录ID
        request: 分析请求
        
    Returns:
        保存结果
    """
    from deepanalyze.core.task_manager.overall_analysis_service import OverallAnalysisService
    
    # 直接保存传入的分析文本（前端已生成）
    if hasattr(request, 'analysis_text') and request.analysis_text:
        success = OverallAnalysisService.save_analysis(
            record_id=record_id,
            task_type=request.task_type,
            analysis_text=request.analysis_text,
            model_used=getattr(request, 'model_used', None)
        )
        
        return {
            "success": success,
            "message": "分析已保存" if success else "保存失败"
        }
    
    return {
        "success": False,
        "message": "缺少分析文本"
    }


@router.delete("/history/{record_id}/overall-analysis")
async def delete_overall_analysis(record_id: str) -> Dict[str, Any]:
    """
    删除任务整体 AI 分析结果
    
    Args:
        record_id: 任务记录ID
        
    Returns:
        删除结果
    """
    from deepanalyze.core.task_manager.overall_analysis_service import OverallAnalysisService
    
    success = OverallAnalysisService.delete_analysis(record_id)
    
    return {
        "success": success,
        "message": "分析已删除" if success else "删除失败"
    }


@router.post("/overall-analysis/build-prompt")
async def build_overall_analysis_prompt(
    request: OverallAnalysisGenerateRequest
) -> Dict[str, Any]:
    """
    构建整体分析Prompt
    
    根据任务类型和执行结果，构建用于生成整体AI分析的Prompt。
    前端可使用此Prompt调用LLM生成分析。
    
    Args:
        request: 生成请求
        
    Returns:
        Prompt文本
    """
    from deepanalyze.core.task_manager.overall_analysis_service import OverallAnalysisService
    from deepanalyze.core.task_manager.task_result_config import get_task_result_config
    
    # 获取执行结果
    results = {}
    stages_data = {}
    
    if request.execution_id:
        context = get_execution_context(request.execution_id)
        if context:
            # 从执行上下文获取结果
            results = context.result or {}
            # 从阶段获取数据
            for stage_id, stage in context.stages.items():
                if stage.output_preview:
                    stages_data[stage_id] = stage.output_preview
    
    # 构建数据描述
    data_description = OverallAnalysisService.build_data_description(
        task_type=request.task_type,
        results=results,
        stages_data=stages_data
    )
    
    # 构建Prompt
    prompt = OverallAnalysisService.build_prompt(
        task_type=request.task_type,
        data_description=data_description
    )
    
    # 获取配置信息
    config = get_task_result_config(request.task_type)
    
    return {
        "success": True,
        "prompt": prompt,
        "data_description": data_description,
        "config": {
            "task_name": config.task_name if config else request.task_type,
            "max_words": config.ai_analysis_config.max_words if config else 200,
            "focus_areas": config.ai_analysis_config.focus_areas if config else []
        }
    }


# =============================================================================
# P2-7: 先验规则解析
# =============================================================================

@router.post("/prior-rules/parse")
async def parse_prior_rules(
    file: UploadFile | None = File(None),
    text: str | None = Form(None),
    session_id: str | None = Form(None),
    data_file: str | None = Form(None),
):
    """
    解析先验规则（CSV 文件上传或文本输入）
    
    支持：
    - CSV/Excel 文件上传（自动识别结构化/表达式格式）
    - 文本直接输入（每行一条表达式）
    
    可选：提供 session_id + data_file 校验列名
    """
    import tempfile
    import os
    from deepanalyze.analysis.task_SOP.prior_rule_parser import PriorRuleParser
    
    parser = PriorRuleParser()
    
    if file and file.filename:
        # F3 安全修复：用 tempfile 替代 /tmp/{filename}
        suffix = os.path.splitext(file.filename)[1] or '.csv'
        with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
            content = await file.read()
            tmp.write(content)
            tmp_path = tmp.name
        
        try:
            parser.parse_csv(tmp_path)
        finally:
            os.unlink(tmp_path)
    elif text:
        parser.parse_text(text)
    else:
        raise HTTPException(400, "需要提供文件或文本")
    
    # 可选列名校验
    validation = None
    if session_id and data_file and parser.rules:
        try:
            workspace_dir = os.path.join("workspace", session_id)
            file_path = os.path.join(workspace_dir, data_file)
            if os.path.exists(file_path):
                sample_df = pd.read_csv(file_path, nrows=1)
                validation = parser.validate_columns(set(sample_df.columns))
        except Exception as e:
            validation = {"error": str(e)}
    
    result = parser.to_dict()
    if validation:
        result["validation"] = validation
    
    return {"success": True, **result}


# =============================================================================
# Export
# =============================================================================

__all__ = ['router']
