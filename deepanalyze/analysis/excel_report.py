# pyright: reportAny=false, reportExplicitAny=false, reportUnknownMemberType=false, reportUnknownArgumentType=false, reportUnknownVariableType=false, reportUnknownParameterType=false
"""
Excel Report Generator Module

Provides professional Excel report generation for scorecard and rule mining results.

Features:
- Multi-sheet Excel reports
- Professional styling (headers, data bars, conditional formatting)
- Chart embedding (ROC, KS, score distribution)
- Customizable templates
- Section selection for partial exports

This module uses openpyxl for Excel generation with advanced formatting.
"""

import io
import pandas as pd
import numpy as np
from typing import Any, Literal, TYPE_CHECKING
from pathlib import Path

if TYPE_CHECKING:
    from deepanalyze.core.task_manager.task_result_config import StageDataSheetConfig

try:
    from openpyxl import Workbook
    from openpyxl.styles import (
        Font, PatternFill, Border, Side, Alignment,
        NamedStyle, numbers
    )
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.formatting.rule import DataBarRule, ColorScaleRule, FormulaRule
    from openpyxl.chart import LineChart, BarChart, Reference
    from openpyxl.chart.label import DataLabelList
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


# Style definitions
class ExcelStyles:
    """Excel style definitions for professional reports."""
    
    # Colors
    HEADER_BG = "1F4E79"  # Dark blue
    HEADER_FG = "FFFFFF"  # White
    SUBHEADER_BG = "D6DCE4"  # Light gray-blue
    ACCENT_BG = "E2EFDA"  # Light green
    WARNING_BG = "FCE4D6"  # Light orange
    ERROR_BG = "FFC7CE"  # Light red
    
    # Fonts
    TITLE_FONT = Font(name='Arial', size=14, bold=True, color=HEADER_FG)
    HEADER_FONT = Font(name='Arial', size=11, bold=True, color=HEADER_FG)
    SUBHEADER_FONT = Font(name='Arial', size=10, bold=True)
    DATA_FONT = Font(name='Arial', size=10)
    
    # Fills
    HEADER_FILL = PatternFill(start_color=HEADER_BG, end_color=HEADER_BG, fill_type="solid")
    SUBHEADER_FILL = PatternFill(start_color=SUBHEADER_BG, end_color=SUBHEADER_BG, fill_type="solid")
    ACCENT_FILL = PatternFill(start_color=ACCENT_BG, end_color=ACCENT_BG, fill_type="solid")
    
    # Borders
    THIN_BORDER = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Alignments
    CENTER = Alignment(horizontal='center', vertical='center')
    LEFT = Alignment(horizontal='left', vertical='center')
    RIGHT = Alignment(horizontal='right', vertical='center')


class ExcelReportGenerator:
    """
    Excel report generator for scorecard and rule mining results.
    
    Generates professional multi-sheet Excel reports with:
    - Overview summary
    - Detailed data tables
    - Charts and visualizations
    - Conditional formatting
    
    Example:
        >>> generator = ExcelReportGenerator(style_template='professional')
        >>> report_bytes = generator.generate_scorecard_report(
        ...     results=scorecard_results,
        ...     sections=['overview', 'bins', 'scorecard', 'statistics']
        ... )
        >>> with open('report.xlsx', 'wb') as f:
        ...     f.write(report_bytes)
    """
    
    # Sections corresponding to UI tabs
    # 评分卡UI tabs: 评估图表, 特征筛选, 评分卡明细, IV值排序, 模型系数, 统计检验, 评分转换
    AVAILABLE_SECTIONS_SCORECARD = [
        'charts',           # 评估图表
        'selection',        # 特征筛选
        'scorecard',        # 评分卡明细
        'iv',               # IV值排序
        'coefficients',     # 模型系数
        'statistics',       # 统计检验
    ]
    
    # 规则挖掘UI tabs: 评估图表, 最优规则, 过滤后, 全部, 质量验证, 稳定性, 金额分析
    AVAILABLE_SECTIONS_RULE_MINING = [
        'charts',           # 评估图表
        'optimal',          # 最优规则
        'filtered',         # 过滤后
        'all',              # 全部
        'validation',       # 质量验证
        'psi',              # 稳定性
        'amount',           # 金额分析
    ]
    
    def __init__(
        self,
        style_template: Literal['professional', 'simple', 'colorful'] = 'professional'
    ):
        """
        Initialize ExcelReportGenerator.
        
        Args:
            style_template: Style template to use
                - 'professional': Corporate style with subtle colors
                - 'simple': Minimal styling
                - 'colorful': More vibrant colors
        """
        if not OPENPYXL_AVAILABLE:
            raise ImportError(
                "openpyxl is required for Excel report generation. "
                "Install it with: pip install openpyxl"
            )
        
        self.style_template = style_template
        self.styles = ExcelStyles()
        
    def generate_scorecard_report(
        self,
        results: dict[str, Any],
        sections: list[str] | None = None,
        include_charts: bool = True,
        title: str = "评分卡开发报告",
        ai_analysis: str | None = None
    ) -> bytes:
        """
        Generate Excel report for scorecard development results.
        
        与规则挖掘报告结构一致：
        - 首位：任务报告 Sheet（整合 MD/HTML/Word 报告的完整内容）
        - 后续：各阶段 Sheets
        
        Args:
            results: Scorecard development results dictionary
            sections: List of sections to include (None = all)
            include_charts: Whether to include charts
            title: Report title
            
        Returns:
            Excel file as bytes
        """
        # sections 参数保留但不再用于控制独立 Sheet（已整合到任务报告）
        _ = sections
        
        wb = Workbook()
        
        # Remove default sheet
        if 'Sheet' in wb.sheetnames:
            del wb['Sheet']
        
        # 首位：任务报告 Sheet（整合 HTML/Word 报告的完整内容）
        # 优先使用传入的 ai_analysis 参数，否则从 results 中获取
        ai_analysis_text = ai_analysis if ai_analysis is not None else results.get('ai_analysis')
        self._add_scorecard_task_report_sheet(wb, results, title, ai_analysis_text)
        
        # 添加各阶段下载结果Sheets（使用配置驱动）
        if results.get('stages'):
            stage_configs = results.get('stage_configs')  # 从results获取配置
            self._add_stages_sheets(wb, results['stages'], stage_configs)
        
        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()
    
    def generate_rule_mining_report(
        self,
        results: dict[str, Any],
        sections: list[str] | None = None,
        include_charts: bool = True,
        title: str = "规则挖掘报告"
    ) -> bytes:
        """
        Generate Excel report for rule mining results.
        
        Args:
            results: Rule mining results dictionary
            sections: List of sections to include (None = all)
            include_charts: Whether to include charts
            title: Report title
            
        Returns:
            Excel file as bytes
        """
        # sections 参数保留但不再用于控制独立 Sheet（已整合到任务报告）
        _ = sections  # 标记为已使用，避免 lint 警告
        
        wb = Workbook()
        
        if 'Sheet' in wb.sheetnames:
            del wb['Sheet']
        
        # 首位：任务报告 Sheet（整合 HTML/Word 报告的完整内容）
        ai_analysis = results.get('ai_analysis')
        self._add_task_report_sheet(wb, results, title, ai_analysis)
        
        # 添加各阶段 Sheets（1_数据预处理 ~ 6_报告生成）
        # 独立的概览/最优规则/质量验证等 Sheet 已移除（内容整合到任务报告中）
        if results.get('stages'):
            stage_configs = results.get('stage_configs')  # 从results获取配置
            self._add_stages_sheets(wb, results['stages'], stage_configs)
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()
    
    # =========================================================================
    # Task Report Sheet (任务报告 - 整合 HTML/Word 报告的完整内容)
    # =========================================================================
    
    def _add_task_report_sheet(
        self,
        wb: "Workbook",
        results: dict[str, Any],
        title: str,
        ai_analysis: str | None = None
    ) -> None:
        """
        添加任务报告 Sheet，整合 HTML/Word/MD 报告的完整内容。
        
        结构与 HTML/Word 报告保持一致：
        - 一、概览（汇总指标 + AI 分析）
        - 二、样本及特征
        - 三、最优规则
        - 四、规则筛选流程
        - 五、质量验证
        - 六、稳定性分析
        - 七、附加分析
        """
        ws = wb.create_sheet("任务报告", 0)  # 插入到第一个位置
        
        # 设置列宽
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 50
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 15
        ws.column_dimensions['H'].width = 15
        
        row = 1
        
        # ========== 标题 ==========
        ws.merge_cells('A1:H1')
        ws['A1'] = title
        ws['A1'].font = self.styles.TITLE_FONT
        ws['A1'].fill = self.styles.HEADER_FILL
        ws['A1'].alignment = self.styles.CENTER
        ws.row_dimensions[1].height = 35
        row = 3
        
        # ========== 一、概览 ==========
        row = self._write_section_header(ws, row, "一、概览")
        
        # 汇总指标卡片
        optimal_rules = results.get('optimal_rules', results.get('rules', []))
        # Phase 25: 兼容 DataFrame 和 list
        if isinstance(optimal_rules, pd.DataFrame) and not optimal_rules.empty:
            n_rules = len(optimal_rules)
            last_rule = optimal_rules.iloc[-1].to_dict()
            final_recall = last_rule.get('cumulative_recall', last_rule.get('cum_recall', last_rule.get('dev_cum_recall', 0)))
            final_hit_rate = last_rule.get('cumulative_hit_rate', last_rule.get('cum_hit_rate', last_rule.get('dev_cum_hit_rate', 0)))
            final_lift = last_rule.get('cumulative_lift', last_rule.get('cum_lift', last_rule.get('dev_cum_lift', last_rule.get('lift', 0))))
            
            overview_metrics = [
                ("最优规则数", str(n_rules)),
                ("累计召回率", f"{final_recall*100:.1f}%"),
                ("累计命中率", f"{final_hit_rate*100:.1f}%"),
                ("累计提升倍数", f"{final_lift:.2f}x"),
            ]
            
            for name, value in overview_metrics:
                ws.cell(row=row, column=1, value=name).font = self.styles.DATA_FONT
                ws.cell(row=row, column=2, value=value).font = Font(name='Arial', size=10, bold=True)
                row += 1
            row += 1
        
        # AI 分析摘要
        if ai_analysis and ai_analysis.strip():
            ws.cell(row=row, column=1, value="AI 分析评估").font = self.styles.SUBHEADER_FONT
            ws.cell(row=row, column=1).fill = self.styles.SUBHEADER_FILL
            row += 1
            
            # 将 AI 分析文本分段写入（处理 Markdown 格式）
            row = self._write_ai_analysis_to_sheet(ws, ai_analysis, row)
            row += 1
        
        # ========== 二、样本及特征 ==========
        row = self._write_section_header(ws, row, "二、样本及特征")
        stages = results.get('stages', {})
        row = self._write_sample_features_section(ws, stages, row)
        
        # ========== 三、最优规则 ==========
        row = self._write_section_header(ws, row, "三、最优规则")
        if isinstance(optimal_rules, pd.DataFrame) and not optimal_rules.empty:
            row = self._write_rules_table_to_report(ws, optimal_rules, row, max_rules=15)
        else:
            ws.cell(row=row, column=1, value="暂无数据").font = self.styles.DATA_FONT
            row += 1
        row += 1
        
        # ========== 四、规则筛选流程 ==========
        row = self._write_section_header(ws, row, "四、规则筛选流程")
        row = self._write_filtering_flow_section(ws, stages, row)
        
        # ========== 五、质量验证 ==========
        row = self._write_section_header(ws, row, "五、质量验证")
        validation_report = results.get('validation_report')
        if validation_report:
            row = self._write_validation_section(ws, validation_report, row)
        else:
            ws.cell(row=row, column=1, value="暂无数据").font = self.styles.DATA_FONT
            row += 1
        row += 1
        
        # ========== 六、稳定性分析 ==========
        row = self._write_section_header(ws, row, "六、稳定性分析 (PSI)")
        psi_report = results.get('psi_report')
        if psi_report and isinstance(psi_report, list) and len(psi_report) > 0:
            row = self._write_psi_section(ws, psi_report, row)
        else:
            ws.cell(row=row, column=1, value="暂无数据").font = self.styles.DATA_FONT
            row += 1
        row += 1
        
        # ========== 七、附加分析 ==========
        row = self._write_section_header(ws, row, "七、附加分析")
        amount_analysis = results.get('amount_analysis')
        prior_analysis = results.get('prior_analysis')
        if amount_analysis or prior_analysis:
            row = self._write_advanced_analysis_section(ws, amount_analysis, prior_analysis, row)
        else:
            ws.cell(row=row, column=1, value="暂无数据").font = self.styles.DATA_FONT
            row += 1
    
    # =========================================================================
    # Scorecard Task Report Sheet (评分卡任务报告 - 整合 HTML/Word/MD 报告内容)
    # =========================================================================
    
    def _add_scorecard_task_report_sheet(
        self,
        wb: "Workbook",
        results: dict[str, Any],
        title: str,
        ai_analysis: str | None = None
    ) -> None:
        """
        添加评分卡任务报告 Sheet，整合 HTML/Word/MD 报告的完整内容。
        
        章节结构（2026-02-12 重构完成，与 HTML/Word/MD 报告对齐）：
        - 一、概览（汇总指标卡 + 数据集对比 + AI 分析）
        - 二、样本与特征（对应 Tab: sample-data）
        - 三、评估图表（对应 Tab: charts）- Excel中展示数据表格
        - 四、评分卡明细（对应 Tab: scorecard）
        - 五、变量筛选（对应 Tab: selection）- 合并特征筛选+IV排序
        - 六、模型系数（对应 Tab: statistics）- 合并系数+统计检验
        """
        import logging
        logger = logging.getLogger(__name__)
        
        # 调试日志：检查关键数据字段
        logger.info(f"[Excel Scorecard Report] results keys: {list(results.keys())}")
        logger.info(f"[Excel Scorecard Report] metrics: {results.get('metrics')}")
        logger.info(f"[Excel Scorecard Report] stages: {list(results.get('stages', {}).keys()) if results.get('stages') else 'None'}")
        
        ws = wb.create_sheet("任务报告", 0)  # 插入到第一个位置
        
        # 设置列宽
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 35
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 15
        ws.column_dimensions['H'].width = 15
        
        row = 1
        
        # ========== 标题 ==========
        ws.merge_cells('A1:H1')
        ws['A1'] = title
        ws['A1'].font = self.styles.TITLE_FONT
        ws['A1'].fill = self.styles.HEADER_FILL
        ws['A1'].alignment = self.styles.CENTER
        ws.row_dimensions[1].height = 35
        row = 3
        
        # ==========================================================================
        # 一、概览（重构：汇总指标卡 + 数据集对比 + AI分析）
        # ==========================================================================
        row = self._write_section_header(ws, row, "一、概览")
        row = self._write_scorecard_overview_section(ws, results, row)
        
        # AI 分析摘要
        if ai_analysis and ai_analysis.strip():
            ws.cell(row=row, column=1, value="AI 整体分析").font = self.styles.SUBHEADER_FONT
            ws.cell(row=row, column=1).fill = self.styles.SUBHEADER_FILL
            row += 1
            row = self._write_ai_analysis_to_sheet(ws, ai_analysis, row)
            row += 1
        
        # ==========================================================================
        # 二、样本与特征（新增，与 HTML/Word/MD 报告对齐）
        # ==========================================================================
        stages = results.get('stages', {})
        if stages:
            row = self._write_section_header(ws, row, "二、样本与特征")
            row = self._write_scorecard_sample_features_section(ws, stages, results, row)
        
        # ==========================================================================
        # 三、评分卡明细（原"四、评分卡明细"，章节号调整）
        # ==========================================================================
        row = self._write_section_header(ws, row, "三、评分卡明细")
        row = self._write_scorecard_details_section(ws, results, row)
        
        # ==========================================================================
        # 四、变量筛选（原"五、变量筛选"，章节号调整）
        # ==========================================================================
        row = self._write_section_header(ws, row, "四、变量筛选")
        row = self._write_scorecard_feature_selection_section(ws, results, row)
        
        # ==========================================================================
        # 五、模型系数（原"六、模型系数"，章节号调整）
        # ==========================================================================
        row = self._write_section_header(ws, row, "五、模型系数")
        row = self._write_scorecard_model_coefficients_section(ws, results, row)
    
    def _write_scorecard_overview_section(self, ws, results: dict, start_row: int) -> int:
        """
        写入评分卡概览章节（一、概览）
        包含：汇总指标卡 + 数据集指标对比
        """
        row = start_row
        
        metrics = results.get('metrics', {})
        multi_dataset_metrics = results.get('multi_dataset_metrics', {})
        psi_result = results.get('psi_result', {})
        
        # 动态获取指标数据来源标签
        metrics_source = metrics.get('source', 'test')
        source_label = 'OOT验证集' if metrics_source == 'oot' else '测试集'
        
        # 汇总指标卡
        ws.cell(row=row, column=1, value=f"数据来源：{source_label}").font = self.styles.DATA_FONT
        row += 1
        
        # 指标表格
        headers = ['指标', '值', '评级']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = self.styles.HEADER_FONT
            cell.fill = self.styles.HEADER_FILL
            cell.border = self.styles.THIN_BORDER
        row += 1
        
        # 指标评估函数
        def get_ks_level(ks):
            if ks is None:
                return '-'
            if ks >= 0.4:
                return '优秀'
            elif ks >= 0.3:
                return '良好'
            elif ks >= 0.2:
                return '可用'
            else:
                return '较差'
        
        def get_auc_level(auc):
            if auc is None:
                return '-'
            if auc >= 0.8:
                return '优秀'
            elif auc >= 0.75:
                return '良好'
            elif auc >= 0.7:
                return '可用'
            else:
                return '较差'
        
        def get_gini_level(gini):
            if gini is None:
                return '-'
            if gini >= 0.6:
                return '优秀'
            elif gini >= 0.5:
                return '良好'
            elif gini >= 0.4:
                return '可用'
            else:
                return '较差'
        
        def get_psi_level(psi):
            if psi is None:
                return '-'
            if psi < 0.1:
                return '稳定'
            elif psi < 0.25:
                return '轻微变化'
            else:
                return '显著变化'
        
        ks_val = metrics.get('ks')
        auc_val = metrics.get('auc')
        gini_val = metrics.get('gini')
        psi_val = psi_result.get('value') if psi_result and isinstance(psi_result, dict) and 'value' in psi_result else metrics.get('psi')
        
        # KS
        row_data = [
            'KS值',
            f"{ks_val * 100:.2f}%" if ks_val is not None else '-',
            get_ks_level(ks_val)
        ]
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.font = self.styles.DATA_FONT
            cell.border = self.styles.THIN_BORDER
        row += 1
        
        # AUC
        row_data = [
            'AUC',
            f"{auc_val:.4f}" if auc_val is not None else '-',
            get_auc_level(auc_val)
        ]
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.font = self.styles.DATA_FONT
            cell.border = self.styles.THIN_BORDER
        row += 1
        
        # Gini
        row_data = [
            'Gini系数',
            f"{gini_val * 100:.2f}%" if gini_val is not None else '-',
            get_gini_level(gini_val)
        ]
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.font = self.styles.DATA_FONT
            cell.border = self.styles.THIN_BORDER
        row += 1
        
        # PSI
        psi_comparison = psi_result.get('comparison', '稳定性') if psi_result and isinstance(psi_result, dict) else '稳定性'
        row_data = [
            f'PSI ({psi_comparison})',
            f"{psi_val:.4f}" if psi_val is not None else '-',
            get_psi_level(psi_val)
        ]
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.font = self.styles.DATA_FONT
            cell.border = self.styles.THIN_BORDER
        row += 2
        
        # 数据集指标对比
        ws.cell(row=row, column=1, value="数据集指标对比").font = self.styles.SUBHEADER_FONT
        ws.cell(row=row, column=1).fill = self.styles.SUBHEADER_FILL
        row += 1
        
        headers = ['数据集', '样本数', '坏账率', 'KS', 'AUC', 'Gini']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = self.styles.HEADER_FONT
            cell.fill = self.styles.HEADER_FILL
            cell.border = self.styles.THIN_BORDER
        row += 1
        
        dataset_names = {'train': '训练集', 'test': '测试集', 'oot': 'OOT验证集'}
        for dataset_key, dataset_name in dataset_names.items():
            dataset_metrics = multi_dataset_metrics.get(dataset_key) if multi_dataset_metrics else None
            
            if dataset_metrics:
                ks = dataset_metrics.get('ks')
                auc = dataset_metrics.get('auc')
                gini = dataset_metrics.get('gini')
                samples = dataset_metrics.get('samples')
                bad_rate = dataset_metrics.get('bad_rate')
                
                samples_str = f"{samples:,}" if samples is not None else '-'
                bad_rate_str = f"{bad_rate:.2f}%" if bad_rate is not None else '-'
                ks_str = f"{ks * 100:.2f}%" if ks is not None else '-'
                auc_str = f"{auc:.4f}" if auc is not None else '-'
                gini_str = f"{gini * 100:.2f}%" if gini is not None else '-'
            else:
                samples_str = bad_rate_str = ks_str = auc_str = gini_str = '-'
            
            row_data = [dataset_name, samples_str, bad_rate_str, ks_str, auc_str, gini_str]
            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.font = self.styles.DATA_FONT
                cell.border = self.styles.THIN_BORDER
            row += 1
        
        return row + 1
    
    def _write_scorecard_sample_features_section(self, ws, stages: dict, results: dict, start_row: int) -> int:
        """
        写入评分卡样本与特征章节（二、样本与特征）
        与 HTML/Word/MD 报告对齐
        """
        import logging
        logger = logging.getLogger(__name__)
        
        row = start_row
        
        data_loading = stages.get('data_loading', {})
        data_loading_preview = data_loading.get('output_preview', {}) if isinstance(data_loading, dict) else {}
        
        # 调试日志
        logger.info(f"[Excel Scorecard SampleFeatures] stages keys: {list(stages.keys())}")
        logger.info(f"[Excel Scorecard SampleFeatures] data_loading type: {type(data_loading)}")
        logger.info(f"[Excel Scorecard SampleFeatures] data_loading_preview keys: {list(data_loading_preview.keys()) if isinstance(data_loading_preview, dict) else 'not dict'}")
        logger.info(f"[Excel Scorecard SampleFeatures] rows: {data_loading_preview.get('rows')}")
        logger.info(f"[Excel Scorecard SampleFeatures] target_rate: {data_loading_preview.get('target_rate')}")
        logger.info(f"[Excel Scorecard SampleFeatures] split_info: {data_loading_preview.get('split_info')}")
        
        # 2.1 样本概览
        ws.cell(row=row, column=1, value="样本概览").font = self.styles.SUBHEADER_FONT
        ws.cell(row=row, column=1).fill = self.styles.SUBHEADER_FILL
        row += 1
        
        headers = ['指标', '值']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = self.styles.HEADER_FONT
            cell.fill = self.styles.HEADER_FILL
            cell.border = self.styles.THIN_BORDER
        row += 1
        
        total_rows = data_loading_preview.get('rows')
        target_rate = data_loading_preview.get('target_rate')
        split_info = data_loading_preview.get('split_info', {})
        
        # 总样本数
        ws.cell(row=row, column=1, value="总样本数").font = self.styles.DATA_FONT
        ws.cell(row=row, column=1).border = self.styles.THIN_BORDER
        ws.cell(row=row, column=2, value=f"{total_rows:,}" if isinstance(total_rows, int) else str(total_rows)).font = self.styles.DATA_FONT
        ws.cell(row=row, column=2).border = self.styles.THIN_BORDER
        row += 1
        
        # 总体坏账率
        ws.cell(row=row, column=1, value="总体坏账率").font = self.styles.DATA_FONT
        ws.cell(row=row, column=1).border = self.styles.THIN_BORDER
        ws.cell(row=row, column=2, value=f"{target_rate * 100:.2f}%" if target_rate is not None else '-').font = self.styles.DATA_FONT
        ws.cell(row=row, column=2).border = self.styles.THIN_BORDER
        row += 1
        
        # 训练集
        if split_info.get('train'):
            train_count = split_info.get('train', 0)
            train_rate = split_info.get('train_target_rate')
            rate_str = f"{train_rate * 100:.2f}%" if isinstance(train_rate, (int, float)) else "-"
            ws.cell(row=row, column=1, value="训练集").font = self.styles.DATA_FONT
            ws.cell(row=row, column=1).border = self.styles.THIN_BORDER
            ws.cell(row=row, column=2, value=f"{train_count:,} (坏账率: {rate_str})").font = self.styles.DATA_FONT
            ws.cell(row=row, column=2).border = self.styles.THIN_BORDER
            row += 1
        
        # 测试集
        if split_info.get('test'):
            test_count = split_info.get('test', 0)
            test_rate = split_info.get('test_target_rate')
            rate_str = f"{test_rate * 100:.2f}%" if isinstance(test_rate, (int, float)) else "-"
            ws.cell(row=row, column=1, value="测试集").font = self.styles.DATA_FONT
            ws.cell(row=row, column=1).border = self.styles.THIN_BORDER
            ws.cell(row=row, column=2, value=f"{test_count:,} (坏账率: {rate_str})").font = self.styles.DATA_FONT
            ws.cell(row=row, column=2).border = self.styles.THIN_BORDER
            row += 1
        
        # OOT验证集
        oot_count = split_info.get('oot', 0)
        ws.cell(row=row, column=1, value="OOT验证集").font = self.styles.DATA_FONT
        ws.cell(row=row, column=1).border = self.styles.THIN_BORDER
        if oot_count and oot_count > 0:
            oot_rate = split_info.get('oot_target_rate')
            rate_str = f"{oot_rate * 100:.2f}%" if isinstance(oot_rate, (int, float)) else "-"
            ws.cell(row=row, column=2, value=f"{oot_count:,} (坏账率: {rate_str})").font = self.styles.DATA_FONT
        else:
            ws.cell(row=row, column=2, value="未划分").font = self.styles.DATA_FONT
        ws.cell(row=row, column=2).border = self.styles.THIN_BORDER
        row += 2
        
        # 2.2 时间范围
        time_range_info = data_loading_preview.get('time_range_info', {})
        if time_range_info:
            time_col = time_range_info.get('column', '')
            time_col_display = f"（{time_col}）" if time_col else ""
            ws.cell(row=row, column=1, value=f"时间范围{time_col_display}").font = self.styles.SUBHEADER_FONT
            ws.cell(row=row, column=1).fill = self.styles.SUBHEADER_FILL
            row += 1
            
            headers = ['数据集', '起始时间', '截止时间']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.font = self.styles.HEADER_FONT
                cell.fill = self.styles.HEADER_FILL
                cell.border = self.styles.THIN_BORDER
            row += 1
            
            train_range = time_range_info.get('train', {})
            if train_range:
                row_data = ['训练集', str(train_range.get('min', '-')), str(train_range.get('max', '-'))]
                for col, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row, column=col, value=value)
                    cell.font = self.styles.DATA_FONT
                    cell.border = self.styles.THIN_BORDER
                row += 1
            
            test_range = time_range_info.get('test', {})
            if test_range:
                row_data = ['测试集', str(test_range.get('min', '-')), str(test_range.get('max', '-'))]
                for col, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row, column=col, value=value)
                    cell.font = self.styles.DATA_FONT
                    cell.border = self.styles.THIN_BORDER
                row += 1
            
            oot_range = time_range_info.get('oot', {})
            if oot_range and oot_range.get('min'):
                row_data = ['OOT验证集', str(oot_range.get('min', '-')), str(oot_range.get('max', '-'))]
                for col, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row, column=col, value=value)
                    cell.font = self.styles.DATA_FONT
                    cell.border = self.styles.THIN_BORDER
                row += 1
            row += 1
        
        # 2.3 特征概览
        ws.cell(row=row, column=1, value="特征概览").font = self.styles.SUBHEADER_FONT
        ws.cell(row=row, column=1).fill = self.styles.SUBHEADER_FILL
        row += 1
        
        headers = ['指标', '值']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = self.styles.HEADER_FONT
            cell.fill = self.styles.HEADER_FILL
            cell.border = self.styles.THIN_BORDER
        row += 1
        
        # 原始特征数
        var_filter_result = data_loading_preview.get('var_filter_result', {})
        original_features = var_filter_result.get('input_features') or data_loading_preview.get('columns')
        ws.cell(row=row, column=1, value="原始特征数").font = self.styles.DATA_FONT
        ws.cell(row=row, column=1).border = self.styles.THIN_BORDER
        ws.cell(row=row, column=2, value=str(original_features) if original_features else '-').font = self.styles.DATA_FONT
        ws.cell(row=row, column=2).border = self.styles.THIN_BORDER
        row += 1
        
        # 异常值特征数
        outlier_count = data_loading_preview.get('outlier_count', '-')
        ws.cell(row=row, column=1, value="异常值特征数").font = self.styles.DATA_FONT
        ws.cell(row=row, column=1).border = self.styles.THIN_BORDER
        ws.cell(row=row, column=2, value=str(outlier_count)).font = self.styles.DATA_FONT
        ws.cell(row=row, column=2).border = self.styles.THIN_BORDER
        row += 1
        
        # 平均缺失率
        missing_rate = data_loading_preview.get('missing_rate')
        ws.cell(row=row, column=1, value="平均缺失率").font = self.styles.DATA_FONT
        ws.cell(row=row, column=1).border = self.styles.THIN_BORDER
        ws.cell(row=row, column=2, value=f"{missing_rate * 100:.1f}%" if missing_rate is not None else '-').font = self.styles.DATA_FONT
        ws.cell(row=row, column=2).border = self.styles.THIN_BORDER
        row += 1
        
        # 入模变量数（从 coefficients 获取）
        coefficients = results.get('coefficients', [])
        n_vars = len(coefficients) if coefficients else 0
        ws.cell(row=row, column=1, value="入模变量数").font = self.styles.DATA_FONT
        ws.cell(row=row, column=1).border = self.styles.THIN_BORDER
        ws.cell(row=row, column=2, value=str(n_vars)).font = self.styles.DATA_FONT
        ws.cell(row=row, column=2).border = self.styles.THIN_BORDER
        row += 1
        
        return row + 1
    
    def _write_scorecard_metrics_section(self, ws, results: dict, start_row: int) -> int:
        """
        写入评分卡模型评估指标（三、评估图表）
        与WORD报告对齐：按数据集分组展示，包含PSI稳定性分析和排序性分析
        """
        row = start_row
        
        metrics = results.get('metrics', {})
        multi_dataset_metrics = results.get('multi_dataset_metrics', {})
        multi_dataset_chart_data = results.get('multi_dataset_chart_data', {})
        psi_result = results.get('psi_result', {})
        psi_train_vs_test = results.get('psi_train_vs_test')
        psi_train_vs_oot = results.get('psi_train_vs_oot')
        stages = results.get('stages', {})
        
        # 从psi_result获取PSI数据（与WORD报告逻辑一致）
        if psi_train_vs_test is None and psi_result and isinstance(psi_result, dict) and psi_result.get('comparison') == '训练集 vs 测试集':
            psi_train_vs_test = psi_result
        if psi_train_vs_oot is None and psi_result and isinstance(psi_result, dict) and psi_result.get('comparison') == '训练集 vs OOT':
            psi_train_vs_oot = psi_result
        
        # 获取评分分布数据（用于单调性分析）
        model_eval_preview = stages.get('model_evaluation', {}).get('output_preview', {}) if stages else {}
        score_dist_from_stage = model_eval_preview.get('score_distribution', {})
        
        # 过拟合警告
        overfit_warning = results.get('overfit_warning')
        if overfit_warning and str(overfit_warning).strip() and overfit_warning != 'None':
            ws.cell(row=row, column=1, value="⚠️ 过拟合警告").font = Font(name='Arial', size=10, bold=True, color='DC2626')
            row += 1
            ws.cell(row=row, column=1, value=str(overfit_warning)).font = Font(name='Arial', size=10, color='DC2626')
            row += 2
        
        # 核心指标表格（按数据集对比）
        # 优先从 multi_dataset_metrics 获取各数据集指标
        train_metrics = multi_dataset_metrics.get('train', {}) if multi_dataset_metrics else {}
        test_metrics = multi_dataset_metrics.get('test', {}) if multi_dataset_metrics else {}
        oot_metrics = multi_dataset_metrics.get('oot', {}) if multi_dataset_metrics else {}
        
        ws.cell(row=row, column=1, value="核心性能指标").font = self.styles.SUBHEADER_FONT
        ws.cell(row=row, column=1).fill = self.styles.SUBHEADER_FILL
        row += 1
        
        headers = ['指标', '训练集', '测试集', 'OOT']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = self.styles.HEADER_FONT
            cell.fill = self.styles.HEADER_FILL
            cell.border = self.styles.THIN_BORDER
        row += 1
        
        # AUC - 优先从 multi_dataset_metrics 获取
        train_auc = train_metrics.get('auc') if train_metrics else metrics.get('train_auc', metrics.get('auc'))
        test_auc = test_metrics.get('auc') if test_metrics else metrics.get('test_auc')
        oot_auc = oot_metrics.get('auc') if oot_metrics else metrics.get('oot_auc')
        row_data = [
            'AUC',
            f"{train_auc:.4f}" if isinstance(train_auc, (int, float)) else '-',
            f"{test_auc:.4f}" if isinstance(test_auc, (int, float)) else '-',
            f"{oot_auc:.4f}" if isinstance(oot_auc, (int, float)) else '-',
        ]
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.font = self.styles.DATA_FONT
            cell.border = self.styles.THIN_BORDER
            # 高亮最佳性能
            if col > 1 and isinstance(row_data[col-1], str) and row_data[col-1] != '-':
                val = float(row_data[col-1])
                if val >= 0.8:
                    cell.font = Font(name='Arial', size=10, color='16A34A', bold=True)
        row += 1
        
        # KS - 优先从 multi_dataset_metrics 获取
        train_ks = train_metrics.get('ks') if train_metrics else metrics.get('train_ks', metrics.get('ks'))
        test_ks = test_metrics.get('ks') if test_metrics else metrics.get('test_ks')
        oot_ks = oot_metrics.get('ks') if oot_metrics else metrics.get('oot_ks')
        row_data = [
            'KS值',
            f"{train_ks*100:.2f}%" if isinstance(train_ks, (int, float)) else '-',
            f"{test_ks*100:.2f}%" if isinstance(test_ks, (int, float)) else '-',
            f"{oot_ks*100:.2f}%" if isinstance(oot_ks, (int, float)) else '-',
        ]
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.font = self.styles.DATA_FONT
            cell.border = self.styles.THIN_BORDER
        row += 1
        
        # Gini - 优先从 multi_dataset_metrics 获取
        train_gini = train_metrics.get('gini') if train_metrics else metrics.get('train_gini', metrics.get('gini'))
        test_gini = test_metrics.get('gini') if test_metrics else metrics.get('test_gini')
        oot_gini = oot_metrics.get('gini') if oot_metrics else metrics.get('oot_gini')
        row_data = [
            'Gini系数',
            f"{train_gini*100:.2f}%" if isinstance(train_gini, (int, float)) else '-',
            f"{test_gini*100:.2f}%" if isinstance(test_gini, (int, float)) else '-',
            f"{oot_gini*100:.2f}%" if isinstance(oot_gini, (int, float)) else '-',
        ]
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.font = self.styles.DATA_FONT
            cell.border = self.styles.THIN_BORDER
        row += 2
        
        # 排序性分析（与Word报告对齐）
        # 按数据集展示排序性分析结果
        datasets_to_show = []
        has_oot = bool(multi_dataset_metrics.get('oot') or multi_dataset_chart_data.get('oot'))
        
        if has_oot:
            datasets_to_show.append(('oot', 'OOT验证集'))
        datasets_to_show.append(('test', '测试集'))
        datasets_to_show.append(('train', '训练集'))
        
        for dataset_key, dataset_label in datasets_to_show:
            ds_chart_data = (multi_dataset_chart_data.get(dataset_key) or {}) if multi_dataset_chart_data else {}
            
            # 回退：从stage获取评分分布数据
            if not ds_chart_data.get('score_distribution') and score_dist_from_stage:
                stage_score_dist = score_dist_from_stage.get(dataset_key, {})
                if stage_score_dist:
                    ds_chart_data = {'score_distribution': stage_score_dist}
            
            score_dist = ds_chart_data.get('score_distribution') or {}
            ranking_bins = (score_dist.get('ranking_analysis') or {}).get('bins') or score_dist.get('bins')
            
            if ranking_bins and len(ranking_bins) > 0:
                ws.cell(row=row, column=1, value=f"{dataset_label} - 排序性分析").font = self.styles.SUBHEADER_FONT
                ws.cell(row=row, column=1).fill = self.styles.SUBHEADER_FILL
                row += 1
                
                # 排序性分析表格
                headers = ['评分区间', '样本数', '占比', '坏样本数', '坏样本率', 'Lift']
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=row, column=col, value=header)
                    cell.font = self.styles.HEADER_FONT
                    cell.fill = self.styles.HEADER_FILL
                    cell.border = self.styles.THIN_BORDER
                row += 1
                
                for bin_data in ranking_bins:
                    row_data = [
                        str(bin_data.get('bin', bin_data.get('score_range', '-'))),
                        f"{bin_data.get('count', bin_data.get('total', 0)):,}",
                        f"{bin_data.get('pct', bin_data.get('percent', 0)):.1f}%",
                        f"{bin_data.get('bad', bin_data.get('bad_count', 0)):,}",
                        f"{bin_data.get('bad_rate', 0):.2f}%",
                        f"{bin_data.get('lift', 0):.2f}",
                    ]
                    for col, value in enumerate(row_data, 1):
                        cell = ws.cell(row=row, column=col, value=value)
                        cell.font = self.styles.DATA_FONT
                        cell.border = self.styles.THIN_BORDER
                    row += 1
                
                # 单调性分析
                rank_analysis = score_dist.get('rank_ordering_analysis', {})
                monotonicity = rank_analysis.get('monotonicity', {})
                if monotonicity:
                    mono_pass = monotonicity.get('is_monotonic', False)
                    mono_violations = monotonicity.get('violations', 0)
                    first_lift = ranking_bins[0].get('lift') if ranking_bins else None
                    last_lift = ranking_bins[-1].get('lift') if ranking_bins else None
                    
                    row += 1
                    mono_text = f"单调性：{'✓ 通过' if mono_pass else f'不通过（{mono_violations}处违反）'}"
                    ws.cell(row=row, column=1, value=mono_text).font = Font(
                        name='Arial', size=9, 
                        color='16A34A' if mono_pass else 'DC2626'
                    )
                    row += 1
                    
                    first_str = f"{first_lift:.2f}" if first_lift is not None else "-"
                    last_str = f"{last_lift:.2f}" if last_lift is not None else "-"
                    lift_text = f"首组Lift：{first_str}  |  末组Lift：{last_str}"
                    ws.cell(row=row, column=1, value=lift_text).font = Font(name='Arial', size=9, color='666666')
                    row += 2
        
        # PSI稳定性指标
        ws.cell(row=row, column=1, value="PSI稳定性分析").font = self.styles.SUBHEADER_FONT
        ws.cell(row=row, column=1).fill = self.styles.SUBHEADER_FILL
        row += 1
        
        # PSI指标说明
        ws.cell(row=row, column=1, value="PSI < 0.1 稳定 | 0.1-0.25 轻微变化 | ≥0.25 显著变化").font = Font(name='Arial', size=9, italic=True, color='666666')
        row += 1
        
        # PSI表格
        psi_headers = ['对比数据集', 'PSI值', '稳定性评估']
        for col, header in enumerate(psi_headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = self.styles.HEADER_FONT
            cell.fill = self.styles.HEADER_FILL
            cell.border = self.styles.THIN_BORDER
        row += 1
        
        def get_psi_level(psi):
            if psi is None:
                return '-'
            if psi < 0.1:
                return '稳定'
            elif psi < 0.25:
                return '轻微变化'
            else:
                return '显著变化'
        
        # 训练集 vs 测试集
        if psi_train_vs_test and isinstance(psi_train_vs_test, dict):
            psi_value = psi_train_vs_test.get('value')
            row_data = [
                '训练集 vs 测试集',
                f"{psi_value:.4f}" if psi_value is not None else '-',
                get_psi_level(psi_value)
            ]
            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.font = self.styles.DATA_FONT
                cell.border = self.styles.THIN_BORDER
                if col == 3:
                    if value == '稳定':
                        cell.font = Font(name='Arial', size=10, color='16A34A')
                    elif value == '显著变化':
                        cell.font = Font(name='Arial', size=10, color='DC2626')
            row += 1
        
        # 训练集 vs OOT
        if psi_train_vs_oot and isinstance(psi_train_vs_oot, dict):
            psi_value = psi_train_vs_oot.get('value')
            row_data = [
                '训练集 vs OOT',
                f"{psi_value:.4f}" if psi_value is not None else '-',
                get_psi_level(psi_value)
            ]
            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.font = self.styles.DATA_FONT
                cell.border = self.styles.THIN_BORDER
                if col == 3:
                    if value == '稳定':
                        cell.font = Font(name='Arial', size=10, color='16A34A')
                    elif value == '显著变化':
                        cell.font = Font(name='Arial', size=10, color='DC2626')
            row += 1
        
        # 如果没有PSI数据，显示暂无数据
        if not psi_train_vs_test and not psi_train_vs_oot:
            ws.cell(row=row, column=1, value="暂无PSI数据").font = self.styles.DATA_FONT
            row += 1
        
        row += 1
        return row
    
    def _write_scorecard_feature_selection_section(self, ws, results: dict, start_row: int) -> int:
        """
        写入特征筛选详情（与Word/HTML报告对齐）
        包含：
        1. 特征筛选漏斗
        2. 变量IV排行（带状态和淘汰信息）
        """
        row = start_row
        
        stages = results.get('stages', {})
        selected_features = results.get('selected_features', [])
        selection_detail = results.get('selection_detail', {})
        
        # 5.1 特征筛选漏斗
        if stages:
            ws.cell(row=row, column=1, value="特征筛选漏斗").font = self.styles.SUBHEADER_FONT
            ws.cell(row=row, column=1).fill = self.styles.SUBHEADER_FILL
            row += 1
            
            # 获取各阶段数据
            data_loading_preview = stages.get('data_loading', {}).get('output_preview', {}) if stages else {}
            woe_binning_preview = stages.get('woe_binning', {}).get('output_preview', {}) if stages else {}
            feature_selection_preview = stages.get('feature_selection', {}).get('output_preview', {}) if stages else {}
            model_training_preview = stages.get('model_training', {}).get('output_preview', {}) if stages else {}
            
            var_filter_result = data_loading_preview.get('var_filter_result', {})
            
            # 阶段1: 原始特征数
            original_count = var_filter_result.get('input_features') or data_loading_preview.get('feature_count') or data_loading_preview.get('columns', 0)
            if isinstance(original_count, list):
                original_count = len(original_count)
            
            # 阶段2: 质量筛选后
            after_var_filter = var_filter_result.get('output_features', 0)
            if isinstance(after_var_filter, list):
                after_var_filter = len(after_var_filter)
            
            # 阶段3: WOE分箱后
            woe_output = woe_binning_preview.get('total_features', 0)
            if isinstance(woe_output, list):
                woe_output = len(woe_output)
            elif woe_output == 0:
                woe_results = woe_binning_preview.get('woe_results', {})
                if isinstance(woe_results, dict):
                    woe_output = len([k for k in woe_results.keys() if not k.startswith('_')])
                else:
                    bins = woe_binning_preview.get('bins', {})
                    if isinstance(bins, dict):
                        woe_output = len(bins)
            
            # 阶段4: 特征筛选后
            fe_after = feature_selection_preview.get('after_count') or feature_selection_preview.get('selected_count') or 0
            
            # 阶段5: 最终入模特征数
            coefficients = results.get('coefficients', [])
            if isinstance(coefficients, list):
                final_count = len([c for c in coefficients if c.get('feature') != 'intercept'])
            else:
                mt_coefficients = model_training_preview.get('coefficients', []) or model_training_preview.get('all_coefficients', [])
                final_count = len([c for c in mt_coefficients if c.get('feature') != 'intercept']) if mt_coefficients else 0
            
            # 创建漏斗指标卡
            if original_count > 0:
                funnel_steps = [
                    ("原始特征", original_count),
                    ("质量筛选", after_var_filter),
                    ("WOE分箱", woe_output),
                    ("IV/相关/VIF", fe_after),
                    ("最终入模", final_count),
                ]
                
                # 表头
                headers = [step[0] for step in funnel_steps]
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=row, column=col, value=header)
                    cell.font = Font(name='Arial', size=9, color='6B7280')
                    cell.alignment = self.styles.CENTER
                row += 1
                
                # 数量
                for col, (label, count) in enumerate(funnel_steps, 1):
                    display_count = count if count is not None else 0
                    cell = ws.cell(row=row, column=col, value=f"{display_count}")
                    cell.font = Font(name='Arial', size=12, bold=True)
                    cell.alignment = self.styles.CENTER
                row += 1
                
                # 百分比
                for col, (label, count) in enumerate(funnel_steps, 1):
                    display_count = count if count is not None else 0
                    pct = f"{display_count/original_count*100:.0f}%" if original_count > 0 else "0%"
                    cell = ws.cell(row=row, column=col, value=pct)
                    cell.font = Font(name='Arial', size=8, color='6B7280')
                    cell.alignment = self.styles.CENTER
                row += 2
        
        # 5.2 变量IV排行（带状态和淘汰信息）
        iv_table = results.get('iv_table', [])
        if iv_table and len(iv_table) > 0:
            ws.cell(row=row, column=1, value="变量IV排行").font = self.styles.SUBHEADER_FONT
            ws.cell(row=row, column=1).fill = self.styles.SUBHEADER_FILL
            row += 1
            
            # 构建淘汰信息映射
            elimination_map = {}
            
            # 从var_filter获取淘汰信息
            removed_by_missing = var_filter_result.get('removed_by_missing', [])
            for item in removed_by_missing:
                if isinstance(item, dict) and item.get('feature'):
                    elimination_map[item['feature']] = {
                        'stage': '数据质量(var_filter)',
                        'reason': item.get('reason', f"缺失率{(item.get('missing_rate', 0) * 100):.0f}%")
                    }
            removed_by_identical = var_filter_result.get('removed_by_identical', [])
            for item in removed_by_identical:
                if isinstance(item, dict) and item.get('feature'):
                    elimination_map[item['feature']] = {
                        'stage': '数据质量(var_filter)',
                        'reason': item.get('reason', f"同值率{(item.get('identical_rate', 0) * 100):.0f}%")
                    }
            
            # 从WOE阶段获取
            woe_filtered = woe_binning_preview.get('woe_filtered', {})
            woe_filtered_features = woe_filtered.get('features', [])
            for feat in woe_filtered_features:
                if feat not in elimination_map:
                    elimination_map[feat] = {
                        'stage': 'WOE分箱',
                        'reason': woe_filtered.get('reason', '常量/分箱失败')
                    }
            
            # 从feature_selection阶段获取
            all_features_detail = feature_selection_preview.get('all_features_detail', [])
            for item in all_features_detail:
                if isinstance(item, dict) and item.get('feature') and item.get('remove_reason'):
                    feat = item['feature']
                    reason = item['remove_reason']
                    stage = '特征筛选(IV)' if 'IV' in reason else '特征筛选(相关性)' if '相关性' in reason else '特征筛选(VIF)' if 'VIF' in reason else '特征筛选'
                    elimination_map[feat] = {'stage': stage, 'reason': reason}
                    elimination_map[feat + '_woe'] = {'stage': stage, 'reason': reason}
            
            # 从model_training阶段获取
            stepwise_result_data = model_training_preview.get('stepwise_result', {})
            stepwise_steps = stepwise_result_data.get('steps', [])
            for s in stepwise_steps:
                if s.get('action') == 'remove' and s.get('feature'):
                    feat = s['feature']
                    base_name = feat.replace('_woe', '')
                    reason = f"P值={s.get('pvalue', 0):.4f}"
                    elimination_map[feat] = {'stage': '模型训练(逐步回归)', 'reason': reason}
                    elimination_map[base_name] = {'stage': '模型训练(逐步回归)', 'reason': reason}
            
            # 从系数验证获取
            coef_validation = model_training_preview.get('coefficient_validation', {})
            for feat in coef_validation.get('invalid_direction', []):
                base_name = feat.replace('_woe', '')
                if base_name not in elimination_map and feat not in elimination_map:
                    elimination_map[feat] = {'stage': '模型训练(系数验证)', 'reason': '系数方向异常'}
                    elimination_map[base_name] = {'stage': '模型训练(系数验证)', 'reason': '系数方向异常'}
            
            # 入模特征集合
            model_features_set = set(f.replace('_woe', '') for f in (selected_features or []))
            
            # 表头（7列）
            headers = ['序号', '变量', 'IV值', '预测能力', '状态', '淘汰阶段', '淘汰原因']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.font = self.styles.HEADER_FONT
                cell.fill = self.styles.HEADER_FILL
                cell.border = self.styles.THIN_BORDER
            row += 1
            
            # 按IV值降序排列
            if isinstance(iv_table, list):
                iv_list = sorted(iv_table, key=lambda x: x.get('iv', x.get('IV', 0)), reverse=True)
            else:
                iv_list = []
            
            # 填充数据
            for idx, item in enumerate(iv_list, 1):
                var_name = item.get('variable', item.get('Variable', ''))
                iv_val = item.get('iv', item.get('IV', 0))
                
                # 预测能力评级
                if isinstance(iv_val, (int, float)):
                    if iv_val >= 0.5:
                        power = "过高"
                    elif iv_val >= 0.3:
                        power = "强"
                    elif iv_val >= 0.1:
                        power = "中等"
                    elif iv_val >= 0.02:
                        power = "弱"
                    else:
                        power = "无"
                else:
                    power = "-"
                
                # 状态判断
                base_name = var_name.replace('_woe', '')
                is_model_var = base_name in model_features_set
                elim_info = elimination_map.get(base_name) or elimination_map.get(var_name)
                
                if is_model_var:
                    status = "入模"
                    elim_stage = "-"
                    elim_reason = "-"
                elif elim_info:
                    status = "淘汰"
                    elim_stage = elim_info.get('stage', '-')
                    elim_reason = elim_info.get('reason', '-')
                else:
                    status = "-"
                    elim_stage = "-"
                    elim_reason = "-"
                
                row_data = [
                    idx,
                    var_name,
                    f"{iv_val:.4f}" if isinstance(iv_val, (int, float)) else str(iv_val),
                    power,
                    status,
                    elim_stage,
                    elim_reason
                ]
                
                for col, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row, column=col, value=value)
                    cell.font = self.styles.DATA_FONT
                    cell.border = self.styles.THIN_BORDER
                    # 状态列着色
                    if col == 5:
                        if status == "入模":
                            cell.font = Font(name='Arial', size=10, color='16A34A', bold=True)
                        elif status == "淘汰":
                            cell.font = Font(name='Arial', size=10, color='DC2626', bold=True)
                row += 1
            row += 1
        
        if not stages and not iv_table:
            ws.cell(row=row, column=1, value="暂无特征筛选数据").font = self.styles.DATA_FONT
            row += 1
        
        return row
    
    def _write_scorecard_details_section(self, ws, results: dict, start_row: int) -> int:
        """
        写入评分卡明细（与Word/HTML报告对齐）
        包含：
        1. 评分卡核心参数指标卡
        2. 入模变量评分贡献
        3. 完整评分卡表格
        """
        import logging
        logger = logging.getLogger(__name__)
        
        row = start_row
        
        # 获取stages数据
        stages = results.get('stages', {})
        score_scaling = stages.get('score_scaling', {}) if stages else {}
        score_scaling_preview = score_scaling.get('output_preview', {}) if isinstance(score_scaling, dict) else {}
        
        # 4.1 评分卡核心参数（指标卡形式）
        num_vars = score_scaling_preview.get('num_variables', 0)
        theoretical_range = score_scaling_preview.get('theoretical_score_range', {})
        base_score = score_scaling_preview.get('base_score', 600)
        pdo = score_scaling_preview.get('pdo', 50)
        target_odds = score_scaling_preview.get('target_odds', 20)
        
        # 获取实际评分统计数据
        actual_stats = None
        stats_dataset_label = ""
        multi_dataset_chart_data = results.get('multi_dataset_chart_data', {})
        
        if multi_dataset_chart_data:
            for ds_key, ds_label in [('oot', 'OOT验证集'), ('test', '测试集'), ('train', '训练集')]:
                ds_chart = multi_dataset_chart_data.get(ds_key) or {}
                score_dist = ds_chart.get('score_distribution') or {}
                summary = score_dist.get('summary') or {}
                if summary and (summary.get('good_mean') is not None or summary.get('bad_mean') is not None):
                    actual_stats = summary
                    stats_dataset_label = ds_label
                    break
        
        # 回退获取方式
        if not actual_stats:
            score_stats_by_dataset = score_scaling_preview.get('score_stats_by_dataset', {})
            if score_stats_by_dataset:
                for ds_key, ds_label in [('oot', 'OOT验证集'), ('test', '测试集'), ('train', '训练集')]:
                    stats = score_stats_by_dataset.get(ds_key)
                    if stats and (stats.get('good_mean') is not None or stats.get('bad_mean') is not None):
                        actual_stats = stats
                        stats_dataset_label = ds_label
                        break
        
        if not actual_stats:
            actual_stats = score_scaling_preview.get('actual_score_stats')
            if actual_stats:
                stats_dataset_label = "训练集"
        
        if num_vars or theoretical_range or actual_stats:
            ws.cell(row=row, column=1, value="评分卡核心参数").font = self.styles.SUBHEADER_FONT
            ws.cell(row=row, column=1).fill = self.styles.SUBHEADER_FILL
            row += 1
            
            # 计算分离度
            good_mean = actual_stats.get('good_mean') if actual_stats else None
            bad_mean = actual_stats.get('bad_mean') if actual_stats else None
            separation = abs(good_mean - bad_mean) if good_mean is not None and bad_mean is not None else None
            
            min_score = theoretical_range.get('min', 0)
            max_score = theoretical_range.get('max', 0)
            
            # 创建2x3指标卡布局
            metrics_data = [
                ["入模变量数", f"{num_vars} 个", "评分区间（理论）", f"{min_score:.0f} ~ {max_score:.0f}"],
                ["基准配置", f"{base_score:.0f}/{pdo:.0f}/{target_odds:.0f}", "好样本均分", f"{good_mean:.1f}" if good_mean is not None else "-"],
                ["坏样本均分", f"{bad_mean:.1f}" if bad_mean is not None else "-", "分离度", f"{separation:.1f}" if separation is not None else "-"],
            ]
            
            for metric_row in metrics_data:
                ws.cell(row=row, column=1, value=metric_row[0]).font = Font(name='Arial', size=9, color='6B7280')
                ws.cell(row=row, column=2, value=metric_row[1]).font = Font(name='Arial', size=11, bold=True)
                ws.cell(row=row, column=3, value=metric_row[2]).font = Font(name='Arial', size=9, color='6B7280')
                ws.cell(row=row, column=4, value=metric_row[3]).font = Font(name='Arial', size=11, bold=True)
                row += 1
            
            if stats_dataset_label:
                ws.cell(row=row, column=1, value=f"* 数据统计来源: {stats_dataset_label}").font = Font(name='Arial', size=8, italic=True, color='9CA3AF')
                row += 1
            row += 1
        
        # 4.2 入模变量评分贡献
        scorecard_preview = score_scaling_preview.get('scorecard_preview', [])
        if scorecard_preview and len(scorecard_preview) > 0:
            valid_vars = [v for v in scorecard_preview if v.get('variable') not in ('basepoints', '常数项')]
            
            if valid_vars:
                ws.cell(row=row, column=1, value="入模变量评分贡献").font = self.styles.SUBHEADER_FONT
                ws.cell(row=row, column=1).fill = self.styles.SUBHEADER_FILL
                row += 1
                
                # 计算波动幅度并排序
                for v in valid_vars:
                    min_score = v.get('min_score', 0) or 0
                    max_score = v.get('max_score', 0) or 0
                    v['_score_range'] = abs(max_score - min_score)
                
                sorted_vars = sorted(valid_vars, key=lambda x: x.get('_score_range', 0), reverse=True)[:10]
                
                # 表头
                headers = ['变量', '最低分', '最高分', '波动幅度']
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=row, column=col, value=header)
                    cell.font = self.styles.HEADER_FONT
                    cell.fill = self.styles.HEADER_FILL
                    cell.border = self.styles.THIN_BORDER
                row += 1
                
                # 写入数据
                for v in sorted_vars:
                    var_name = v.get('variable', '')
                    min_score = v.get('min_score', 0) or 0
                    max_score = v.get('max_score', 0) or 0
                    score_range = v.get('_score_range', 0)
                    
                    row_data = [
                        var_name[:30] + '...' if len(var_name) > 30 else var_name,
                        f"{min_score:.0f}",
                        f"{max_score:.0f}",
                        f"{score_range:.0f}分"
                    ]
                    for col, value in enumerate(row_data, 1):
                        cell = ws.cell(row=row, column=col, value=value)
                        cell.font = self.styles.DATA_FONT
                        cell.border = self.styles.THIN_BORDER
                    row += 1
                
                ws.cell(row=row, column=1, value="* 波动幅度 = 最高分 - 最低分，反映变量对评分的影响程度").font = Font(name='Arial', size=8, italic=True, color='6B7280')
                row += 2
        
        # 4.3 完整评分卡表格
        full_scorecard_csv = score_scaling_preview.get('full_scorecard_csv', [])
        scorecard = results.get('scorecard')
        
        if full_scorecard_csv or scorecard:
            ws.cell(row=row, column=1, value="完整评分卡").font = self.styles.SUBHEADER_FONT
            ws.cell(row=row, column=1).fill = self.styles.SUBHEADER_FILL
            row += 1
            
            # 使用 full_scorecard_csv 优先
            if full_scorecard_csv and len(full_scorecard_csv) > 0:
                # 表头 - 包含完整字段
                headers = ['变量', 'IV', '系数', '分箱', '样本数', '占比', '坏样本数', '坏样本率', '分数']
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=row, column=col, value=header)
                    cell.font = self.styles.HEADER_FONT
                    cell.fill = self.styles.HEADER_FILL
                    cell.border = self.styles.THIN_BORDER
                row += 1
                
                # 预计算每个变量的行数（用于判断是否第一行）
                variable_row_counts = {}
                for item in full_scorecard_csv:
                    var_name = item.get('variable', '')
                    variable_row_counts[var_name] = variable_row_counts.get(var_name, 0) + 1
                
                current_var = None
                remaining_rows = 0
                
                for item in full_scorecard_csv[:50]:  # 最多50行
                    var_name = item.get('variable', '')
                    
                    if var_name != current_var:
                        current_var = var_name
                        remaining_rows = variable_row_counts.get(var_name, 1)
                        is_first_row = True
                    else:
                        is_first_row = False
                    
                    # 变量名、IV、系数（只在第一行显示）
                    if is_first_row:
                        var_display = var_name
                        iv = item.get('total_iv', item.get('iv', 0))
                        iv_display = f"{iv:.4f}" if isinstance(iv, (int, float)) else str(iv)
                        coef = item.get('cof', item.get('coef', item.get('coefficient', 0)))
                        coef_display = f"{coef:.4f}" if isinstance(coef, (int, float)) else str(coef)
                    else:
                        var_display = ""
                        iv_display = ""
                        coef_display = ""
                    
                    # 其他字段
                    bin_val = item.get('bin', '-')
                    total_samples = item.get('total_samples', item.get('count', 0))
                    percentage = item.get('percentage', 0)
                    bad_count = item.get('bad_count', item.get('bad', 0))
                    bad_rate = item.get('bad_rate', 0)
                    points = item.get('points', item.get('score', 0))
                    
                    row_data = [
                        var_display,
                        iv_display,
                        coef_display,
                        str(bin_val),
                        f"{int(total_samples):,}" if total_samples else "-",
                        f"{percentage:.2f}%" if isinstance(percentage, (int, float)) else str(percentage),
                        f"{int(bad_count):,}" if bad_count else "-",
                        f"{bad_rate:.2f}%" if isinstance(bad_rate, (int, float)) else str(bad_rate),
                        f"{points:.0f}" if isinstance(points, (int, float)) else str(points)
                    ]
                    
                    for col, value in enumerate(row_data, 1):
                        cell = ws.cell(row=row, column=col, value=value)
                        cell.font = self.styles.DATA_FONT
                        cell.border = self.styles.THIN_BORDER
                    row += 1
                
                if len(full_scorecard_csv) > 50:
                    ws.cell(row=row, column=1, value=f"（仅显示前50条，共{len(full_scorecard_csv)}条）").font = Font(name='Arial', size=9, italic=True)
                    row += 1
            else:
                # 简化版评分卡表格（后备方案）
                headers = ['变量', '分箱', 'WOE值', '评分']
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=row, column=col, value=header)
                    cell.font = self.styles.HEADER_FONT
                    cell.fill = self.styles.HEADER_FILL
                    cell.border = self.styles.THIN_BORDER
                row += 1
                
                items = []
                if isinstance(scorecard, dict):
                    for variable, var_data in scorecard.items():
                        bins = []
                        if isinstance(var_data, dict) and 'data' in var_data:
                            bins = var_data['data']
                        elif isinstance(var_data, list):
                            bins = var_data
                        
                        for bin_item in bins:
                            items.append({
                                'variable': variable,
                                'bin': bin_item.get('bin', bin_item.get('Bin', '-')),
                                'woe': bin_item.get('woe', bin_item.get('WOE', 0)),
                                'points': bin_item.get('points', bin_item.get('Points', 0)),
                            })
                elif isinstance(scorecard, list):
                    items = scorecard
                
                for item in items[:50]:
                    row_data = [
                        item.get('variable', '-'),
                        item.get('bin', '-'),
                        f"{item.get('woe', 0):.4f}" if isinstance(item.get('woe'), (int, float)) else str(item.get('woe', '-')),
                        f"{item.get('points', 0):.0f}" if isinstance(item.get('points'), (int, float)) else str(item.get('points', '-')),
                    ]
                    for col, value in enumerate(row_data, 1):
                        cell = ws.cell(row=row, column=col, value=value)
                        cell.font = self.styles.DATA_FONT
                        cell.border = self.styles.THIN_BORDER
                    row += 1
        else:
            ws.cell(row=row, column=1, value="暂无评分卡数据").font = self.styles.DATA_FONT
            row += 1
        
        return row + 1
    
    def _write_iv_table_section(self, ws, results: dict, start_row: int) -> int:
        """写入IV值表格（对应前端 iv Tab）"""
        row = start_row
        
        iv_table = results.get('iv_table', [])
        if not iv_table:
            ws.cell(row=row, column=1, value="暂无IV数据").font = self.styles.DATA_FONT
            return row + 1
        
        # 表头
        headers = ['排名', '变量', 'IV值', '预测能力']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = self.styles.HEADER_FONT
            cell.fill = self.styles.HEADER_FILL
            cell.border = self.styles.THIN_BORDER
        row += 1
        
        # 处理数据
        items = iv_table if isinstance(iv_table, list) else []
        for i, item in enumerate(items[:20], 1):
            var = item.get('variable', item.get('feature', '-'))
            iv = item.get('iv', item.get('IV', 0))
            
            # 判断预测能力
            if isinstance(iv, (int, float)):
                if iv >= 0.5:
                    power = "极强"
                elif iv >= 0.3:
                    power = "强"
                elif iv >= 0.1:
                    power = "中等"
                elif iv >= 0.02:
                    power = "弱"
                else:
                    power = "极弱"
            else:
                power = "-"
            
            row_data = [
                i,
                var,
                f"{iv:.4f}" if isinstance(iv, (int, float)) else str(iv),
                power,
            ]
            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.font = self.styles.DATA_FONT
                cell.border = self.styles.THIN_BORDER
            row += 1
        
        if len(items) > 20:
            ws.cell(row=row, column=1, value=f"（仅显示前20个，共{len(items)}个）").font = Font(name='Arial', size=9, italic=True)
            row += 1
        
        row += 1
        
        return row
    
    def _write_coefficients_section(self, ws, results: dict, start_row: int) -> int:
        """写入模型系数（对应前端 coefficients Tab）"""
        row = start_row
        
        coefficients = results.get('coefficients', [])
        if not coefficients:
            ws.cell(row=row, column=1, value="暂无系数数据").font = self.styles.DATA_FONT
            return row + 1
        
        # 表头
        headers = ['变量', '系数', '标准误', 'z值', 'P值', '显著性']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = self.styles.HEADER_FONT
            cell.fill = self.styles.HEADER_FILL
            cell.border = self.styles.THIN_BORDER
        row += 1
        
        items = coefficients if isinstance(coefficients, list) else []
        for item in items:
            var = item.get('feature', item.get('variable', '-'))
            coef = item.get('coef', item.get('coefficient', 0))
            std_err = item.get('std_err', item.get('std_error'))
            z_val = item.get('z', item.get('z_value'))
            p_val = item.get('p_value', item.get('pvalue'))
            sig = item.get('significance', '')
            
            # 自动计算显著性
            if not sig and isinstance(p_val, (int, float)):
                if p_val < 0.001:
                    sig = '***'
                elif p_val < 0.01:
                    sig = '**'
                elif p_val < 0.05:
                    sig = '*'
                elif p_val < 0.1:
                    sig = '.'
            
            row_data = [
                str(var).replace('_woe', ''),
                f"{coef:.4f}" if isinstance(coef, (int, float)) else str(coef),
                f"{std_err:.4f}" if isinstance(std_err, (int, float)) else '-',
                f"{z_val:.4f}" if isinstance(z_val, (int, float)) else '-',
                f"<0.001" if isinstance(p_val, (int, float)) and p_val < 0.001 else (f"{p_val:.4f}" if isinstance(p_val, (int, float)) else '-'),
                sig,
            ]
            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.font = self.styles.DATA_FONT
                cell.border = self.styles.THIN_BORDER
            row += 1
        
        row += 1
        ws.cell(row=row, column=1, value="显著性标记：*** p<0.001, ** p<0.01, * p<0.05, . p<0.1").font = Font(name='Arial', size=9, italic=True)
        row += 2
        
        return row
    
    def _write_scorecard_model_coefficients_section(self, ws, results: dict, start_row: int) -> int:
        """
        写入评分卡模型系数章节（与Word/HTML报告对齐）
        包含：
        1. 模型概览指标卡（似然比检验、显著变量、系数方向、截距项）
        2. 模型拟合指标（Pseudo R²、Log-Likelihood、AIC、BIC）
        3. 系数统计表（完整字段：变量、系数、标准误、z值、P值、95%CI、显著性）
        """
        row = start_row
        
        stages = results.get('stages', {})
        coefficients = results.get('coefficients', [])
        model_statistics = results.get('model_statistics', {})
        selection_detail = results.get('selection_detail', {})
        model_training_preview = stages.get('model_training', {}).get('output_preview', {}) if stages else {}
        
        # 6.1 模型概览指标卡
        if coefficients or model_statistics:
            ws.cell(row=row, column=1, value="模型概览").font = self.styles.SUBHEADER_FONT
            ws.cell(row=row, column=1).fill = self.styles.SUBHEADER_FILL
            row += 1
            
            # 获取统计数据源
            stats_summary = model_statistics.get('summary', []) if model_statistics else []
            if stats_summary and isinstance(stats_summary, list) and len(stats_summary) > 0:
                coef_list = stats_summary
            else:
                preview_coefficients = model_training_preview.get('coefficients', [])
                if preview_coefficients and isinstance(preview_coefficients, list):
                    coef_list = preview_coefficients
                else:
                    coef_list = coefficients if isinstance(coefficients, list) else []
            
            # 计算指标
            n_features = sum(1 for item in coef_list 
                            if item.get('feature', item.get('variable', '')) not in ('const', '常数项', 'intercept'))
            
            significant_count = 0
            for item in coef_list:
                feature = item.get('feature', item.get('variable', ''))
                p_val = item.get('p_value', item.get('pvalue'))
                if feature not in ('const', '常数项', 'intercept') and p_val is not None and isinstance(p_val, (int, float)) and p_val < 0.05:
                    significant_count += 1
            
            # 系数方向验证
            coef_validation = model_training_preview.get('coefficient_validation', {})
            if not coef_validation and selection_detail:
                coef_validation = selection_detail.get('coefficient_validation', {})
            valid_direction = coef_validation.get('valid_direction', [])
            invalid_direction = coef_validation.get('invalid_direction', [])
            
            # 截距项
            intercept = model_training_preview.get('intercept') or results.get('intercept') or (model_statistics.get('intercept') if model_statistics else None)
            
            # 似然比检验
            lr_pvalue = model_statistics.get('lr_pvalue') if model_statistics else None
            
            # 创建1行4列指标卡
            metrics_labels = ['似然比检验', '显著变量 (p<0.05)', '系数方向', '截距项']
            metrics_values = []
            
            # 似然比检验
            if lr_pvalue is not None and isinstance(lr_pvalue, (int, float)):
                lr_significant = lr_pvalue < 0.05
                lr_p_str = '<0.001' if lr_pvalue < 0.001 else f"{lr_pvalue:.4f}"
                metrics_values.append(f"{lr_p_str}\n{'✓ 显著' if lr_significant else '不显著'}")
            else:
                metrics_values.append("-")
            
            # 显著变量
            metrics_values.append(f"{significant_count}\n/{n_features}个")
            
            # 系数方向
            total_direction = len(valid_direction) + len(invalid_direction)
            if total_direction > 0:
                direction_str = f"{len(valid_direction)}/{total_direction}"
                if len(invalid_direction) == 0:
                    direction_str += "\n✓ 全部正确"
                else:
                    direction_str += f"\n⚠ {len(invalid_direction)}个异常"
                metrics_values.append(direction_str)
            else:
                metrics_values.append("-")
            
            # 截距项
            intercept_str = f"{intercept:.4f}" if isinstance(intercept, (int, float)) else '-'
            metrics_values.append(intercept_str)
            
            # 写入指标卡（标签行）
            for col, label in enumerate(metrics_labels, 1):
                cell = ws.cell(row=row, column=col, value=label)
                cell.font = Font(name='Arial', size=9, color='6B7280')
                cell.alignment = self.styles.CENTER
            row += 1
            
            # 写入指标卡（数值行）
            for col, value in enumerate(metrics_values, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.font = Font(name='Arial', size=11, bold=True)
                cell.alignment = self.styles.CENTER
            row += 2
        
        # 6.2 模型拟合指标
        if model_statistics:
            pseudo_r2 = model_statistics.get('pseudo_r2')
            log_likelihood = model_statistics.get('log_likelihood')
            aic = model_statistics.get('aic')
            bic = model_statistics.get('bic')
            
            if any([pseudo_r2 is not None, log_likelihood is not None, aic is not None, bic is not None]):
                ws.cell(row=row, column=1, value="模型拟合指标").font = self.styles.SUBHEADER_FONT
                ws.cell(row=row, column=1).fill = self.styles.SUBHEADER_FILL
                row += 1
                
                # 创建1行4列指标卡
                fit_labels = ['Pseudo R²', 'Log-Likelihood', 'AIC', 'BIC']
                fit_values = []
                
                fit_values.append(f"{pseudo_r2:.4f}" if pseudo_r2 is not None else '-')
                fit_values.append(f"{log_likelihood:.4f}" if log_likelihood is not None else '-')
                fit_values.append(f"{aic:.2f}" if aic is not None else '-')
                fit_values.append(f"{bic:.2f}" if bic is not None else '-')
                
                # 写入标签行
                for col, label in enumerate(fit_labels, 1):
                    cell = ws.cell(row=row, column=col, value=label)
                    cell.font = Font(name='Arial', size=9, color='6B7280')
                    cell.alignment = self.styles.CENTER
                row += 1
                
                # 写入数值行
                for col, value in enumerate(fit_values, 1):
                    cell = ws.cell(row=row, column=col, value=value)
                    cell.font = Font(name='Arial', size=11, bold=True)
                    cell.alignment = self.styles.CENTER
                row += 2
        
        # 6.3 系数统计表
        ws.cell(row=row, column=1, value="系数统计").font = self.styles.SUBHEADER_FONT
        ws.cell(row=row, column=1).fill = self.styles.SUBHEADER_FILL
        row += 1
        
        # 优先使用 model_statistics.summary 作为数据源
        stats_summary = model_statistics.get('summary', []) if model_statistics else []
        has_valid_z = stats_summary and isinstance(stats_summary, list) and len(stats_summary) > 0 and \
                      any(item.get('z') is not None for item in stats_summary if isinstance(item, dict))
        
        if has_valid_z:
            coef_data_source = stats_summary
        else:
            preview_coefficients = model_training_preview.get('coefficients', [])
            if preview_coefficients and isinstance(preview_coefficients, list) and len(preview_coefficients) > 0:
                coef_data_source = preview_coefficients
            elif coefficients and isinstance(coefficients, list) and len(coefficients) > 0:
                coef_data_source = coefficients
            else:
                coef_data_source = []
        
        if coef_data_source:
            # 表头（8列）
            headers = ['变量', '系数', '标准误', 'z值', 'P值', '95%CI下限', '95%CI上限', '显著性']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.font = self.styles.HEADER_FONT
                cell.fill = self.styles.HEADER_FILL
                cell.border = self.styles.THIN_BORDER
            row += 1
            
            for item in coef_data_source:
                feature_name = str(item.get('feature', item.get('variable', '')))
                if feature_name.lower() in ('const', '常数项', 'intercept'):
                    continue
                
                coef = item.get('coef', item.get('coefficient', 0))
                std_err = item.get('std_err', item.get('std_error', item.get('se', None)))
                z_val = item.get('z', item.get('z_value', item.get('zval', None)))
                p_val = item.get('p_value', item.get('pvalue', item.get('p_val', None)))
                ci_lower = item.get('ci_lower', item.get('conf_int_lower', item.get('ci_025', None)))
                ci_upper = item.get('ci_upper', item.get('conf_int_upper', item.get('ci_975', None)))
                sig = item.get('significance', '')
                
                # 自动计算显著性
                if not sig and isinstance(p_val, (int, float)):
                    if p_val < 0.001:
                        sig = '***'
                    elif p_val < 0.01:
                        sig = '**'
                    elif p_val < 0.05:
                        sig = '*'
                    elif p_val < 0.1:
                        sig = '.'
                
                row_data = [
                    feature_name,
                    f"{coef:.4f}" if isinstance(coef, (int, float)) else str(coef),
                    f"{std_err:.4f}" if isinstance(std_err, (int, float)) else '-',
                    f"{z_val:.4f}" if isinstance(z_val, (int, float)) else '-',
                    "<0.001" if isinstance(p_val, (int, float)) and p_val < 0.001 else (f"{p_val:.4f}" if isinstance(p_val, (int, float)) else '-'),
                    f"{ci_lower:.4f}" if isinstance(ci_lower, (int, float)) else '-',
                    f"{ci_upper:.4f}" if isinstance(ci_upper, (int, float)) else '-',
                    sig,
                ]
                
                for col, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row, column=col, value=value)
                    cell.font = self.styles.DATA_FONT
                    cell.border = self.styles.THIN_BORDER
                row += 1
            
            ws.cell(row=row, column=1, value="显著性标记: *** p<0.001, ** p<0.01, * p<0.05, . p<0.1").font = Font(name='Arial', size=9, italic=True)
            row += 2
        else:
            ws.cell(row=row, column=1, value="暂无系数统计数据").font = self.styles.DATA_FONT
            row += 1
        
        return row
    
    def _write_statistics_section(self, ws, results: dict, start_row: int) -> int:
        """写入模型统计检验（对应前端 statistics Tab）"""
        row = start_row
        
        model_statistics = results.get('model_statistics', {})
        if not model_statistics:
            ws.cell(row=row, column=1, value="暂无统计检验数据").font = self.styles.DATA_FONT
            return row + 1
        
        # 统计量映射
        stat_names = {
            'n_observations': '观测数',
            'n_params': '参数数量',
            'log_likelihood': '对数似然',
            'null_log_likelihood': '零模型对数似然',
            'pseudo_r2': '伪R²',
            'aic': 'AIC',
            'bic': 'BIC',
            'lr_stat': '似然比统计量',
            'lr_pvalue': '似然比p值',
        }
        
        headers = ['统计量', '值']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = self.styles.HEADER_FONT
            cell.fill = self.styles.HEADER_FILL
            cell.border = self.styles.THIN_BORDER
        row += 1
        
        for key, label in stat_names.items():
            value = model_statistics.get(key)
            if value is not None:
                row_data = [
                    label,
                    f"{value:.4f}" if isinstance(value, float) else str(value),
                ]
                for col, val in enumerate(row_data, 1):
                    cell = ws.cell(row=row, column=col, value=val)
                    cell.font = self.styles.DATA_FONT
                    cell.border = self.styles.THIN_BORDER
                row += 1
        
        row += 1
        return row
    
    def _write_stages_summary_section(self, ws, stages: dict, start_row: int) -> int:
        """写入阶段执行摘要"""
        row = start_row
        
        if not stages:
            ws.cell(row=row, column=1, value="暂无阶段数据").font = self.styles.DATA_FONT
            return row + 1
        
        for stage_id, stage_data in stages.items():
            stage_name = stage_data.get('stage_name', stage_id)
            status = stage_data.get('status', 'unknown')
            
            # 状态图标
            status_icon = "✅" if status == 'completed' else ("⏸️" if status == 'paused' else ("❌" if status == 'failed' else "⏳"))
            
            ws.cell(row=row, column=1, value=f"{stage_name} {status_icon}").font = self.styles.SUBHEADER_FONT
            row += 1
            
            output_preview = stage_data.get('output_preview')
            if output_preview and isinstance(output_preview, dict):
                # 显示概要字段
                exclude_fields = {
                    '_full_stage_data', '_skip_expert_pause', '_skipped_during_retry',
                    'retry_message', 'feature_details', 'bins_data', 'scorecard_data'
                }
                
                count = 0
                for key, value in output_preview.items():
                    if key.startswith('_') or key in exclude_fields:
                        continue
                    if isinstance(value, (list, dict)) and len(str(value)) > 200:
                        continue
                    if count >= 5:
                        break
                    
                    if isinstance(value, (int, float)):
                        value_str = f"{value:.4f}" if isinstance(value, float) else str(value)
                    elif isinstance(value, str) and len(value) < 80:
                        value_str = value
                    else:
                        continue
                    
                    ws.cell(row=row, column=1, value=f"  {key}").font = self.styles.DATA_FONT
                    ws.cell(row=row, column=2, value=value_str).font = self.styles.DATA_FONT
                    row += 1
                    count += 1
            
            row += 1
        
        return row
    
    def _write_section_header(self, ws, row: int, title: str) -> int:
        """写入章节标题"""
        ws.merge_cells(f'A{row}:H{row}')
        cell = ws.cell(row=row, column=1, value=title)
        cell.font = Font(name='Arial', size=12, bold=True, color="1F4E79")
        cell.fill = PatternFill(start_color="D6DCE4", end_color="D6DCE4", fill_type="solid")
        ws.row_dimensions[row].height = 22
        return row + 2
    
    def _write_ai_analysis_to_sheet(self, ws, ai_analysis: str, start_row: int) -> int:
        """将 AI 分析文本写入 Sheet（处理 Markdown 格式）"""
        row = start_row
        
        # 按段落分割
        paragraphs = ai_analysis.strip().split('\n\n')
        
        for para in paragraphs:
            para = para.strip()
            if not para:
                continue
            
            # 处理标题行（**xxx**）
            if para.startswith('**') and para.endswith('**'):
                # 加粗标题
                title_text = para.strip('*')
                ws.cell(row=row, column=1, value=title_text).font = Font(name='Arial', size=10, bold=True)
                row += 1
            elif para.startswith('**'):
                # 混合格式：**标题** 内容
                parts = para.split('**')
                content = ''.join(parts)
                # 合并单元格以容纳长文本
                ws.merge_cells(f'A{row}:H{row}')
                cell = ws.cell(row=row, column=1, value=content)
                cell.font = self.styles.DATA_FONT
                cell.alignment = Alignment(wrap_text=True, vertical='top')
                ws.row_dimensions[row].height = max(20, len(content) // 80 * 15 + 20)
                row += 1
            elif para.startswith('1.') or para.startswith('2.') or para.startswith('3.'):
                # 编号列表
                ws.merge_cells(f'A{row}:H{row}')
                cell = ws.cell(row=row, column=1, value=para)
                cell.font = self.styles.DATA_FONT
                cell.alignment = Alignment(wrap_text=True, vertical='top')
                ws.row_dimensions[row].height = max(20, len(para) // 80 * 15 + 20)
                row += 1
            else:
                # 普通段落
                ws.merge_cells(f'A{row}:H{row}')
                cell = ws.cell(row=row, column=1, value=para)
                cell.font = self.styles.DATA_FONT
                cell.alignment = Alignment(wrap_text=True, vertical='top')
                ws.row_dimensions[row].height = max(20, len(para) // 80 * 15 + 20)
                row += 1
        
        return row
    
    def _write_sample_features_section(self, ws, stages: dict, start_row: int) -> int:
        """
        写入样本及特征章节（对齐 MD/HTML/Word 报告内容）
        
        2026-02-10: 精简内容，与前端Tab对齐，移除冗余内容
        """
        row = start_row
        
        preprocessing = stages.get('preprocessing', {}).get('output_preview', {})
        feature_eng = stages.get('feature_engineering', {}).get('output_preview', {})
        # 放宽条件：只要有 before_count 或 after_count 或 selection_flow，就认为有特征工程数据
        has_feature_engineering = bool(feature_eng) and (
            feature_eng.get('before_count') is not None or 
            feature_eng.get('after_count') is not None or 
            feature_eng.get('selection_flow')
        )
        
        # ===== 样本概览（与前端一致）=====
        ws.cell(row=row, column=1, value="样本概览").font = self.styles.SUBHEADER_FONT
        row += 1
        
        # 总样本数
        rows_count = preprocessing.get('rows')
        if rows_count is not None:
            ws.cell(row=row, column=1, value="总样本数").font = self.styles.DATA_FONT
            ws.cell(row=row, column=2, value=f"{rows_count:,}").font = self.styles.DATA_FONT
            row += 1
        
        # 总体坏账率
        target_rate = preprocessing.get('target_rate')
        if target_rate is not None:
            ws.cell(row=row, column=1, value="总体坏账率").font = self.styles.DATA_FONT
            ws.cell(row=row, column=2, value=f"{target_rate*100:.2f}%").font = self.styles.DATA_FONT
            row += 1
        
        # 训练集/测试集（从 split_info 获取，与 MD 报告一致）
        split_info = preprocessing.get('split_info', {})
        if split_info:
            train_count = split_info.get('train')
            train_rate = split_info.get('train_target_rate')
            if train_count is not None:
                rate_str = f" (坏账率: {train_rate*100:.2f}%)" if train_rate else ""
                ws.cell(row=row, column=1, value="训练集").font = self.styles.DATA_FONT
                ws.cell(row=row, column=2, value=f"{train_count:,}{rate_str}").font = self.styles.DATA_FONT
                row += 1
            
            test_count = split_info.get('test')
            test_rate = split_info.get('test_target_rate')
            if test_count is not None:
                rate_str = f" (坏账率: {test_rate*100:.2f}%)" if test_rate else ""
                ws.cell(row=row, column=1, value="测试集").font = self.styles.DATA_FONT
                ws.cell(row=row, column=2, value=f"{test_count:,}{rate_str}").font = self.styles.DATA_FONT
                row += 1
        row += 1
        
        # ===== 时间范围（与前端一致，新增）=====
        time_range_info = preprocessing.get('time_range_info', {})
        if time_range_info:
            time_col = time_range_info.get('column', '')
            time_col_display = f"（{time_col}）" if time_col else ""
            
            ws.cell(row=row, column=1, value=f"时间范围{time_col_display}").font = self.styles.SUBHEADER_FONT
            row += 1
            
            # 表头
            headers = ['数据集', '起始时间', '截止时间']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.font = self.styles.HEADER_FONT
                cell.fill = self.styles.HEADER_FILL
                cell.border = self.styles.THIN_BORDER
            row += 1
            
            train_range = time_range_info.get('train', {})
            if train_range:
                ws.cell(row=row, column=1, value="训练集").font = self.styles.DATA_FONT
                ws.cell(row=row, column=2, value=str(train_range.get('min', '-'))).font = self.styles.DATA_FONT
                ws.cell(row=row, column=3, value=str(train_range.get('max', '-'))).font = self.styles.DATA_FONT
                for col in range(1, 4):
                    ws.cell(row=row, column=col).border = self.styles.THIN_BORDER
                row += 1
            
            test_range = time_range_info.get('test', {})
            if test_range:
                ws.cell(row=row, column=1, value="测试集").font = self.styles.DATA_FONT
                ws.cell(row=row, column=2, value=str(test_range.get('min', '-'))).font = self.styles.DATA_FONT
                ws.cell(row=row, column=3, value=str(test_range.get('max', '-'))).font = self.styles.DATA_FONT
                for col in range(1, 4):
                    ws.cell(row=row, column=col).border = self.styles.THIN_BORDER
                row += 1
            
            oot_range = time_range_info.get('oot', {})
            if oot_range and oot_range.get('min'):
                ws.cell(row=row, column=1, value="OOT验证集").font = self.styles.DATA_FONT
                ws.cell(row=row, column=2, value=str(oot_range.get('min', '-'))).font = self.styles.DATA_FONT
                ws.cell(row=row, column=3, value=str(oot_range.get('max', '-'))).font = self.styles.DATA_FONT
                for col in range(1, 4):
                    ws.cell(row=row, column=col).border = self.styles.THIN_BORDER
                row += 1
            
            row += 1
        
        # ===== 特征概览（与前端一致：原始特征数、筛选后特征、平均缺失率）=====
        ws.cell(row=row, column=1, value="特征概览").font = self.styles.SUBHEADER_FONT
        row += 1
        
        # 原始特征数
        feature_count = preprocessing.get('feature_count')
        if feature_count is not None:
            ws.cell(row=row, column=1, value="原始特征数").font = self.styles.DATA_FONT
            ws.cell(row=row, column=2, value=str(feature_count)).font = self.styles.DATA_FONT
            row += 1
        
        # 筛选后特征
        if has_feature_engineering and feature_eng.get('after_count') is not None:
            ws.cell(row=row, column=1, value="筛选后特征").font = self.styles.DATA_FONT
            ws.cell(row=row, column=2, value=str(feature_eng['after_count'])).font = self.styles.DATA_FONT
            row += 1
        
        # 平均缺失率
        missing_rate = preprocessing.get('missing_rate')
        if missing_rate is not None:
            ws.cell(row=row, column=1, value="平均缺失率").font = self.styles.DATA_FONT
            ws.cell(row=row, column=2, value=f"{missing_rate*100:.1f}%").font = self.styles.DATA_FONT
            row += 1
        row += 1
        
        # ===== 特征变化流程（与前端一致：初始 → 缺失率筛选 → One-Hot后 → IV筛选）=====
        if has_feature_engineering:
            before_count = feature_eng.get('before_count')
            after_count = feature_eng.get('after_count')
            
            # 优先使用 selection_flow（规则挖掘任务使用此格式）
            selection_flow = feature_eng.get('selection_flow', [])
            
            ws.cell(row=row, column=1, value="特征变化流程").font = self.styles.SUBHEADER_FONT
            row += 1
            
            steps = []
            
            if selection_flow:
                # 使用 selection_flow 格式（规则挖掘任务）
                for step in selection_flow:
                    step_name = step.get('step', '')
                    step_count = step.get('count', 0)
                    step_removed = step.get('removed', 0)
                    step_added = step.get('added', 0)
                    
                    diff_str = ''
                    if step_removed > 0:
                        diff_str += f" (-{step_removed})"
                    if step_added > 0:
                        diff_str += f" (+{step_added})"
                    
                    steps.append(f"{step_count}{diff_str} ({step_name})")
            else:
                # 使用旧格式（评分卡任务兼容）
                var_filter = feature_eng.get('var_filter_result', {})
                onehot_info = feature_eng.get('onehot_info', {})
                
                # 初始特征
                initial_count = feature_count or before_count
                if initial_count:
                    steps.append(f"{initial_count} (初始)")
                
                # 缺失率筛选后（如有）
                missing_filtered = var_filter.get('after_missing_filter') or before_count
                if missing_filtered and missing_filtered != initial_count:
                    steps.append(f"{missing_filtered} (缺失率筛选)")
                
                # One-Hot后（如有）
                onehot_after = onehot_info.get('after_count')
                if onehot_after:
                    onehot_before = onehot_info.get('before_count', 0)
                    diff = onehot_after - onehot_before if onehot_before else 0
                    diff_str = f" (+{diff})" if diff > 0 else ""
                    steps.append(f"{onehot_after}{diff_str} (One-Hot后)")
                
                # IV筛选后
                if after_count is not None:
                    base_count = onehot_after or before_count or initial_count or 0
                    removed = base_count - after_count
                    removed_str = f" (-{removed})" if removed > 0 else ""
                    steps.append(f"{after_count}{removed_str} (IV筛选)")
            
            if steps:
                ws.cell(row=row, column=1, value=" → ".join(steps)).font = self.styles.DATA_FONT
                row += 1
            
            row += 1
        
        return row
    
    def _write_rules_table_to_report(self, ws, rules: list, start_row: int, max_rules: int = 15) -> int:
        """写入规则表格到任务报告"""
        row = start_row
        
        # 表头
        headers = ['序号', '规则条件', '召回率', '命中率', '坏账率', 'Lift', '累计召回']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = self.styles.HEADER_FONT
            cell.fill = self.styles.HEADER_FILL
            cell.border = self.styles.THIN_BORDER
            cell.alignment = self.styles.CENTER
        row += 1
        
        # 数据行
        for i, rule in enumerate(rules[:max_rules], 1):
            rule_text = str(rule.get('rule', rule.get('condition', '-')))
            if len(rule_text) > 60:
                rule_text = rule_text[:57] + "..."
            
            recall = rule.get('recall', 0)
            hit_rate = rule.get('hit_rate', 0)
            bad_rate = rule.get('bad_rate', 0)
            lift = rule.get('lift', 0)
            cum_recall = rule.get('cumulative_recall', rule.get('cum_recall', rule.get('dev_cum_recall', 0)))
            
            row_data = [
                i,
                rule_text,
                f"{recall*100:.2f}%" if isinstance(recall, (int, float)) else '-',
                f"{hit_rate*100:.2f}%" if isinstance(hit_rate, (int, float)) else '-',
                f"{bad_rate*100:.2f}%" if isinstance(bad_rate, (int, float)) else '-',
                f"{lift:.2f}" if isinstance(lift, (int, float)) else '-',
                f"{cum_recall*100:.2f}%" if isinstance(cum_recall, (int, float)) else '-',
            ]
            
            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.font = self.styles.DATA_FONT
                cell.border = self.styles.THIN_BORDER
            row += 1
        
        if len(rules) > max_rules:
            ws.cell(row=row, column=1, value=f"... 共 {len(rules)} 条规则，仅显示前 {max_rules} 条").font = Font(name='Arial', size=9, italic=True)
            row += 1
        
        return row
    
    def _write_filtering_flow_section(self, ws, stages: dict, start_row: int) -> int:
        """写入规则筛选流程章节"""
        row = start_row
        
        rule_filtering = stages.get('rule_filtering', {}).get('output_preview', {})
        selecting_rules = stages.get('selecting_rules', {}).get('output_preview', {})
        
        # 漏斗概览
        generated_count = rule_filtering.get('generated_count', 0)
        filtered_count = rule_filtering.get('after_count', 0)
        optimal_count = selecting_rules.get('after_count', 0)
        
        if generated_count > 0 or filtered_count > 0 or optimal_count > 0:
            ws.cell(row=row, column=1, value="漏斗概览").font = self.styles.SUBHEADER_FONT
            row += 1
            
            # 表头
            for col, header in enumerate(['阶段', '规则数', '比例'], 1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.font = self.styles.HEADER_FONT
                cell.fill = self.styles.HEADER_FILL
                cell.border = self.styles.THIN_BORDER
            row += 1
            
            funnel_data = [
                ('规则生成', generated_count, '100%' if generated_count > 0 else '-'),
                ('规则筛选', filtered_count, f"{filtered_count/generated_count*100:.1f}%" if generated_count > 0 else '-'),
                ('最优选择', optimal_count, f"{optimal_count/generated_count*100:.1f}%" if generated_count > 0 else '-'),
            ]
            
            for stage, count, pct in funnel_data:
                ws.cell(row=row, column=1, value=stage).font = self.styles.DATA_FONT
                ws.cell(row=row, column=1).border = self.styles.THIN_BORDER
                ws.cell(row=row, column=2, value=count).font = self.styles.DATA_FONT
                ws.cell(row=row, column=2).border = self.styles.THIN_BORDER
                ws.cell(row=row, column=3, value=pct).font = self.styles.DATA_FONT
                ws.cell(row=row, column=3).border = self.styles.THIN_BORDER
                row += 1
            row += 1
        
        # 4.1 规则筛选阶段
        ws.cell(row=row, column=1, value="4.1 规则筛选阶段").font = Font(name='Arial', size=10, bold=True)
        row += 1
        
        filter_criteria = rule_filtering.get('filter_criteria', {})
        if filter_criteria:
            ws.cell(row=row, column=1, value="筛选条件:").font = self.styles.DATA_FONT
            row += 1
            min_lift = filter_criteria.get('min_lift')
            max_hit_rate = filter_criteria.get('max_hit_rate')
            ws.cell(row=row, column=1, value=f"  - 最小Lift阈值: {min_lift if min_lift is not None else '未设置'}").font = self.styles.DATA_FONT
            row += 1
            ws.cell(row=row, column=1, value=f"  - 最大命中率: {f'{max_hit_rate*100:.1f}%' if max_hit_rate is not None else '未设置'}").font = self.styles.DATA_FONT
            row += 1
        
        filter_summary = rule_filtering.get('filter_summary', {})
        if filter_summary:
            row += 1
            ws.cell(row=row, column=1, value="筛选结果:").font = self.styles.DATA_FONT
            row += 1
            
            filter_results = [
                ('单调性校验移除', filter_summary.get('direction_removed', 0)),
                ('坏账率为0移除', filter_summary.get('bad_rate_zero_removed', 0)),
                ('最小Lift阈值移除', filter_summary.get('lift_removed', 0)),
                ('最大命中率移除', filter_summary.get('hit_rate_removed', 0)),
                ('总移除', filter_summary.get('total_removed', 0)),
            ]
            
            for name, count in filter_results:
                ws.cell(row=row, column=1, value=f"  - {name}").font = self.styles.DATA_FONT
                ws.cell(row=row, column=2, value=count).font = self.styles.DATA_FONT
                row += 1
        row += 1
        
        # 4.2 最优选择阶段
        ws.cell(row=row, column=1, value="4.2 最优选择阶段").font = Font(name='Arial', size=10, bold=True)
        row += 1
        
        if selecting_rules:
            ws.cell(row=row, column=1, value="选择条件:").font = self.styles.DATA_FONT
            row += 1
            
            allow_overlap = selecting_rules.get('allow_overlap', False)
            selection_mode_text = "允许重叠（独立选择）" if allow_overlap else "贪婪算法（不允许重叠）"
            ws.cell(row=row, column=1, value=f"  - 选择模式: {selection_mode_text}").font = self.styles.DATA_FONT
            row += 1
            
            max_hit_rate = selecting_rules.get('max_hit_rate')
            ws.cell(row=row, column=1, value=f"  - 最大命中率（规则集）: {f'{max_hit_rate*100:.1f}%' if max_hit_rate else '未设置'}").font = self.styles.DATA_FONT
            row += 1
            
            # 放弃原因统计
            rejected_stats = selecting_rules.get('rejected_rules_stats', {})
            reason_distribution = rejected_stats.get('reason_distribution', {})
            selection_mode = rejected_stats.get('selection_mode', 'greedy' if not allow_overlap else 'overlap')
            
            # 注：坏账率为0的规则在规则筛选阶段已被过滤，最优选择阶段不再需要该原因
            # v2.5: 贪婪模式下移除"未被选中"，所有未被选中的规则都归类为"样本被消耗"
            greedy_reasons = ["命中率达上限", "样本被消耗（贪婪模式）", "目标坏账率已达成", "召回率目标已达成"]
            overlap_reasons = ["命中率达上限", "目标坏账率已达成", "召回率目标已达成", "排序靠后", "异常情况（请检查数据）"]
            possible_reasons = overlap_reasons if selection_mode == 'overlap' else greedy_reasons
            
            row += 1
            ws.cell(row=row, column=1, value="选择结果（放弃原因）:").font = self.styles.DATA_FONT
            row += 1
            
            total_rejected = 0
            for reason in possible_reasons:
                count = reason_distribution.get(reason, 0)
                ws.cell(row=row, column=1, value=f"  - {reason}").font = self.styles.DATA_FONT
                ws.cell(row=row, column=2, value=count).font = self.styles.DATA_FONT
                total_rejected += count
                row += 1
            
            ws.cell(row=row, column=1, value="  - 总放弃").font = Font(name='Arial', size=10, bold=True)
            ws.cell(row=row, column=2, value=total_rejected).font = Font(name='Arial', size=10, bold=True)
            row += 1
        
        return row + 1
    
    def _write_validation_section(self, ws, validation_report: dict, start_row: int) -> int:
        """写入质量验证章节（对齐 MD/HTML 报告和前端UI结构）"""
        row = start_row
        
        # 状态映射
        status_labels = {
            'excellent': '优秀',
            'good': '良好', 
            'acceptable': '可接受',
            'warning': '需优化',
            'warning_low': '偏低',
            'warning_high': '偏高',
            'error': '异常',
            'ok': '正常',
        }
        
        # 综合评分（字段名: quality_score，不是 total_score）
        quality_score = validation_report.get('quality_score', 0)
        ws.cell(row=row, column=1, value=f"综合质量评分: {quality_score:.1f} / 100").font = Font(name='Arial', size=11, bold=True)
        row += 1
        
        # 评分明细（加权得分）
        score_breakdown = validation_report.get('score_breakdown', {})
        if score_breakdown:
            dimension_names = {
                'discrimination': '提升度(30分)',
                'recall': '召回率(25分)',
                'coverage': '命中率(15分)',
                'independence': '独立性(15分)',
                'complexity': '复杂度(15分)',
            }
            scores = [f"{dimension_names.get(k, k)}: {v:.1f}" for k, v in score_breakdown.items() if k in dimension_names]
            if scores:
                ws.cell(row=row, column=1, value=f"得分明细: {' | '.join(scores)}").font = self.styles.DATA_FONT
                row += 2
        
        # 评估详情表格（与 MD 报告结构一致）
        headers = ['评估维度', '核心指标', '状态', '说明']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = self.styles.HEADER_FONT
            cell.fill = self.styles.HEADER_FILL
            cell.border = self.styles.THIN_BORDER
        row += 1
        
        # 提升度
        disc = validation_report.get('discrimination_report', {})
        if disc and isinstance(disc, dict):
            status = disc.get('status', 'error')
            avg_lift = disc.get('avg_lift', 0)
            min_lift = disc.get('min_lift', 0)
            max_lift = disc.get('max_lift', 0)
            row_data = ['提升度 (Lift)', f'平均: {avg_lift:.2f}x', status_labels.get(status, status), f'范围[{min_lift:.2f}, {max_lift:.2f}]']
            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.font = self.styles.DATA_FONT
                cell.border = self.styles.THIN_BORDER
            row += 1
        
        # 召回率
        recall = validation_report.get('recall_report', {})
        if recall and isinstance(recall, dict):
            status = recall.get('status', 'error')
            cumulative_recall = recall.get('cumulative_recall', 0)
            row_data = ['召回率', f'累计: {cumulative_recall*100:.2f}%', status_labels.get(status, status), '对坏客户的捕获能力']
            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.font = self.styles.DATA_FONT
                cell.border = self.styles.THIN_BORDER
            row += 1
        
        # 覆盖率/命中率
        coverage = validation_report.get('coverage_report', {})
        if coverage and isinstance(coverage, dict):
            status = coverage.get('status', 'error')
            total_coverage = coverage.get('total_coverage', 0)
            row_data = ['命中率/覆盖率', f'累计: {total_coverage*100:.2f}%', status_labels.get(status, status), '规则命中样本比例']
            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.font = self.styles.DATA_FONT
                cell.border = self.styles.THIN_BORDER
            row += 1
        
        # 重叠度
        overlap = validation_report.get('overlap_report', {})
        if overlap and isinstance(overlap, dict):
            status = overlap.get('status', 'ok')
            avg_overlap = overlap.get('avg_overlap', 0)
            desc = "无重叠" if avg_overlap == 0 else f"平均重叠{avg_overlap*100:.1f}%"
            row_data = ['重叠度', f'平均: {avg_overlap*100:.1f}%', status_labels.get(status, status), desc]
            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.font = self.styles.DATA_FONT
                cell.border = self.styles.THIN_BORDER
            row += 1
        
        # 冗余度
        redundancy = validation_report.get('redundancy_report', {})
        if redundancy and isinstance(redundancy, dict):
            status = redundancy.get('status', 'ok')
            redundant_count = redundancy.get('redundant_count', 0)
            desc = "无冗余" if redundant_count == 0 else f"{redundant_count}对冗余规则"
            row_data = ['冗余度', f'{redundant_count}对', status_labels.get(status, status), desc]
            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.font = self.styles.DATA_FONT
                cell.border = self.styles.THIN_BORDER
            row += 1
        
        # 复杂度
        complexity = validation_report.get('complexity_report', {})
        if complexity and isinstance(complexity, dict):
            status = complexity.get('status', 'ok')
            avg_complexity = complexity.get('avg_complexity', 0)
            max_complexity = complexity.get('max_complexity', 0)
            row_data = ['复杂度', f'平均: {avg_complexity:.1f}条件', status_labels.get(status, status), f'最大{max_complexity}个条件']
            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.font = self.styles.DATA_FONT
                cell.border = self.styles.THIN_BORDER
            row += 1
        
        # 优化建议/警告
        warnings = validation_report.get('warnings', [])
        if warnings:
            row += 1
            ws.cell(row=row, column=1, value="优化建议:").font = self.styles.SUBHEADER_FONT
            row += 1
            for warning in warnings[:5]:
                ws.cell(row=row, column=1, value=f"  • {warning}").font = self.styles.DATA_FONT
                row += 1
        elif quality_score >= 80:
            row += 1
            ws.cell(row=row, column=1, value="规则集质量优秀，各项指标均达标").font = Font(name='Arial', size=10, italic=True, color='228B22')
            row += 1
        
        return row
    
    def _write_psi_section(self, ws, psi_report: list, start_row: int) -> int:
        """写入 PSI 稳定性章节"""
        row = start_row
        
        headers = ['序号', '规则', 'PSI值', '稳定性']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = self.styles.HEADER_FONT
            cell.fill = self.styles.HEADER_FILL
            cell.border = self.styles.THIN_BORDER
        row += 1
        
        for i, item in enumerate(psi_report[:15], 1):
            rule_text = str(item.get('rule', ''))
            if len(rule_text) > 40:
                rule_text = rule_text[:37] + "..."
            
            psi_value = item.get('psi', 0)
            stability = "✅ 稳定" if psi_value < 0.1 else ("⚠️ 轻微变化" if psi_value < 0.25 else "❌ 显著变化")
            
            row_data = [i, rule_text, f"{psi_value:.4f}", stability]
            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.font = self.styles.DATA_FONT
                cell.border = self.styles.THIN_BORDER
            row += 1
        
        # PSI 说明
        row += 1
        ws.cell(row=row, column=1, value="PSI指标说明：PSI < 0.1 稳定 | 0.1-0.25 轻微变化 | ≥0.25 显著变化").font = Font(name='Arial', size=9, italic=True)
        row += 1
        
        return row
    
    def _write_advanced_analysis_section(self, ws, amount_analysis: dict | None, prior_analysis: dict | None, start_row: int) -> int:
        """写入附加分析章节（FIX-4: 匹配 FIX-1 扁平化后的 amount_analysis 结构）"""
        row = start_row
        
        # 金额分析
        if amount_analysis and isinstance(amount_analysis, dict) and amount_analysis.get('enabled'):
            ws.cell(row=row, column=1, value="💰 金额维度分析").font = self.styles.SUBHEADER_FONT
            row += 1
            
            # 汇总指标
            summary_items = [
                ("总金额", self._format_value(amount_analysis.get('total_amount', 0))),
                ("总坏账金额", self._format_value(amount_analysis.get('total_bad_amount', 0))),
                ("整体金额坏账率", self._format_percent(amount_analysis.get('overall_amount_bad_rate', 0))),
            ]
            cumulative = amount_analysis.get('cumulative', {})
            if cumulative:
                summary_items.append(("累计命中金额", self._format_value(cumulative.get('cum_hit_amount', 0))))
                summary_items.append(("金额召回率", self._format_percent(cumulative.get('amount_recall', 0))))
            
            for label, value in summary_items:
                ws.cell(row=row, column=1, value=label).font = self.styles.DATA_FONT
                ws.cell(row=row, column=2, value=str(value)).font = self.styles.DATA_FONT
                row += 1
            
            # 规则金额明细
            rules_amount = amount_analysis.get('rules_amount', [])
            if rules_amount:
                row += 1
                ws.cell(row=row, column=1, value="规则金额明细").font = self.styles.SUBHEADER_FONT
                row += 1
                headers = ['规则', '命中金额', '金额占比', '坏账金额', '金额坏账率', '金额Lift']
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=row, column=col, value=header)
                    cell.font = self.styles.HEADER_FONT
                    cell.fill = self.styles.HEADER_FILL
                    cell.border = self.styles.THIN_BORDER
                row += 1
                for item in rules_amount[:20]:
                    ws.cell(row=row, column=1, value=str(item.get('rule', '')))
                    ws.cell(row=row, column=2, value=self._format_value(item.get('hit_amount', 0)))
                    ws.cell(row=row, column=3, value=self._format_percent(item.get('hit_amount_pct', 0)))
                    ws.cell(row=row, column=4, value=self._format_value(item.get('bad_amount', 0)))
                    ws.cell(row=row, column=5, value=self._format_percent(item.get('amount_bad_rate', 0)))
                    ws.cell(row=row, column=6, value=round(item.get('amount_lift', 0), 2))
                    for col in range(1, 7):
                        ws.cell(row=row, column=col).border = self.styles.THIN_BORDER
                        ws.cell(row=row, column=col).font = self.styles.DATA_FONT
                    row += 1
            row += 1
        
        # 先验规则分析
        if prior_analysis and isinstance(prior_analysis, dict):
            ws.cell(row=row, column=1, value="📋 先验规则分析").font = self.styles.SUBHEADER_FONT
            row += 1
            
            for key, value in list(prior_analysis.items())[:10]:
                if not key.startswith('_'):
                    ws.cell(row=row, column=1, value=str(key)).font = self.styles.DATA_FONT
                    ws.cell(row=row, column=2, value=self._format_value(value)).font = self.styles.DATA_FONT
                    row += 1
        
        return row

    def _add_overview_sheet(
        self,
        wb: "Workbook",
        results: dict[str, Any],
        title: str
    ) -> None:
        """Add overview sheet with summary metrics."""
        ws = wb.create_sheet("概览")
        
        # Title
        ws.merge_cells('A1:F1')
        ws['A1'] = title
        ws['A1'].font = self.styles.TITLE_FONT
        ws['A1'].fill = self.styles.HEADER_FILL
        ws['A1'].alignment = self.styles.CENTER
        ws.row_dimensions[1].height = 30
        
        # Summary metrics
        row = 3
        ws['A3'] = "模型概览"
        ws['A3'].font = self.styles.SUBHEADER_FONT
        ws['A3'].fill = self.styles.SUBHEADER_FILL
        ws.merge_cells('A3:F3')
        
        # Extract metrics
        metrics = [
            ("样本量", results.get('n_samples', 'N/A')),
            ("特征数", results.get('n_features', 'N/A')),
            ("训练集AUC", self._format_metric(results.get('train_auc'))),
            ("测试集AUC", self._format_metric(results.get('test_auc'))),
            ("训练集KS", self._format_metric(results.get('train_ks'))),
            ("测试集KS", self._format_metric(results.get('test_ks'))),
        ]
        
        row = 5
        for i, (name, value) in enumerate(metrics):
            col = (i % 3) * 2 + 1
            if i > 0 and i % 3 == 0:
                row += 2
            ws.cell(row=row, column=col, value=name).font = self.styles.DATA_FONT
            ws.cell(row=row, column=col+1, value=value).font = self.styles.DATA_FONT
            ws.cell(row=row, column=col+1).alignment = self.styles.RIGHT
        
        # Model statistics if available
        if 'model_statistics' in results:
            stats = results['model_statistics']
            row += 4
            ws.cell(row=row, column=1, value="模型统计").font = self.styles.SUBHEADER_FONT
            ws.cell(row=row, column=1).fill = self.styles.SUBHEADER_FILL
            ws.merge_cells(f'A{row}:F{row}')
            
            stat_metrics = [
                ("伪R²", self._format_metric(stats.get('pseudo_r2'))),
                ("对数似然", self._format_metric(stats.get('log_likelihood'))),
                ("AIC", self._format_metric(stats.get('aic'))),
                ("BIC", self._format_metric(stats.get('bic'))),
            ]
            
            row += 2
            for i, (name, value) in enumerate(stat_metrics):
                col = (i % 2) * 3 + 1
                if i > 0 and i % 2 == 0:
                    row += 1
                ws.cell(row=row, column=col, value=name).font = self.styles.DATA_FONT
                ws.cell(row=row, column=col+1, value=value).font = self.styles.DATA_FONT
        
        # Adjust column widths
        for col in ['A', 'B', 'C', 'D', 'E', 'F']:
            ws.column_dimensions[col].width = 15
    
    def _add_bins_sheet(self, wb: "Workbook", results: dict[str, Any]) -> None:
        """Add binning details sheet."""
        ws = wb.create_sheet("分箱详情")
        
        woe_results = results.get('woe_results', {})
        
        row = 1
        for var_name, var_data in woe_results.items():
            # Variable header
            ws.cell(row=row, column=1, value=f"变量: {var_name}")
            ws.cell(row=row, column=1).font = self.styles.SUBHEADER_FONT
            ws.cell(row=row, column=1).fill = self.styles.SUBHEADER_FILL
            ws.merge_cells(f'A{row}:H{row}')
            row += 1
            
            # Get binning data
            if isinstance(var_data, dict) and 'bins' in var_data:
                bins_df = var_data['bins']
                if isinstance(bins_df, pd.DataFrame):
                    # Write headers
                    headers = ['分箱', '样本数', '坏样本数', '坏账率', 'WOE', 'IV']
                    for col, header in enumerate(headers, 1):
                        cell = ws.cell(row=row, column=col, value=header)
                        cell.font = self.styles.HEADER_FONT
                        cell.fill = self.styles.HEADER_FILL
                        cell.alignment = self.styles.CENTER
                        cell.border = self.styles.THIN_BORDER
                    row += 1
                    
                    # Write data
                    for _, data_row in bins_df.iterrows():
                        ws.cell(row=row, column=1, value=str(data_row.get('bin', '')))
                        ws.cell(row=row, column=2, value=data_row.get('count', 0))
                        ws.cell(row=row, column=3, value=data_row.get('bad', 0))
                        ws.cell(row=row, column=4, value=self._format_percent(data_row.get('bad_rate', 0)))
                        ws.cell(row=row, column=5, value=round(data_row.get('woe', 0), 4))
                        ws.cell(row=row, column=6, value=round(data_row.get('iv', 0), 4))
                        
                        for col in range(1, 7):
                            ws.cell(row=row, column=col).border = self.styles.THIN_BORDER
                            ws.cell(row=row, column=col).font = self.styles.DATA_FONT
                        row += 1
            
            row += 2  # Space between variables
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 25
        for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H']:
            ws.column_dimensions[col].width = 12
    
    def _add_scorecard_sheet(self, wb: "Workbook", results: dict[str, Any]) -> None:
        """Add scorecard sheet."""
        ws = wb.create_sheet("评分卡")
        
        scorecard = results.get('scorecard', [])
        
        # Title
        ws['A1'] = "评分卡"
        ws['A1'].font = self.styles.TITLE_FONT
        ws['A1'].fill = self.styles.HEADER_FILL
        ws.merge_cells('A1:D1')
        ws.row_dimensions[1].height = 25
        
        # Headers
        headers = ['变量', '分箱', '分数', 'WOE']
        row = 3
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = self.styles.HEADER_FONT
            cell.fill = self.styles.HEADER_FILL
            cell.alignment = self.styles.CENTER
            cell.border = self.styles.THIN_BORDER
        
        # Data
        row = 4
        if isinstance(scorecard, list):
            for item in scorecard:
                ws.cell(row=row, column=1, value=item.get('variable', ''))
                ws.cell(row=row, column=2, value=item.get('bin', ''))
                ws.cell(row=row, column=3, value=item.get('points', 0))
                ws.cell(row=row, column=4, value=round(item.get('woe', 0), 4))
                
                for col in range(1, 5):
                    ws.cell(row=row, column=col).border = self.styles.THIN_BORDER
                    ws.cell(row=row, column=col).font = self.styles.DATA_FONT
                row += 1
        elif isinstance(scorecard, pd.DataFrame):
            for _, data_row in scorecard.iterrows():
                ws.cell(row=row, column=1, value=data_row.get('variable', ''))
                ws.cell(row=row, column=2, value=str(data_row.get('bin', '')))
                ws.cell(row=row, column=3, value=data_row.get('points', 0))
                ws.cell(row=row, column=4, value=round(data_row.get('woe', 0), 4))
                
                for col in range(1, 5):
                    ws.cell(row=row, column=col).border = self.styles.THIN_BORDER
                    ws.cell(row=row, column=col).font = self.styles.DATA_FONT
                row += 1
        
        # Column widths
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 12
    
    def _add_statistics_sheet(self, wb: "Workbook", results: dict[str, Any]) -> None:
        """Add model statistics sheet."""
        ws = wb.create_sheet("模型统计")
        
        stats = results.get('model_statistics', {})
        
        # Title
        ws['A1'] = "模型统计信息"
        ws['A1'].font = self.styles.TITLE_FONT
        ws['A1'].fill = self.styles.HEADER_FILL
        ws.merge_cells('A1:H1')
        ws.row_dimensions[1].height = 25
        
        # Model fit metrics
        ws['A3'] = "模型拟合指标"
        ws['A3'].font = self.styles.SUBHEADER_FONT
        ws['A3'].fill = self.styles.SUBHEADER_FILL
        ws.merge_cells('A3:D3')
        
        fit_metrics = [
            ("样本量", stats.get('n_observations', 'N/A')),
            ("参数数量", stats.get('n_params', 'N/A')),
            ("对数似然", self._format_metric(stats.get('log_likelihood'))),
            ("零模型对数似然", self._format_metric(stats.get('null_log_likelihood'))),
            ("伪R²", self._format_metric(stats.get('pseudo_r2'))),
            ("AIC", self._format_metric(stats.get('aic'))),
            ("BIC", self._format_metric(stats.get('bic'))),
            ("似然比统计量", self._format_metric(stats.get('lr_stat'))),
            ("似然比p值", self._format_metric(stats.get('lr_pvalue'))),
        ]
        
        row = 5
        for name, value in fit_metrics:
            ws.cell(row=row, column=1, value=name).font = self.styles.DATA_FONT
            ws.cell(row=row, column=2, value=value).font = self.styles.DATA_FONT
            row += 1
        
        # Coefficient statistics
        row += 2
        ws.cell(row=row, column=1, value="系数统计").font = self.styles.SUBHEADER_FONT
        ws.cell(row=row, column=1).fill = self.styles.SUBHEADER_FILL
        ws.merge_cells(f'A{row}:H{row}')
        row += 1
        
        # Headers
        headers = ['变量', '系数', '标准误', 'z值', 'p值', '95%CI下限', '95%CI上限', '显著性']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = self.styles.HEADER_FONT
            cell.fill = self.styles.HEADER_FILL
            cell.alignment = self.styles.CENTER
            cell.border = self.styles.THIN_BORDER
        row += 1
        
        # Data
        summary = stats.get('summary', [])
        for item in summary:
            ws.cell(row=row, column=1, value=item.get('feature', ''))
            ws.cell(row=row, column=2, value=round(item.get('coef', 0), 4))
            ws.cell(row=row, column=3, value=round(item.get('std_err', 0), 4))
            ws.cell(row=row, column=4, value=round(item.get('z', 0), 4))
            ws.cell(row=row, column=5, value=self._format_pvalue(item.get('p_value', 1)))
            ws.cell(row=row, column=6, value=round(item.get('ci_lower', 0), 4))
            ws.cell(row=row, column=7, value=round(item.get('ci_upper', 0), 4))
            ws.cell(row=row, column=8, value=item.get('significance', ''))
            
            for col in range(1, 9):
                ws.cell(row=row, column=col).border = self.styles.THIN_BORDER
                ws.cell(row=row, column=col).font = self.styles.DATA_FONT
            
            # Highlight significant coefficients
            p_value = item.get('p_value', 1)
            if p_value is not None and p_value < 0.05:
                ws.cell(row=row, column=5).fill = self.styles.ACCENT_FILL
            
            row += 1
        
        # Significance legend
        row += 2
        ws.cell(row=row, column=1, value="显著性标记: *** p<0.001, ** p<0.01, * p<0.05, . p<0.1")
        ws.cell(row=row, column=1).font = Font(name='Arial', size=9, italic=True)
        
        # Column widths
        ws.column_dimensions['A'].width = 18
        for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H']:
            ws.column_dimensions[col].width = 12
    
    def _add_evaluation_sheet(self, wb: "Workbook", results: dict[str, Any]) -> None:
        """Add model evaluation sheet."""
        ws = wb.create_sheet("模型评估")
        
        # Title
        ws['A1'] = "模型评估指标"
        ws['A1'].font = self.styles.TITLE_FONT
        ws['A1'].fill = self.styles.HEADER_FILL
        ws.merge_cells('A1:F1')
        ws.row_dimensions[1].height = 25
        
        # Performance metrics
        metrics = [
            ("训练集", "train"),
            ("测试集", "test"),
            ("OOT", "oot"),
        ]
        
        row = 3
        headers = ['数据集', 'AUC', 'KS', 'Gini', 'Accuracy', 'F1']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = self.styles.HEADER_FONT
            cell.fill = self.styles.HEADER_FILL
            cell.alignment = self.styles.CENTER
            cell.border = self.styles.THIN_BORDER
        
        row = 4
        for name, prefix in metrics:
            auc = results.get(f'{prefix}_auc')
            ks = results.get(f'{prefix}_ks')
            if auc is not None or ks is not None:
                ws.cell(row=row, column=1, value=name)
                ws.cell(row=row, column=2, value=self._format_metric(auc))
                ws.cell(row=row, column=3, value=self._format_metric(ks))
                ws.cell(row=row, column=4, value=self._format_metric(results.get(f'{prefix}_gini')))
                ws.cell(row=row, column=5, value=self._format_metric(results.get(f'{prefix}_accuracy')))
                ws.cell(row=row, column=6, value=self._format_metric(results.get(f'{prefix}_f1')))
                
                for col in range(1, 7):
                    ws.cell(row=row, column=col).border = self.styles.THIN_BORDER
                    ws.cell(row=row, column=col).font = self.styles.DATA_FONT
                row += 1
        
        # Column widths
        for col in ['A', 'B', 'C', 'D', 'E', 'F']:
            ws.column_dimensions[col].width = 12
    
    def _add_charts_sheet(self, wb: "Workbook", results: dict[str, Any]) -> None:
        """Add charts sheet (placeholder for chart data)."""
        ws = wb.create_sheet("图表数据")
        
        ws['A1'] = "图表数据"
        ws['A1'].font = self.styles.TITLE_FONT
        ws['A1'].fill = self.styles.HEADER_FILL
        ws.merge_cells('A1:F1')
        
        # ROC data if available
        roc_data = results.get('roc_data')
        if roc_data is not None:
            ws['A3'] = "ROC曲线数据"
            ws['A3'].font = self.styles.SUBHEADER_FONT
            ws['A3'].fill = self.styles.SUBHEADER_FILL
            ws.merge_cells('A3:C3')
            
            headers = ['FPR', 'TPR', 'Threshold']
            for col, header in enumerate(headers, 1):
                ws.cell(row=4, column=col, value=header)
            
            if isinstance(roc_data, dict):
                fpr = roc_data.get('fpr', [])
                tpr = roc_data.get('tpr', [])
                thresholds = roc_data.get('thresholds', [])
                
                for i, (f, t, th) in enumerate(zip(fpr, tpr, thresholds)):
                    ws.cell(row=5+i, column=1, value=round(f, 4))
                    ws.cell(row=5+i, column=2, value=round(t, 4))
                    ws.cell(row=5+i, column=3, value=round(th, 4))
        
        # KS data if available
        ks_data = results.get('ks_data')
        if ks_data is not None:
            start_row = ws.max_row + 3
            ws.cell(row=start_row, column=1, value="KS曲线数据")
            ws.cell(row=start_row, column=1).font = self.styles.SUBHEADER_FONT
            ws.cell(row=start_row, column=1).fill = self.styles.SUBHEADER_FILL
    
    def _add_feature_selection_sheet(self, wb: "Workbook", results: dict[str, Any]) -> None:
        """Add feature selection details sheet."""
        ws = wb.create_sheet("特征选择")
        
        # Title
        ws['A1'] = "特征选择详情"
        ws['A1'].font = self.styles.TITLE_FONT
        ws['A1'].fill = self.styles.HEADER_FILL
        ws.merge_cells('A1:F1')
        ws.row_dimensions[1].height = 25
        
        row = 3
        
        # Overfit warning
        overfit_warning = results.get('overfit_warning')
        if overfit_warning:
            ws.cell(row=row, column=1, value="⚠️ 过拟合警告")
            ws.cell(row=row, column=1).font = self.styles.SUBHEADER_FONT
            ws.cell(row=row, column=1).fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
            ws.merge_cells(f'A{row}:F{row}')
            row += 1
            ws.cell(row=row, column=1, value=str(overfit_warning))
            ws.cell(row=row, column=1).font = self.styles.DATA_FONT
            ws.merge_cells(f'A{row}:F{row}')
            row += 2
        
        # Selected features
        selected_features = results.get('selected_features', [])
        if selected_features:
            ws.cell(row=row, column=1, value="入模特征列表")
            ws.cell(row=row, column=1).font = self.styles.SUBHEADER_FONT
            ws.cell(row=row, column=1).fill = self.styles.SUBHEADER_FILL
            ws.merge_cells(f'A{row}:F{row}')
            row += 1
            
            # Headers
            headers = ['序号', '特征名称']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.font = self.styles.HEADER_FONT
                cell.fill = self.styles.HEADER_FILL
                cell.border = self.styles.THIN_BORDER
            row += 1
            
            # Data
            for i, feature in enumerate(selected_features, 1):
                ws.cell(row=row, column=1, value=i)
                ws.cell(row=row, column=2, value=feature)
                for col in range(1, 3):
                    ws.cell(row=row, column=col).border = self.styles.THIN_BORDER
                    ws.cell(row=row, column=col).font = self.styles.DATA_FONT
                row += 1
            row += 1
        
        # Selection detail (stepwise regression, significance tests, etc.)
        selection_detail = results.get('selection_detail')
        if selection_detail and isinstance(selection_detail, dict):
            # Stepwise regression steps
            if 'steps' in selection_detail:
                ws.cell(row=row, column=1, value="逐步回归过程")
                ws.cell(row=row, column=1).font = self.styles.SUBHEADER_FONT
                ws.cell(row=row, column=1).fill = self.styles.SUBHEADER_FILL
                ws.merge_cells(f'A{row}:F{row}')
                row += 1
                
                headers = ['步骤', '操作', '变量', 'AIC']
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=row, column=col, value=header)
                    cell.font = self.styles.HEADER_FONT
                    cell.fill = self.styles.HEADER_FILL
                    cell.border = self.styles.THIN_BORDER
                row += 1
                
                for step in selection_detail['steps']:
                    ws.cell(row=row, column=1, value=step.get('step', ''))
                    ws.cell(row=row, column=2, value=step.get('action', ''))
                    ws.cell(row=row, column=3, value=step.get('variable', ''))
                    ws.cell(row=row, column=4, value=self._format_metric(step.get('aic')))
                    for col in range(1, 5):
                        ws.cell(row=row, column=col).border = self.styles.THIN_BORDER
                        ws.cell(row=row, column=col).font = self.styles.DATA_FONT
                    row += 1
                row += 1
            
            # Significance test
            if 'significance' in selection_detail:
                ws.cell(row=row, column=1, value="显著性检验")
                ws.cell(row=row, column=1).font = self.styles.SUBHEADER_FONT
                ws.cell(row=row, column=1).fill = self.styles.SUBHEADER_FILL
                ws.merge_cells(f'A{row}:F{row}')
                row += 1
                
                headers = ['变量', 'P值', '显著性']
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=row, column=col, value=header)
                    cell.font = self.styles.HEADER_FONT
                    cell.fill = self.styles.HEADER_FILL
                    cell.border = self.styles.THIN_BORDER
                row += 1
                
                for item in selection_detail['significance']:
                    ws.cell(row=row, column=1, value=item.get('feature', ''))
                    ws.cell(row=row, column=2, value=self._format_pvalue(item.get('p_value')))
                    ws.cell(row=row, column=3, value=item.get('significance', ''))
                    for col in range(1, 4):
                        ws.cell(row=row, column=col).border = self.styles.THIN_BORDER
                        ws.cell(row=row, column=col).font = self.styles.DATA_FONT
                    row += 1
                row += 1
            
            # Coefficient direction validation
            if 'coef_validation' in selection_detail:
                ws.cell(row=row, column=1, value="系数方向验证")
                ws.cell(row=row, column=1).font = self.styles.SUBHEADER_FONT
                ws.cell(row=row, column=1).fill = self.styles.SUBHEADER_FILL
                ws.merge_cells(f'A{row}:F{row}')
                row += 1
                
                coef_val = selection_detail['coef_validation']
                if 'valid' in coef_val:
                    ws.cell(row=row, column=1, value="✓ 系数方向正确的特征:")
                    ws.cell(row=row, column=1).font = Font(name='Arial', size=10, color="228B22")
                    row += 1
                    for feature in coef_val['valid']:
                        ws.cell(row=row, column=2, value=feature)
                        row += 1
                
                if 'invalid' in coef_val and coef_val['invalid']:
                    ws.cell(row=row, column=1, value="✗ 系数方向异常的特征:")
                    ws.cell(row=row, column=1).font = Font(name='Arial', size=10, color="DC143C")
                    row += 1
                    for feature in coef_val['invalid']:
                        ws.cell(row=row, column=2, value=feature)
                        row += 1
        
        # Outlier info
        outlier_info = results.get('outlier_info')
        if outlier_info and isinstance(outlier_info, list) and len(outlier_info) > 0:
            row += 1
            ws.cell(row=row, column=1, value="异常值检测结果")
            ws.cell(row=row, column=1).font = self.styles.SUBHEADER_FONT
            ws.cell(row=row, column=1).fill = self.styles.SUBHEADER_FILL
            ws.merge_cells(f'A{row}:F{row}')
            row += 1
            
            headers = ['变量', '异常值数量', '异常值占比']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.font = self.styles.HEADER_FONT
                cell.fill = self.styles.HEADER_FILL
                cell.border = self.styles.THIN_BORDER
            row += 1
            
            # Show top 15 by percentage
            sorted_outliers = sorted(outlier_info, key=lambda x: x.get('outlier_pct', 0), reverse=True)[:15]
            for item in sorted_outliers:
                ws.cell(row=row, column=1, value=item.get('variable', ''))
                ws.cell(row=row, column=2, value=item.get('outlier_count', 0))
                ws.cell(row=row, column=3, value=self._format_percent(item.get('outlier_pct', 0)))
                for col in range(1, 4):
                    ws.cell(row=row, column=col).border = self.styles.THIN_BORDER
                    ws.cell(row=row, column=col).font = self.styles.DATA_FONT
                row += 1
        
        # Column widths
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 15
    
    def _add_iv_table_sheet(self, wb: "Workbook", results: dict[str, Any]) -> None:
        """Add IV table sheet."""
        ws = wb.create_sheet("IV值表")
        
        iv_table = results.get('iv_table', [])
        if not iv_table:
            return
        
        # Title
        ws['A1'] = "变量IV值表"
        ws['A1'].font = self.styles.TITLE_FONT
        ws['A1'].fill = self.styles.HEADER_FILL
        ws.merge_cells('A1:D1')
        ws.row_dimensions[1].height = 25
        
        # Headers
        headers = ['变量', 'IV值', '预测能力']
        row = 3
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = self.styles.HEADER_FONT
            cell.fill = self.styles.HEADER_FILL
            cell.alignment = self.styles.CENTER
            cell.border = self.styles.THIN_BORDER
        
        # Data
        row = 4
        for item in iv_table:
            var_name = item.get('variable', item.get('Variable', ''))
            iv_value = item.get('iv', item.get('IV', 0))
            
            # Determine predictive power
            if iv_value >= 0.5:
                power = "过高（可能过拟合）"
            elif iv_value >= 0.3:
                power = "强"
            elif iv_value >= 0.1:
                power = "中等"
            elif iv_value >= 0.02:
                power = "弱"
            else:
                power = "极弱"
            
            ws.cell(row=row, column=1, value=var_name)
            ws.cell(row=row, column=2, value=round(iv_value, 4) if isinstance(iv_value, (int, float)) else iv_value)
            ws.cell(row=row, column=3, value=power)
            
            for col in range(1, 4):
                ws.cell(row=row, column=col).border = self.styles.THIN_BORDER
                ws.cell(row=row, column=col).font = self.styles.DATA_FONT
            row += 1
        
        # Column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 18
    
    def _add_coefficients_sheet(self, wb: "Workbook", results: dict[str, Any]) -> None:
        """Add model coefficients sheet."""
        ws = wb.create_sheet("模型系数")
        
        coefficients = results.get('coefficients', [])
        if not coefficients:
            return
        
        # Title
        ws['A1'] = "模型系数详情"
        ws['A1'].font = self.styles.TITLE_FONT
        ws['A1'].fill = self.styles.HEADER_FILL
        ws.merge_cells('A1:H1')
        ws.row_dimensions[1].height = 25
        
        # Headers
        headers = ['变量', '系数', '标准误', 'z值', 'P值', '95%CI下限', '95%CI上限', '显著性']
        row = 3
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = self.styles.HEADER_FONT
            cell.fill = self.styles.HEADER_FILL
            cell.alignment = self.styles.CENTER
            cell.border = self.styles.THIN_BORDER
        
        # Data
        row = 4
        for item in coefficients:
            feature = item.get('feature', item.get('variable', ''))
            coef = item.get('coef', item.get('coefficient', 0))
            std_err = item.get('std_err', item.get('std_error', None))
            z_val = item.get('z', item.get('z_value', None))
            p_val = item.get('p_value', item.get('pvalue', None))
            ci_lower = item.get('ci_lower', item.get('conf_int_lower', None))
            ci_upper = item.get('ci_upper', item.get('conf_int_upper', None))
            sig = item.get('significance', '')
            
            ws.cell(row=row, column=1, value=feature)
            ws.cell(row=row, column=2, value=round(coef, 4) if isinstance(coef, (int, float)) else coef)
            ws.cell(row=row, column=3, value=round(std_err, 4) if std_err and isinstance(std_err, (int, float)) else std_err or 'N/A')
            ws.cell(row=row, column=4, value=round(z_val, 4) if z_val and isinstance(z_val, (int, float)) else z_val or 'N/A')
            ws.cell(row=row, column=5, value=self._format_pvalue(p_val) if p_val else 'N/A')
            ws.cell(row=row, column=6, value=round(ci_lower, 4) if ci_lower and isinstance(ci_lower, (int, float)) else ci_lower or 'N/A')
            ws.cell(row=row, column=7, value=round(ci_upper, 4) if ci_upper and isinstance(ci_upper, (int, float)) else ci_upper or 'N/A')
            ws.cell(row=row, column=8, value=sig)
            
            for col in range(1, 9):
                ws.cell(row=row, column=col).border = self.styles.THIN_BORDER
                ws.cell(row=row, column=col).font = self.styles.DATA_FONT
            
            # Highlight significant coefficients
            if p_val is not None and isinstance(p_val, (int, float)) and p_val < 0.05:
                ws.cell(row=row, column=5).fill = self.styles.ACCENT_FILL
            
            row += 1
        
        # Significance legend
        row += 2
        ws.cell(row=row, column=1, value="显著性标记: *** p<0.001, ** p<0.01, * p<0.05, . p<0.1")
        ws.cell(row=row, column=1).font = Font(name='Arial', size=9, italic=True)
        
        # Column widths
        ws.column_dimensions['A'].width = 20
        for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H']:
            ws.column_dimensions[col].width = 12
    
    def _add_rule_overview_sheet(
        self,
        wb: "Workbook",
        results: dict[str, Any],
        title: str
    ) -> None:
        """Add rule mining overview sheet."""
        ws = wb.create_sheet("概览")
        
        ws['A1'] = title
        ws['A1'].font = self.styles.TITLE_FONT
        ws['A1'].fill = self.styles.HEADER_FILL
        ws.merge_cells('A1:F1')
        ws.row_dimensions[1].height = 25
        
        # Summary metrics
        metrics = [
            ("样本量", results.get('n_samples', 'N/A')),
            ("坏样本数", results.get('n_bad', 'N/A')),
            ("基础坏账率", self._format_percent(results.get('base_bad_rate', 0))),
            ("生成规则数", results.get('n_rules_generated', 'N/A')),
            ("最终规则数", results.get('n_rules_selected', 'N/A')),
            ("累计召回率", self._format_percent(results.get('total_recall', 0))),
        ]
        
        row = 3
        for i, (name, value) in enumerate(metrics):
            col = (i % 3) * 2 + 1
            if i > 0 and i % 3 == 0:
                row += 2
            ws.cell(row=row, column=col, value=name).font = self.styles.DATA_FONT
            ws.cell(row=row, column=col+1, value=value).font = self.styles.DATA_FONT
        
        for col in ['A', 'B', 'C', 'D', 'E', 'F']:
            ws.column_dimensions[col].width = 15
    
    def _add_rules_sheet(self, wb: "Workbook", results: dict[str, Any]) -> None:
        """Add rules detail sheet."""
        ws = wb.create_sheet("规则详情")
        
        rules = results.get('selected_rules', results.get('rules', []))
        
        headers = ['序号', '规则', '召回率', '命中率', '坏账率', 'Lift', '累计召回']
        row = 1
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = self.styles.HEADER_FONT
            cell.fill = self.styles.HEADER_FILL
            cell.alignment = self.styles.CENTER
            cell.border = self.styles.THIN_BORDER
        
        row = 2
        for i, rule in enumerate(rules, 1):
            if isinstance(rule, dict):
                ws.cell(row=row, column=1, value=i)
                ws.cell(row=row, column=2, value=rule.get('rule', ''))
                ws.cell(row=row, column=3, value=self._format_percent(rule.get('recall', 0)))
                ws.cell(row=row, column=4, value=self._format_percent(rule.get('hit_rate', 0)))
                ws.cell(row=row, column=5, value=self._format_percent(rule.get('bad_rate', 0)))
                ws.cell(row=row, column=6, value=round(rule.get('lift', 0), 2))
                ws.cell(row=row, column=7, value=self._format_percent(rule.get('cum_recall', rule.get('cumulative_recall', rule.get('dev_cum_recall', 0)))))
                
                for col in range(1, 8):
                    ws.cell(row=row, column=col).border = self.styles.THIN_BORDER
                    ws.cell(row=row, column=col).font = self.styles.DATA_FONT
                row += 1
        
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 50
        for col in ['C', 'D', 'E', 'F', 'G']:
            ws.column_dimensions[col].width = 12
    
    def _add_rule_evaluation_sheet(self, wb: "Workbook", results: dict[str, Any]) -> None:
        """Add rule evaluation sheet."""
        ws = wb.create_sheet("规则评估")
        
        # Similar to rules sheet but with more evaluation metrics
        self._add_rules_sheet(wb, results)
    
    def _add_amount_analysis_sheet(self, wb: "Workbook", results: dict[str, Any]) -> None:
        """Add amount analysis sheet."""
        ws = wb.create_sheet("金额分析")
        
        amount = results.get('amount_analysis', {})
        
        ws['A1'] = "金额维度分析"
        ws['A1'].font = self.styles.TITLE_FONT
        ws['A1'].fill = self.styles.HEADER_FILL
        ws.merge_cells('A1:G1')
        ws.row_dimensions[1].height = 25
        
        # Summary
        ws['A3'] = "汇总指标"
        ws['A3'].font = self.styles.SUBHEADER_FONT
        ws['A3'].fill = self.styles.SUBHEADER_FILL
        ws.merge_cells('A3:D3')
        
        summary_metrics = [
            ("总金额", self._format_currency(amount.get('total_amount', 0))),
            ("总坏账金额", self._format_currency(amount.get('total_bad_amount', 0))),
            ("累计命中金额", self._format_currency(amount.get('cumulative', {}).get('cum_hit_amount', 0))),
            ("金额召回率", self._format_percent(amount.get('cumulative', {}).get('amount_recall', 0))),
        ]
        
        row = 5
        for name, value in summary_metrics:
            ws.cell(row=row, column=1, value=name).font = self.styles.DATA_FONT
            ws.cell(row=row, column=2, value=value).font = self.styles.DATA_FONT
            row += 1
        
        # Rules amount detail
        row += 2
        ws.cell(row=row, column=1, value="规则金额明细").font = self.styles.SUBHEADER_FONT
        ws.cell(row=row, column=1).fill = self.styles.SUBHEADER_FILL
        ws.merge_cells(f'A{row}:G{row}')
        row += 1
        
        headers = ['规则', '命中金额', '金额占比', '坏账金额', '坏账占比', '金额坏账率', '金额Lift']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = self.styles.HEADER_FONT
            cell.fill = self.styles.HEADER_FILL
            cell.border = self.styles.THIN_BORDER
        row += 1
        
        rules_amount = amount.get('rules_amount', [])
        for item in rules_amount:
            ws.cell(row=row, column=1, value=item.get('rule', ''))
            ws.cell(row=row, column=2, value=self._format_currency(item.get('hit_amount', 0)))
            ws.cell(row=row, column=3, value=self._format_percent(item.get('hit_amount_pct', 0)))
            ws.cell(row=row, column=4, value=self._format_currency(item.get('bad_amount', 0)))
            ws.cell(row=row, column=5, value=self._format_percent(item.get('bad_amount_pct', 0)))
            ws.cell(row=row, column=6, value=self._format_percent(item.get('amount_bad_rate', 0)))
            ws.cell(row=row, column=7, value=round(item.get('amount_lift', 0), 2))
            
            for col in range(1, 8):
                ws.cell(row=row, column=col).border = self.styles.THIN_BORDER
                ws.cell(row=row, column=col).font = self.styles.DATA_FONT
            row += 1
        
        ws.column_dimensions['A'].width = 40
        for col in ['B', 'C', 'D', 'E', 'F', 'G']:
            ws.column_dimensions[col].width = 14
    
    def _add_charts_metrics_sheet(
        self,
        wb: "Workbook",
        results: dict[str, Any],
        title: str
    ) -> None:
        """Add charts/metrics sheet (Tab 1: 评估图表)."""
        ws = wb.create_sheet("评估图表")
        
        # Title
        ws.merge_cells('A1:F1')
        ws['A1'] = title
        ws['A1'].font = self.styles.TITLE_FONT
        ws['A1'].fill = self.styles.HEADER_FILL
        ws['A1'].alignment = self.styles.CENTER
        ws.row_dimensions[1].height = 30
        
        # Summary metrics
        row = 3
        ws['A3'] = "模型评估指标"
        ws['A3'].font = self.styles.SUBHEADER_FONT
        ws['A3'].fill = self.styles.SUBHEADER_FILL
        ws.merge_cells('A3:F3')
        
        # Extract metrics
        metrics = [
            ("样本量", results.get('n_samples', 'N/A')),
            ("特征数", results.get('n_features', 'N/A')),
            ("训练集AUC", self._format_metric(results.get('train_auc'))),
            ("测试集AUC", self._format_metric(results.get('test_auc'))),
            ("训练集KS", self._format_metric(results.get('train_ks'))),
            ("测试集KS", self._format_metric(results.get('test_ks'))),
        ]
        
        row = 5
        for i, (name, value) in enumerate(metrics):
            col = (i % 3) * 2 + 1
            if i > 0 and i % 3 == 0:
                row += 2
            ws.cell(row=row, column=col, value=name).font = self.styles.DATA_FONT
            ws.cell(row=row, column=col+1, value=value).font = self.styles.DATA_FONT
            ws.cell(row=row, column=col+1).alignment = self.styles.RIGHT
        
        # Model statistics if available
        if results.get('model_statistics'):
            stats = results['model_statistics']
            row += 4
            ws.cell(row=row, column=1, value="模型统计").font = self.styles.SUBHEADER_FONT
            ws.cell(row=row, column=1).fill = self.styles.SUBHEADER_FILL
            ws.merge_cells(f'A{row}:F{row}')
            
            stat_metrics = [
                ("伪R²", self._format_metric(stats.get('pseudo_r2'))),
                ("对数似然", self._format_metric(stats.get('log_likelihood'))),
                ("AIC", self._format_metric(stats.get('aic'))),
                ("BIC", self._format_metric(stats.get('bic'))),
            ]
            
            row += 2
            for i, (name, value) in enumerate(stat_metrics):
                col = (i % 2) * 3 + 1
                if i > 0 and i % 2 == 0:
                    row += 1
                ws.cell(row=row, column=col, value=name).font = self.styles.DATA_FONT
                ws.cell(row=row, column=col+1, value=value).font = self.styles.DATA_FONT
        
        # Adjust column widths
        for col in ['A', 'B', 'C', 'D', 'E', 'F']:
            ws.column_dimensions[col].width = 15
    
    def _add_optimal_rules_sheet(self, wb: "Workbook", results: dict[str, Any]) -> None:
        """Add optimal rules sheet (Tab 2: 最优规则)."""
        ws = wb.create_sheet("最优规则")
        
        rules = results.get('selected_rules', results.get('rules', []))
        self._write_rules_to_sheet(ws, rules, "最优规则列表")
    
    def _add_filtered_rules_sheet(self, wb: "Workbook", results: dict[str, Any]) -> None:
        """Add filtered rules sheet (Tab 3: 过滤后)."""
        ws = wb.create_sheet("过滤后规则")
        
        rules = results.get('filtered_rules', [])
        self._write_rules_to_sheet(ws, rules, "过滤后规则列表")
    
    def _add_all_rules_sheet(self, wb: "Workbook", results: dict[str, Any]) -> None:
        """Add all rules sheet (Tab 4: 全部)."""
        ws = wb.create_sheet("全部规则")
        
        rules = results.get('all_rules', results.get('rules', []))
        self._write_rules_to_sheet(ws, rules, "全部规则列表")
    
    def _write_rules_to_sheet(self, ws, rules: list, title: str) -> None:
        """Helper to write rules to a worksheet."""
        # Title
        ws['A1'] = title
        ws['A1'].font = self.styles.TITLE_FONT
        ws['A1'].fill = self.styles.HEADER_FILL
        ws.merge_cells('A1:G1')
        ws.row_dimensions[1].height = 25
        
        headers = ['序号', '规则', '召回率', '命中率', '坏账率', 'Lift', '累计召回']
        row = 3
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = self.styles.HEADER_FONT
            cell.fill = self.styles.HEADER_FILL
            cell.alignment = self.styles.CENTER
            cell.border = self.styles.THIN_BORDER
        
        row = 4
        for i, rule in enumerate(rules, 1):
            if isinstance(rule, dict):
                ws.cell(row=row, column=1, value=i)
                ws.cell(row=row, column=2, value=rule.get('rule', ''))
                ws.cell(row=row, column=3, value=self._format_percent(rule.get('recall', 0)))
                ws.cell(row=row, column=4, value=self._format_percent(rule.get('hit_rate', 0)))
                ws.cell(row=row, column=5, value=self._format_percent(rule.get('bad_rate', 0)))
                ws.cell(row=row, column=6, value=round(rule.get('lift', 0), 2))
                ws.cell(row=row, column=7, value=self._format_percent(rule.get('cum_recall', rule.get('cumulative_recall', rule.get('dev_cum_recall', 0)))))
                
                for col in range(1, 8):
                    ws.cell(row=row, column=col).border = self.styles.THIN_BORDER
                    ws.cell(row=row, column=col).font = self.styles.DATA_FONT
                row += 1
        
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 50
        for col in ['C', 'D', 'E', 'F', 'G']:
            ws.column_dimensions[col].width = 12
    
    def _add_validation_sheet(self, wb: "Workbook", results: dict[str, Any]) -> None:
        """Add validation report sheet (Tab 5: 质量验证)."""
        ws = wb.create_sheet("质量验证")
        
        validation = results.get('validation_report', {})
        
        ws['A1'] = "规则质量验证报告"
        ws['A1'].font = self.styles.TITLE_FONT
        ws['A1'].fill = self.styles.HEADER_FILL
        ws.merge_cells('A1:D1')
        ws.row_dimensions[1].height = 25
        
        row = 3
        if isinstance(validation, dict):
            for category, items in validation.items():
                ws.cell(row=row, column=1, value=str(category)).font = self.styles.SUBHEADER_FONT
                ws.cell(row=row, column=1).fill = self.styles.SUBHEADER_FILL
                ws.merge_cells(f'A{row}:D{row}')
                row += 1
                
                if isinstance(items, dict):
                    for key, value in items.items():
                        ws.cell(row=row, column=1, value=str(key)).font = self.styles.DATA_FONT
                        ws.cell(row=row, column=2, value=str(value)).font = self.styles.DATA_FONT
                        row += 1
                else:
                    ws.cell(row=row, column=1, value=str(items)).font = self.styles.DATA_FONT
                    row += 1
                row += 1
        elif isinstance(validation, list):
            headers = list(validation[0].keys()) if validation else []
            for col, header in enumerate(headers, 1):
                ws.cell(row=row, column=col, value=header).font = self.styles.HEADER_FONT
            row += 1
            for item in validation:
                for col, key in enumerate(headers, 1):
                    ws.cell(row=row, column=col, value=item.get(key, '')).font = self.styles.DATA_FONT
                row += 1
        
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
    
    def _add_psi_sheet(self, wb: "Workbook", results: dict[str, Any]) -> None:
        """Add PSI report sheet (Tab 6: 稳定性)."""
        ws = wb.create_sheet("稳定性PSI")
        
        psi_report = results.get('psi_report', [])
        
        ws['A1'] = "规则稳定性（PSI）报告"
        ws['A1'].font = self.styles.TITLE_FONT
        ws['A1'].fill = self.styles.HEADER_FILL
        ws.merge_cells('A1:E1')
        ws.row_dimensions[1].height = 25
        
        if isinstance(psi_report, list) and len(psi_report) > 0:
            # Get headers from first item
            headers = list(psi_report[0].keys()) if psi_report else ['规则', 'PSI', '稳定性']
            row = 3
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.font = self.styles.HEADER_FONT
                cell.fill = self.styles.HEADER_FILL
                cell.border = self.styles.THIN_BORDER
            
            row = 4
            for item in psi_report:
                for col, key in enumerate(headers, 1):
                    value = item.get(key, '')
                    ws.cell(row=row, column=col, value=value).font = self.styles.DATA_FONT
                    ws.cell(row=row, column=col).border = self.styles.THIN_BORDER
                row += 1
        
        ws.column_dimensions['A'].width = 40
        for col in ['B', 'C', 'D', 'E']:
            ws.column_dimensions[col].width = 15
    
    def _add_rule_charts_sheet(self, wb: "Workbook", results: dict[str, Any]) -> None:
        """Add rule mining charts sheet."""
        ws = wb.create_sheet("图表数据")
        
        ws['A1'] = "图表数据"
        ws['A1'].font = self.styles.TITLE_FONT
        ws['A1'].fill = self.styles.HEADER_FILL
        ws.merge_cells('A1:E1')
        
        # Cumulative recall data
        ws['A3'] = "累计召回曲线数据"
        ws['A3'].font = self.styles.SUBHEADER_FONT
        ws['A3'].fill = self.styles.SUBHEADER_FILL
        ws.merge_cells('A3:C3')
        
        headers = ['规则序号', '累计召回率', '累计命中率']
        for col, header in enumerate(headers, 1):
            ws.cell(row=4, column=col, value=header)
        
        rules = results.get('selected_rules', results.get('rules', []))
        for i, rule in enumerate(rules, 1):
            if isinstance(rule, dict):
                ws.cell(row=4+i, column=1, value=i)
                ws.cell(row=4+i, column=2, value=rule.get('cum_recall', 0))
                ws.cell(row=4+i, column=3, value=rule.get('cum_hit_rate', 0))
    
    # =========================================================================
    # Stages Sheet Generation (各阶段下载结果)
    # =========================================================================
    
    def _add_stages_sheets(
        self,
        wb: "Workbook",
        stages: dict[str, Any],
        stage_configs: list["StageDataSheetConfig"] | None = None
    ) -> None:
        """
        为各执行阶段生成独立的Sheet。
        
        使用配置驱动的方式，从stage_configs中读取Sheet名称和下载字段配置。
        如果未提供配置，则使用默认的阶段名称。
        
        Args:
            wb: Excel工作簿
            stages: 阶段数据字典，格式为 {stage_id: {stage_name, status, output_preview, ...}}
            stage_configs: 阶段Sheet配置列表（可选），包含sheet_name、download_field等
        """
        if not stages:
            return
        
        # 构建配置索引：stage_id -> config
        config_map: dict[str, "StageDataSheetConfig"] = {}
        if stage_configs:
            for config in stage_configs:
                config_map[config.stage_id] = config
        
        # 阶段编号映射（用于生成带编号的Sheet名称）
        STAGE_ORDER = {
            # 规则挖掘阶段
            'preprocessing': 1,
            'feature_engineering': 2,
            'generating_rules': 3,
            'rule_filtering': 4,
            'selecting_rules': 5,
            'report_generation': 6,
            # 评分卡阶段
            'data_loading': 1,
            'woe_binning': 2,
            'feature_selection': 3,
            'model_training': 4,
            'score_scaling': 5,
            'model_evaluation': 6,
        }
        
        for stage_id, stage_data in stages.items():
            # 跳过报告生成阶段（该阶段数据已在"任务报告"Sheet中展示）
            if stage_id == 'report_generation':
                continue
            
            # 使用标准化的阶段名称（从翻译映射中获取）
            stage_name = self._get_stage_name(stage_id, stage_data.get("stage_name", stage_id))
            output_preview = stage_data.get("output_preview")
            
            # 只处理有output_preview数据的阶段
            if not output_preview:
                continue
            
            # 从配置获取Sheet名称，或使用默认格式
            config = config_map.get(stage_id)
            if config:
                sheet_name = config.sheet_name[:31]  # Excel最长31字符
            else:
                # 使用编号前缀（如 1_数据预处理, 6_报告生成）
                stage_num = STAGE_ORDER.get(stage_id, 0)
                if stage_num > 0:
                    sheet_name = f"{stage_num}_{stage_name}"[:31]
                else:
                    sheet_name = f"阶段_{stage_name}"[:31]
            
            # 确保Sheet名称唯一
            if sheet_name in wb.sheetnames:
                sheet_name = f"{sheet_name[:27]}_{stage_id[:3]}"
            
            try:
                self._add_single_stage_sheet(
                    wb, sheet_name, stage_id, stage_name, output_preview, config
                )
            except Exception as e:
                # 阶段Sheet生成失败不影响整体报告
                import logging
                logging.getLogger(__name__).warning(f"Failed to add stage sheet {stage_name}: {e}")
    
    def _add_single_stage_sheet(
        self,
        wb: "Workbook",
        sheet_name: str,
        stage_id: str,
        stage_name: str,
        output_preview: dict[str, Any],
        stage_config: "StageDataSheetConfig | None" = None
    ) -> None:
        """
        生成单个阶段的Sheet。
        
        使用配置驱动的方式识别下载数据字段：
        - 如果提供了stage_config，从config.download_field读取字段名
        - 如果未提供配置，使用通用的output_preview展示
        
        Args:
            wb: Excel工作簿
            sheet_name: Sheet名称
            stage_id: 阶段ID
            stage_name: 阶段名称
            output_preview: 阶段输出预览数据
            stage_config: 阶段配置（可选），包含download_field和download_title
        """
        ws = wb.create_sheet(sheet_name)
        
        # 标题行
        ws.merge_cells('A1:J1')
        ws['A1'] = f"阶段: {stage_name}"
        ws['A1'].font = self.styles.TITLE_FONT
        ws['A1'].fill = self.styles.HEADER_FILL
        ws['A1'].alignment = self.styles.CENTER
        ws.row_dimensions[1].height = 30
        
        # 阶段ID
        ws['A2'] = "阶段ID"
        ws['A2'].font = self.styles.SUBHEADER_FONT
        ws['B2'] = stage_id
        ws['B2'].font = self.styles.DATA_FONT
        
        row = 4
        
        if isinstance(output_preview, dict):
            # === 特殊处理：模型评估阶段的多维评分分布数据 ===
            # 评分卡任务的模型评估阶段有6种CSV下载数据：
            # 3个数据集(train/test/oot) × 2种分析视图(ranking_analysis/distribution_view)
            score_distribution = output_preview.get('score_distribution')
            if stage_id == 'model_evaluation' and score_distribution and isinstance(score_distribution, dict):
                row = self._write_model_evaluation_score_distributions(ws, score_distribution, row)
                row += 2  # 空行分隔
                
                # 然后写入阶段概要（排除大型数据）
                exclude_fields = {
                    '_full_stage_data', '_skip_expert_pause', '_skipped_during_retry', 
                    'retry_message', 'score_distribution'  # 已单独处理
                }
                summary_data = {
                    k: v for k, v in output_preview.items()
                    if k not in exclude_fields 
                    and not k.startswith('_')
                    and (not isinstance(v, (list, dict)) or (isinstance(v, dict) and len(str(v)) < 500))
                }
                if summary_data:
                    ws.cell(row=row, column=1, value="阶段概要").font = self.styles.SUBHEADER_FONT
                    ws.cell(row=row, column=1).fill = self.styles.SUBHEADER_FILL
                    row += 1
                    row = self._write_dict_to_sheet(ws, summary_data, row)
            # === 特殊处理：特征筛选阶段的两种CSV下载数据 ===
            # 评分卡任务的特征筛选阶段有两种CSV下载：
            # 1. all_features_detail - 特征筛选明细表
            # 2. all_features_detail[].bin_detail - 分箱明细表（嵌套在each feature中）
            elif stage_id == 'feature_selection':
                all_features_detail = output_preview.get('all_features_detail', [])
                
                if all_features_detail and isinstance(all_features_detail, list) and len(all_features_detail) > 0:
                    # 1. 写入特征筛选明细表
                    row = self._write_stage_download_table(ws, "特征筛选明细", all_features_detail, row)
                    row += 2  # 空行分隔
                    
                    # 2. 写入分箱明细表（展开嵌套的bin_detail）
                    row = self._write_feature_binning_detail_table(ws, all_features_detail, row)
                    row += 2  # 空行分隔
                
                # 写入阶段概要（排除大型数据）
                exclude_fields = {
                    '_full_stage_data', '_skip_expert_pause', '_skipped_during_retry', 
                    'retry_message', 'all_features_detail'  # 已单独处理
                }
                summary_data = {
                    k: v for k, v in output_preview.items()
                    if k not in exclude_fields 
                    and not k.startswith('_')
                    and (not isinstance(v, (list, dict)) or (isinstance(v, dict) and len(str(v)) < 500))
                }
                if summary_data:
                    ws.cell(row=row, column=1, value="阶段概要").font = self.styles.SUBHEADER_FONT
                    ws.cell(row=row, column=1).fill = self.styles.SUBHEADER_FILL
                    row += 1
                    row = self._write_dict_to_sheet(ws, summary_data, row)
            else:
                # === 标准处理：从配置读取下载字段信息（配置驱动，非硬编码）===
                download_data = None
                download_title = None
                download_field = None
                
                if stage_config and stage_config.download_field:
                    download_field = stage_config.download_field
                    download_title = stage_config.download_title or f"{stage_name}数据"
                    download_data = output_preview.get(download_field)
                
                if download_data and isinstance(download_data, list) and len(download_data) > 0:
                    # 优先导出下载数据字段（作为规范的表格）
                    row = self._write_stage_download_table(ws, download_title, download_data, row)
                    row += 2  # 空行分隔
                    
                    # 然后导出阶段概要信息（排除大型数据字段和内部字段）
                    # 需要排除的字段：下载数据字段 + 内部字段 + 预览字段
                    exclude_fields = {
                        '_full_stage_data', '_skip_expert_pause', 
                        '_skipped_during_retry', 'retry_message',
                        'rules_preview', 'top_rules'
                    }
                    if download_field:
                        exclude_fields.add(download_field)
                    
                    summary_data = {
                        k: v for k, v in output_preview.items()
                        if k not in exclude_fields 
                        and not k.startswith('_')
                        and (not isinstance(v, (list, dict)) or (isinstance(v, dict) and len(str(v)) < 500))
                    }
                    if summary_data:
                        ws.cell(row=row, column=1, value="阶段概要").font = self.styles.SUBHEADER_FONT
                        ws.cell(row=row, column=1).fill = self.styles.SUBHEADER_FILL
                        row += 1
                        row = self._write_dict_to_sheet(ws, summary_data, row)
                else:
                    # 没有特定下载数据字段，使用通用处理
                    # 排除内部字段
                    filtered_preview = {
                        k: v for k, v in output_preview.items()
                        if not k.startswith('_') and k not in ['retry_message']
                    }
                    row = self._write_dict_to_sheet(ws, filtered_preview, row)
        elif isinstance(output_preview, list):
            row = self._write_list_to_sheet(ws, output_preview, row)
        else:
            ws.cell(row=row, column=1, value="数据").font = self.styles.SUBHEADER_FONT
            ws.cell(row=row, column=2, value=str(output_preview)).font = self.styles.DATA_FONT
        
        # 调整列宽
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 50
        for col in ['C', 'D', 'E', 'F', 'G', 'H', 'I', 'J']:
            ws.column_dimensions[col].width = 18
    
    def _write_stage_download_table(
        self,
        ws,
        title: str,
        data: list[dict[str, Any]],
        start_row: int
    ) -> int:
        """
        将阶段下载数据以规范表格形式写入Sheet。
        
        Args:
            ws: Worksheet
            title: 表格标题
            data: 下载数据列表（字典列表）
            start_row: 起始行
            
        Returns:
            下一个可用行号
        """
        row = start_row
        
        # 写入标题
        ws.cell(row=row, column=1, value=title).font = self.styles.SUBHEADER_FONT
        ws.cell(row=row, column=1).fill = self.styles.SUBHEADER_FILL
        ws.cell(row=row, column=2, value=f"共 {len(data)} 条").font = self.styles.DATA_FONT
        row += 1
        
        if not data or not isinstance(data[0], dict):
            return row
        
        # 获取表头（优先排序关键字段）
        all_headers = list(data[0].keys())
        # 定义优先字段顺序
        priority_fields = ['feature', 'rule', 'condition', 'hit_rate', 'bad_rate', 
                          'lift', 'recall', 'is_valid', 'is_optimal', 'status',
                          'cumulative_hit_rate', 'cumulative_recall', 'cumulative_lift']
        # 重新排序表头
        sorted_headers = []
        for pf in priority_fields:
            if pf in all_headers:
                sorted_headers.append(pf)
        for h in all_headers:
            if h not in sorted_headers and not h.startswith('_'):
                sorted_headers.append(h)
        headers = sorted_headers[:15]  # 最多15列
        
        # 写入表头
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col_idx, value=self._translate_header(header))
            cell.font = self.styles.HEADER_FONT
            cell.fill = self.styles.HEADER_FILL
            cell.border = self.styles.THIN_BORDER
            cell.alignment = self.styles.CENTER
        row += 1
        
        # 写入数据（无行数限制，全量导出）
        for item in data:
            for col_idx, header in enumerate(headers, 1):
                value = item.get(header)
                cell = ws.cell(row=row, column=col_idx, value=self._format_value(value))
                cell.font = self.styles.DATA_FONT
                cell.border = self.styles.THIN_BORDER
                # 布尔值特殊格式化
                if isinstance(value, bool):
                    cell.value = "是" if value else "否"
            row += 1
        
        return row
    
    def _write_feature_binning_detail_table(
        self,
        ws,
        all_features_detail: list[dict[str, Any]],
        start_row: int
    ) -> int:
        """
        写入特征筛选阶段的分箱明细表。
        
        分箱明细表是从 all_features_detail[].bin_detail 展开生成的，
        与前端 "分箱明细" CSV 下载保持一致。
        
        Args:
            ws: Worksheet
            all_features_detail: 特征筛选明细列表，每个特征包含 bin_detail 字段
            start_row: 起始行
            
        Returns:
            下一个可用行号
        """
        row = start_row
        
        # 收集所有分箱明细数据
        binning_rows = []
        for feature_data in all_features_detail:
            feature_name = feature_data.get('feature', '')
            iv_value = feature_data.get('iv')
            status = feature_data.get('status', '')
            bin_detail = feature_data.get('bin_detail', [])
            
            if bin_detail and isinstance(bin_detail, list):
                for bin_data in bin_detail:
                    binning_rows.append({
                        'feature': feature_name,
                        'bin': bin_data.get('bin', ''),
                        'woe': bin_data.get('woe'),
                        'count': bin_data.get('count'),
                        'bad_count': bin_data.get('bad_count'),
                        'bad_rate': bin_data.get('bad_rate'),
                        'iv': iv_value,
                        'status': status,
                    })
        
        if not binning_rows:
            return row
        
        # 写入标题
        ws.cell(row=row, column=1, value="分箱明细").font = self.styles.SUBHEADER_FONT
        ws.cell(row=row, column=1).fill = self.styles.SUBHEADER_FILL
        ws.cell(row=row, column=2, value=f"共 {len(binning_rows)} 条").font = self.styles.DATA_FONT
        row += 1
        
        # 表头（与前端CSV下载保持一致）
        headers = ['特征名称', '分箱', 'WOE', '样本数', '坏样本数', '坏账率', 'IV值', '筛选状态']
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col_idx, value=header)
            cell.font = self.styles.HEADER_FONT
            cell.fill = self.styles.HEADER_FILL
            cell.border = self.styles.THIN_BORDER
            cell.alignment = self.styles.CENTER
        row += 1
        
        # 写入数据行
        for bin_row in binning_rows:
            row_data = [
                bin_row['feature'],
                bin_row['bin'],
                f"{bin_row['woe']:.4f}" if bin_row['woe'] is not None else '-',
                bin_row['count'] if bin_row['count'] is not None else '-',
                bin_row['bad_count'] if bin_row['bad_count'] is not None else '-',
                f"{bin_row['bad_rate'] * 100:.2f}%" if bin_row['bad_rate'] is not None else '-',
                f"{bin_row['iv']:.4f}" if bin_row['iv'] is not None else '-',
                bin_row['status'],
            ]
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row, column=col_idx, value=value)
                cell.font = self.styles.DATA_FONT
                cell.border = self.styles.THIN_BORDER
            row += 1
        
        return row
    
    def _write_model_evaluation_score_distributions(
        self,
        ws,
        score_distribution: dict[str, Any],
        start_row: int
    ) -> int:
        """
        写入模型评估阶段的多维评分分布数据。
        
        评分卡任务的模型评估阶段有特殊的数据结构：
        - 3个数据集：train（训练集）、test（测试集）、oot（OOT验证集，可选）
        - 2种分析视图：ranking_analysis（排序性分析/等频分箱）、distribution_view（评分分布/等宽分箱）
        - 共计最多6种CSV下载数据
        
        Args:
            ws: Worksheet
            score_distribution: 评分分布数据字典，结构为 {dataset: {ranking_analysis: {...}, distribution_view: {...}}}
            start_row: 起始行
            
        Returns:
            下一个可用行号
        """
        row = start_row
        
        # 数据集配置
        dataset_configs = [
            ('train', '训练集'),
            ('test', '测试集'),
            ('oot', 'OOT验证集'),
        ]
        
        # 分析视图配置
        view_configs = [
            ('ranking_analysis', '排序性分析（等频分箱）'),
            ('distribution_view', '评分分布（等宽分箱）'),
        ]
        
        # 表头定义（与前端CSV下载保持一致）
        headers = ['序号', '分数区间', '样本数', '样本占比', '好样本', '坏样本', '坏样本率', 'Lift', '累计坏样本率']
        
        for dataset_key, dataset_label in dataset_configs:
            dataset_data = score_distribution.get(dataset_key)
            if not dataset_data:
                continue
            
            for view_key, view_label in view_configs:
                view_data = dataset_data.get(view_key)
                if not view_data:
                    continue
                
                bins = view_data.get('bins', [])
                if not bins:
                    continue
                
                # 写入标题
                title = f"{dataset_label} - {view_label}"
                n_bins = view_data.get('n_bins', len(bins))
                ws.cell(row=row, column=1, value=title).font = self.styles.SUBHEADER_FONT
                ws.cell(row=row, column=1).fill = self.styles.SUBHEADER_FILL
                ws.cell(row=row, column=2, value=f"共 {n_bins} 组").font = self.styles.DATA_FONT
                row += 1
                
                # 写入汇总信息（从 summary 获取，如果没有则从 dataset_data 的 summary 获取）
                summary = view_data.get('summary') or dataset_data.get('summary', {})
                if summary:
                    good_mean = summary.get('good_mean')
                    bad_mean = summary.get('bad_mean')
                    separation = abs(good_mean - bad_mean) if good_mean is not None and bad_mean is not None else None
                    
                    # 分离度评级
                    if separation is not None:
                        if separation >= 60:
                            sep_rating = "★优秀"
                        elif separation >= 40:
                            sep_rating = "✓良好"
                        elif separation >= 20:
                            sep_rating = "○合格"
                        else:
                            sep_rating = "△偏低"
                        sep_str = f"{separation:.1f} ({sep_rating})"
                    else:
                        sep_str = None
                    
                    summary_items = [
                        ('总样本', summary.get('total_samples')),
                        ('坏样本率', f"{summary.get('overall_bad_rate', 0):.2f}%"),
                        ('好样本均分', f"{good_mean:.1f}" if good_mean is not None else None),
                        ('坏样本均分', f"{bad_mean:.1f}" if bad_mean is not None else None),
                        ('分离度', sep_str),
                    ]
                    for idx, (label, value) in enumerate(summary_items):
                        if value is not None:
                            ws.cell(row=row, column=1 + idx * 2, value=label).font = self.styles.DATA_FONT
                            ws.cell(row=row, column=2 + idx * 2, value=str(value)).font = self.styles.DATA_FONT
                    row += 1
                
                # 写入表头
                for col_idx, header in enumerate(headers, 1):
                    cell = ws.cell(row=row, column=col_idx, value=header)
                    cell.font = self.styles.HEADER_FONT
                    cell.fill = self.styles.HEADER_FILL
                    cell.border = self.styles.THIN_BORDER
                    cell.alignment = self.styles.CENTER
                row += 1
                
                # 写入数据行
                for idx, bin_data in enumerate(bins, 1):
                    row_data = [
                        idx,
                        bin_data.get('bin', bin_data.get('bin_label', '-')),
                        bin_data.get('total', '-'),
                        f"{bin_data.get('pct_total', 0):.2f}%",
                        bin_data.get('good', '-'),
                        bin_data.get('bad', '-'),
                        f"{bin_data.get('bad_rate', 0):.2f}%",
                        f"{bin_data.get('lift', 0):.2f}",
                        f"{bin_data.get('cum_bad_rate', 0):.2f}%",
                    ]
                    for col_idx, value in enumerate(row_data, 1):
                        cell = ws.cell(row=row, column=col_idx, value=value)
                        cell.font = self.styles.DATA_FONT
                        cell.border = self.styles.THIN_BORDER
                    row += 1
                
                row += 1  # 表格之间空一行
        
        return row
    
    # 阶段名称映射（类级别常量）
    STAGE_NAME_MAP = {
        # 规则挖掘阶段
        'preprocessing': '数据预处理',
        'feature_engineering': '特征工程',
        'generating_rules': '规则生成',
        'rule_filtering': '规则筛选',
        'selecting_rules': '最优选择',
        'report_generation': '报告生成',
        # 评分卡阶段
        'data_loading': '数据加载',
        'woe_binning': 'WOE分箱',
        'feature_selection': '特征筛选',
        'model_training': '模型训练',
        'score_scaling': '评分转换',
        'model_evaluation': '模型评估',
    }
    
    def _get_stage_name(self, stage_id: str, default_name: str) -> str:
        """获取标准化的阶段名称。"""
        return self.STAGE_NAME_MAP.get(stage_id, default_name)
    
    def _translate_header(self, header: str) -> str:
        """翻译英文表头/key为中文（英文key + 中文标注格式）。"""
        translations = {
            # 基本字段
            'feature': '特征名',
            'rule': '规则',
            'condition': '条件',
            'status': '状态',
            'type': '类型',
            'source': '来源',
            
            # 样本/数据相关
            'rows': '样本数',
            'columns': '列数',
            'feature_count': '特征数量',
            'target_rate': '坏账率',
            'n_samples': '样本量',
            'n_bad': '坏样本量',
            
            # 规则指标（与前端UI保持一致）
            'hit_rate': '命中率',
            'bad_rate': '坏账率',
            'lift': 'Lift提升度',  # 前端：Lift提升度
            'recall': '召回率',
            'cumulative_hit_rate': '累计命中率',  # 前端：累计Hit
            'cumulative_recall': '累计召回率',  # 前端：累计Recall
            'cumulative_lift': '累计提升度',
            'cumulative_bad_rate': '累计坏账率',
            'is_valid': '最终有效',  # 前端：最终有效
            'is_optimal': '入选最优规则集',  # 前端：入选最优规则集
            'direction_valid': '单调性校验通过',  # 前端：单调性校验通过
            'lift_valid': 'Lift阈值通过',  # 前端：Lift阈值通过
            'hit_rate_valid': '命中率阈值通过',  # 前端：命中率阈值通过
            'filter_reason': '筛选过滤原因',  # 前端：筛选过滤原因
            'rejection_reason': '最优选择淘汰原因',  # 前端：最优选择淘汰原因
            'rejection_rank': '坏账率排名',  # 前端：坏账率排名
            
            # IV/特征相关
            'original_iv': '原始IV',
            'post_iv': 'IV值',
            'iv': 'IV值',
            'missing_rate': '缺失率',
            'n_bins': '分箱数',
            
            # 数据预处理阶段
            'auto_exclude_report': '自动排除报告',
            'user_specified': '用户指定',
            'auto_detected': '自动检测',
            'total_excluded': '排除总数',
            'missing_summary': '缺失值汇总',
            'avg_missing_rate': '平均缺失率',
            'max_missing_rate': '最大缺失率',
            'min_missing_rate': '最小缺失率',
            'total_features': '特征总数',
            'features_with_missing': '有缺失的特征数',
            'high_missing_features': '高缺失特征',
            'distribution': '分布',
            'outlier_count': '异常值数',
            'special_value_info': '特殊值信息',
            'special_values': '特殊值',
            'affected_features': '受影响特征',
            'total_replaced': '替换总数',
            'details': '详情',
            
            # 特征工程阶段
            'before_count': '筛选前数量',
            'after_count': '筛选后数量',
            'iv_threshold': 'IV阈值',
            'iv_distribution': 'IV分布',
            'top_features': 'Top特征',
            
            # 规则生成阶段
            'generated_count': '生成规则数',
            'rules_preview': '规则预览',
            'top_rules': 'Top规则',
            'all_rules_for_download': '全部规则',
            
            # 规则筛选阶段（与前端UI保持一致）
            'filter_criteria': '筛选条件',
            'filter_summary': '筛选汇总',
            'min_lift': '最小Lift阈值',  # 前端显示：最小Lift阈值移除
            'max_hit_rate': '最大命中率阈值',  # 前端显示：最大命中率移除
            'direction_removed': '单调性校验移除',  # 前端：单调性校验移除
            'bad_rate_zero_removed': '坏账率为0移除',  # 规则坏账率为0（无风险识别能力）
            'lift_removed': '最小Lift阈值移除',  # 前端：最小Lift阈值移除
            'hit_rate_removed': '最大命中率移除',  # 前端：最大命中率移除
            'total_removed': '总移除数',
            'all_rules_with_status': '规则筛选明细',  # 前端显示为表格
            
            # 最优选择阶段（与前端UI保持一致）
            'allow_overlap': '允许规则重叠',  # 前端配置项
            'selection_mode': '选择模式',
            'target_recall': '目标召回率',
            'target_bad_rate': '目标坏账率',
            'rejected_rules_stats': '淘汰规则统计',  # 前端：淘汰原因
            'total_rejected': '淘汰总数',
            'reason_distribution': '淘汰原因分布',
            'all_optimal_rules': '最优规则列表',
            'risk_targets': '风险目标',
            'optimal_rules': '最优规则数',  # 前端摘要卡片
            'total_rules': '生成规则总数',  # 前端漏斗图
            'filtered_rules': '筛选后规则数',  # 前端漏斗图
            'avg_lift': '平均提升度',  # 前端摘要卡片：累计提升度
            
            # 报告生成阶段
            'chart_data': '图表数据',
            'validation_report': '质量验证报告',
            'psi_report': 'PSI稳定性报告',
            
            # 质量验证相关（与前端ValidationReportPanel一致）
            'quality_score': '综合质量评分',
            'score_breakdown': '评分明细',
            'discrimination': '提升度得分',  # 前端：提升度(30分)
            'independence': '独立性得分',  # 前端：独立性(15分)
            'complexity': '复杂度得分',  # 前端：复杂度(15分)
            'coverage': '命中率得分',  # 前端：命中率(15分)
            'discrimination_report': '提升度评估',
            'recall_report': '召回率评估',
            'coverage_report': '命中率/覆盖率评估',
            'overlap_report': '重叠度评估',
            'redundancy_report': '冗余度评估',
            'complexity_report': '复杂度评估',
            'conflict_report': '冲突检测',
            'avg_lift': '平均Lift',
            'min_lift': '最小Lift',
            'max_lift': '最大Lift',
            'total_coverage': '总覆盖率',
            'avg_overlap': '平均重叠率',
            'redundant_count': '冗余规则对数',
            'avg_complexity': '平均条件数',
            'max_complexity': '最大条件数',
            'warnings': '优化建议',
            
            # PSI相关
            'hit_rate_base': '基准命中率',
            'hit_rate_compare': '对比命中率',
            'psi': 'PSI值',
            'stability': '稳定性',
            
            # ==================== 评分卡开发相关 ====================
            
            # 模型评估指标（对应前端 charts Tab）
            'ks': 'KS值',
            'auc': 'AUC值',
            'gini': 'Gini系数',
            'accuracy': '准确率',
            'train_ks': '训练集KS',
            'test_ks': '测试集KS',
            'oot_ks': 'OOT集KS',
            'train_auc': '训练集AUC',
            'test_auc': '测试集AUC',
            'oot_auc': 'OOT集AUC',
            'train_gini': '训练集Gini',
            'test_gini': '测试集Gini',
            'oot_gini': 'OOT集Gini',
            'n_features': '特征数',
            'overfit_warning': '过拟合警告',
            
            # 特征筛选（对应前端 selection Tab）
            'selected_features': '入模特征',
            'selection_detail': '筛选详情',
            'outlier_info': '异常值检测',
            'initial_features': '初始特征',
            'final_features': '最终特征',
            'removed_by_iv': 'IV筛除',
            'removed_by_corr': '相关性筛除',
            'removed_by_vif': 'VIF筛除',
            'removed_by_stepwise': '逐步回归筛除',
            'stepwise_result': '逐步回归结果',
            'coefficient_validation': '系数方向验证',
            'iteration': '迭代步骤',
            'action': '操作',
            'pvalue': 'P值',
            'direction': '方向',
            'significance_level': '显著性水平',
            'selected': '已选择',
            'final_pvalues': '最终P值',
            'coefficients': '系数',
            'valid_direction': '方向正确',
            'invalid_direction': '方向异常',
            'intercept': '截距项',
            'percentage': '占比',
            'lower_bound': '下界',
            'upper_bound': '上界',
            'method': '方法',
            
            # 评分卡明细（对应前端 scorecard Tab）
            'scorecard': '评分卡',
            'variable': '变量',
            'bin': '分箱',
            'woe': 'WOE值',
            'points': '评分',
            'WOE': 'WOE值',
            'Points': '评分',
            'Bin': '分箱',
            
            # IV值排序（对应前端 iv Tab）
            'iv_table': 'IV值表',
            'IV': 'IV值',
            'total_iv': '总IV值',
            
            # 模型系数（对应前端 coefficients Tab）
            'coefficient': '系数',
            'coef': '系数',
            'std_err': '标准误',
            'std_error': '标准误',
            'z': 'z值',
            'z_value': 'z值',
            'p_value': 'P值',
            'significance': '显著性',
            
            # 统计检验（对应前端 statistics Tab）
            'model_statistics': '模型统计',
            'n_observations': '观测数',
            'n_params': '参数数量',
            'log_likelihood': '对数似然',
            'null_log_likelihood': '零模型对数似然',
            'pseudo_r2': '伪R²',
            'aic': 'AIC',
            'bic': 'BIC',
            'lr_stat': '似然比统计量',
            'lr_pvalue': '似然比P值',
            
            # 评分转换（对应前端 converter Tab）
            'base_score': '基准分',
            'pdo': 'PDO',
            'base_odds': '基准Odds',
            'score_range': '评分区间',
            
            # 评分分布（图表数据）
            'roc': 'ROC曲线',
            'fpr': '假阳性率',
            'tpr': '真阳性率',
            'population_pct': '人群比例',
            'cum_bad': '累计坏样本率',
            'cum_good': '累计好样本率',
            'ks_curve': 'KS曲线',
            'ks_max': 'KS最大值',
            'ks_max_position': 'KS最大位置',
            'score_distribution': '评分分布',
            'bins': '分箱',
            'pct_total': '占比',
            'good': '好样本数',
            'bad': '坏样本数',
            'cum_bad_rate': '累计坏样本率',
            'good_mean': '好样本均值',
            'bad_mean': '坏样本均值',
            'score_min': '最低分',
            'score_max': '最高分',
            'overall_bad_rate': '总体坏账率',
            'total_samples': '总样本数',
            'total_bad': '总坏样本数',
            'total_good': '总好样本数',
            
            # 多数据集指标
            'multi_dataset_metrics': '多数据集指标',
            'train': '训练集',
            'test': '测试集',
            'oot': 'OOT集',
            'samples': '样本数',
            
            # ==================== 规则挖掘阶段相关补充 ====================
            
            # 阶段名称（Stage ID）
            'preprocessing': '数据预处理',
            'feature_engineering': '特征工程',
            'generating_rules': '规则生成',
            'rule_filtering': '规则筛选',
            'selecting_rules': '最优选择',
            'report_generation': '报告生成',
            # 评分卡阶段名称
            'data_loading': '数据加载',
            'woe_binning': 'WOE分箱',
            'feature_selection': '特征筛选',
            'model_training': '模型训练',
            'score_scaling': '评分转换',
            'model_evaluation': '模型评估',
            
            # 数据预处理阶段
            'quality_issues': '数据质量问题',
            'needs_feature_engineering': '是否需要特征工程',
            'derived_features': '衍生特征',
            'onehot_count': 'OneHot编码数',
            'datetime_count': '日期衍生数',
            'text_count': '文本衍生数',
            'total_derived': '衍生特征总数',
            'split_info': '数据拆分信息',
            'train_target_rate': '训练集坏账率',
            'test_target_rate': '测试集坏账率',
            'split_method': '拆分方法',
            'test_ratio': '测试集比例',
            'random': '随机拆分',
            'stratified': '分层拆分',
            'time_based': '时间序列拆分',
            
            # 特征工程阶段
            'after_onehot_count': 'OneHot编码后数量',
            'removed_reasons': '移除原因',
            'added_reasons': '添加原因',
            'strong': '强(IV≥0.3)',
            'medium_strong': '中强(0.1≤IV<0.3)',
            'medium': '中等(0.02≤IV<0.1)',
            'weak': '弱(IV<0.02)',
            'onehot_stats': 'OneHot编码统计',
            'original_cols': '原始列名',
            'original_count': '原始数量',
            'derived_count': '衍生数量',
            'retained_derived': '保留衍生数',
            'removed_derived': '移除衍生数',
            'missing_filter_stats': '缺失值筛选统计',
            'threshold': '阈值',
            'removed_count': '移除数量',
            'removed_vars': '移除变量',
            'has_more': '更多数据',
            'feature_details': '特征明细',
            'feature_name': '特征名称',
            'data_type': '数据类型',
            'iv_level': 'IV等级',
            'no': '序号',
            
            # 数据类型
            'int64': '整数型',
            'float64': '浮点型',
            'object': '字符串',
            'bool': '布尔型',
            'datetime64': '日期型',
            
            # 规则生成阶段
            'mining_mode': '挖掘模式',
            'multi': '多变量',
            'single': '单变量',
            'has_full_tree': '包含完整决策树',
            'use_full_tree': '使用完整决策树',
            'n_vars': '变量数',
            'max_depth': '最大深度',
            'min_samples_leaf': '叶节点最小样本',
            'bin_method': '分箱方法',
            'used_var': '使用变量',
            'chimerge': '卡方分箱',
            'quantile': '等频分箱',
            'equal_width': '等宽分箱',
            'decision_tree': '决策树分箱',
            
            # 规则筛选阶段
            'direction_filtered_count': '单调性筛选数',
            
            # 最优选择阶段
            'min_recall_ruleset': '最小召回规则集',
            'min_bad_rate_ruleset': '最低坏账规则集',
            'target_bad_rate_ruleset': '目标坏账规则集',
            'min_lift_ruleset': '最小提升度规则集',
            'ruleset_summary': '规则集汇总',
            'estimated_overall_bad_rate': '预估整体坏账率',
            'original_bad_rate': '原始坏账率',
            'top_rejected_rules': '被淘汰规则示例',
            
            # 报告生成阶段
            'report_sections': '报告章节',
            'chart_types': '图表类型',
            'cumulative_metrics': '累计指标',
            'rules_table': '规则表格',
            'rule_distribution': '规则分布',
            'summary': '摘要',
            'has_chart_data': '包含图表数据',
            'quality_level': '质量等级',
            'validation_passed': '验证通过',
            'validation_issues': '验证问题',
            'overlap_info': '重叠信息',
            'high_overlap_pairs': '高重叠规则对',
            'psi_calculated': '已计算PSI',
            'psi_summary': 'PSI汇总',
            'total_rules_checked': '检查规则总数',
            'stable_rules': '稳定规则数',
            'unstable_rules': '不稳定规则数',
            'avg_psi': '平均PSI',
            'final_rules_count': '最终规则数',
            'total_rules_evaluated': '评估规则总数',
            'has_tree_structure': '包含树结构',
            'has_rule_source_stats': '包含规则来源统计',
            'mining_method': '挖掘方法',
            'full_tree': '完整决策树',
            
            # 其他通用
            'skipped': '已跳过',
            'reason': '原因',
            'enabled': '已启用',
            'count': '数量',
            'total': '总计',
            'average': '平均',
            'max': '最大',
            'min': '最小',
        }
        
        translated = translations.get(header)
        if translated:
            # 格式：英文key (中文标注)
            return f"{header} ({translated})"
        return header
    
    def _write_dict_to_sheet(
        self,
        ws,
        data: dict[str, Any],
        start_row: int
    ) -> int:
        """
        将字典数据写入Sheet（key 带中文标注）。
        
        Args:
            ws: Worksheet
            data: 字典数据
            start_row: 起始行
            
        Returns:
            下一个可用行号
        """
        row = start_row
        
        for key, value in data.items():
            # 翻译 key（英文key + 中文标注格式）
            translated_key = self._translate_header(key)
            
            if isinstance(value, dict):
                # 子字典作为子表
                ws.cell(row=row, column=1, value=translated_key).font = self.styles.SUBHEADER_FONT
                ws.cell(row=row, column=1).fill = self.styles.SUBHEADER_FILL
                row += 1
                for sub_key, sub_value in value.items():
                    translated_sub_key = self._translate_header(sub_key)
                    ws.cell(row=row, column=1, value=f"  {translated_sub_key}").font = self.styles.DATA_FONT
                    ws.cell(row=row, column=2, value=self._format_value(sub_value)).font = self.styles.DATA_FONT
                    row += 1
            elif isinstance(value, list):
                # 列表数据
                ws.cell(row=row, column=1, value=translated_key).font = self.styles.SUBHEADER_FONT
                ws.cell(row=row, column=1).fill = self.styles.SUBHEADER_FILL
                row += 1
                
                if value and isinstance(value[0], dict):
                    # 列表中的字典 - 创建表格
                    headers = list(value[0].keys())
                    for col_idx, header in enumerate(headers[:6], 1):  # 最多6列
                        cell = ws.cell(row=row, column=col_idx, value=self._translate_header(header))
                        cell.font = self.styles.HEADER_FONT
                        cell.fill = self.styles.HEADER_FILL
                        cell.border = self.styles.THIN_BORDER
                    row += 1
                    
                    for item in value[:100]:  # 最多100行
                        for col_idx, header in enumerate(headers[:6], 1):
                            cell = ws.cell(row=row, column=col_idx, value=self._format_value(item.get(header)))
                            cell.font = self.styles.DATA_FONT
                            cell.border = self.styles.THIN_BORDER
                        row += 1
                    
                    if len(value) > 100:
                        ws.cell(row=row, column=1, value=f"...共{len(value)}条数据，仅展示前100条").font = self.styles.DATA_FONT
                        row += 1
                else:
                    # 简单列表
                    for i, item in enumerate(value[:50], 1):
                        ws.cell(row=row, column=1, value=f"  [{i}]").font = self.styles.DATA_FONT
                        ws.cell(row=row, column=2, value=self._format_value(item)).font = self.styles.DATA_FONT
                        row += 1
                    if len(value) > 50:
                        ws.cell(row=row, column=1, value=f"  ...共{len(value)}项").font = self.styles.DATA_FONT
                        row += 1
            else:
                # 简单键值对（应用翻译）
                ws.cell(row=row, column=1, value=translated_key).font = self.styles.DATA_FONT
                ws.cell(row=row, column=2, value=self._format_value(value)).font = self.styles.DATA_FONT
                row += 1
        
        return row
    
    def _write_list_to_sheet(
        self,
        ws,
        data: list[Any],
        start_row: int
    ) -> int:
        """
        将列表数据写入Sheet。
        
        Args:
            ws: Worksheet
            data: 列表数据
            start_row: 起始行
            
        Returns:
            下一个可用行号
        """
        row = start_row
        
        if data and isinstance(data[0], dict):
            # 字典列表 - 创建表格
            headers = list(data[0].keys())
            for col_idx, header in enumerate(headers[:10], 1):  # 最多10列
                cell = ws.cell(row=row, column=col_idx, value=str(header))
                cell.font = self.styles.HEADER_FONT
                cell.fill = self.styles.HEADER_FILL
                cell.border = self.styles.THIN_BORDER
            row += 1
            
            for item in data[:200]:  # 最多200行
                for col_idx, header in enumerate(headers[:10], 1):
                    cell = ws.cell(row=row, column=col_idx, value=self._format_value(item.get(header)))
                    cell.font = self.styles.DATA_FONT
                    cell.border = self.styles.THIN_BORDER
                row += 1
            
            if len(data) > 200:
                ws.cell(row=row, column=1, value=f"...共{len(data)}条数据，仅展示前200条").font = self.styles.DATA_FONT
                row += 1
        else:
            # 简单列表
            for i, item in enumerate(data[:100], 1):
                ws.cell(row=row, column=1, value=i).font = self.styles.DATA_FONT
                ws.cell(row=row, column=2, value=self._format_value(item)).font = self.styles.DATA_FONT
                row += 1
        
        return row
    
    def _format_value(self, value: Any) -> str:
        """格式化任意值为字符串。"""
        if value is None:
            return ""
        if isinstance(value, float):
            if abs(value) < 0.0001 and value != 0:
                return f"{value:.6f}"
            return f"{value:.4f}"
        if isinstance(value, bool):
            return "是" if value else "否"
        if isinstance(value, (list, dict)):
            import json
            try:
                return json.dumps(value, ensure_ascii=False)[:200]  # 截断长JSON
            except:
                return str(value)[:200]
        return str(value)
    
    # =========================================================================
    # Helper methods
    # =========================================================================
    
    def _format_metric(self, value: Any) -> str:
        """Format metric value."""
        if value is None:
            return "N/A"
        if isinstance(value, float):
            return f"{value:.4f}"
        return str(value)
    
    def _format_percent(self, value: Any) -> str:
        """Format percentage value."""
        if value is None:
            return "N/A"
        if isinstance(value, (int, float)):
            return f"{value * 100:.2f}%"
        return str(value)
    
    def _format_pvalue(self, value: Any) -> str:
        """Format p-value."""
        if value is None:
            return "N/A"
        if isinstance(value, float):
            if value < 0.001:
                return "<0.001"
            return f"{value:.4f}"
        return str(value)
    
    def _format_currency(self, value: Any) -> str:
        """Format currency value."""
        if value is None:
            return "N/A"
        if isinstance(value, (int, float)):
            return f"¥{value:,.2f}"
        return str(value)


def generate_excel_report(
    results: dict[str, Any],
    report_type: Literal['scorecard', 'rule_mining'] = 'scorecard',
    sections: list[str] | None = None,
    output_path: str | Path | None = None,
    **kwargs: Any
) -> bytes | None:
    """
    Convenience function to generate Excel report.
    
    Args:
        results: Results dictionary from scorecard or rule mining
        report_type: Type of report ('scorecard' or 'rule_mining')
        sections: Sections to include (None = all)
        output_path: Optional path to save the report
        **kwargs: Additional arguments for ExcelReportGenerator
        
    Returns:
        Excel file as bytes, or None if saved to file
    """
    generator = ExcelReportGenerator(**kwargs)
    
    if report_type == 'scorecard':
        report_bytes = generator.generate_scorecard_report(results, sections)
    else:
        report_bytes = generator.generate_rule_mining_report(results, sections)
    
    if output_path:
        Path(output_path).write_bytes(report_bytes)
        return None
    
    return report_bytes
